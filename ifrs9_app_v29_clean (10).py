pip install plotly
#!/usr/bin/env python3
"""IFRS 9 — Staging Dynamique Avancé v2 | Forvis Mazars"""
import io, warnings
import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy import stats
from scipy.stats import zscore
from scipy.linalg import lstsq
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from sklearn.metrics import silhouette_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ════════════════════════════════════════════════════════════════
# CONFIG PAGE
# ════════════════════════════════════════════════════════════════
st.set_page_config(page_title="IFRS 9 — PFE Implémentation | Forvis Mazars",
                   page_icon=None, layout="wide",
                   initial_sidebar_state="expanded")

LOGO_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 400 72" width="320" height="58">
  <rect width="400" height="72" rx="8" fill="#003087"/>
  <text x="16" y="24" font-family="Arial,sans-serif" font-size="12" font-weight="bold"
        fill="#00A3E0" letter-spacing="2">FORVIS</text>
  <text x="16" y="48" font-family="Arial,sans-serif" font-size="20" font-weight="bold"
        fill="white" letter-spacing="1">MAZARS</text>
  <rect x="118" y="8" width="2" height="56" fill="#00A3E0" rx="1"/>
  <text x="130" y="22" font-family="Arial,sans-serif" font-size="10" fill="#00A3E0" font-weight="700">IFRS 9 — Portefeuille Obligataire Marocain</text>
  <text x="130" y="37" font-family="Arial,sans-serif" font-size="9" fill="white">Staging Dynamique · PD Vasicek · PD CPV · LGD PIT · EAD · ECL IFRS 9</text>
  <text x="130" y="52" font-family="Arial,sans-serif" font-size="8" fill="#7BBFFF">LGD Frye-Jacobs · Courbe BAM · Analyse &amp; Interpretation</text>
  <text x="130" y="65" font-family="Arial,sans-serif" font-size="7" fill="#99C9FF">Date de calcul : 25/04/2025</text>
</svg>"""

# Titre PFE complet affiché dans la page
PFE_TITLE = """
<div style="background:linear-gradient(135deg,#1B1F6B 0%,#2E3580 60%,#009FE3 100%);
     padding:22px 32px;border-radius:10px;margin-bottom:18px;
     display:flex;justify-content:space-between;align-items:center">
  <div>
    <div style="color:white;font-size:1.35rem;font-weight:800;letter-spacing:0.02em;margin-bottom:4px">
      PROJET DE FIN D&#39;ETUDES &mdash; FORVIS MAZARS MAROC
    </div>
    <div style="color:rgba(255,255,255,0.90);font-size:0.90rem;font-weight:500">
      Implementation de la Norme IFRS 9 &nbsp;&middot;&nbsp; Portefeuille Obligataire Marocain &nbsp;&middot;&nbsp; 328 Obligations
    </div>
    <div style="color:rgba(255,255,255,0.70);font-size:0.76rem;margin-top:5px;letter-spacing:0.03em">
      Staging Dynamique &nbsp;&middot;&nbsp; PD Merton-Vasicek &nbsp;&middot;&nbsp; PD CPV &nbsp;&middot;&nbsp; LGD Vasicek &amp; Frye-Jacobs &nbsp;&middot;&nbsp; EAD &nbsp;&middot;&nbsp; ECL IFRS 9 &nbsp;&middot;&nbsp; Analyse &amp; Interpretation
    </div>
  </div>
  <div style="text-align:right">
    <div style="color:#009FE3;font-size:1.6rem;font-weight:900;letter-spacing:0.04em;line-height:1.0">forv/s</div>
    <div style="color:rgba(255,255,255,0.95);font-size:1.25rem;font-weight:700;letter-spacing:0.06em">mazars</div>
  </div>
</div>"""

st.markdown("""<style>
/* ═══════════════════════════════════════════════════════════
   FORVIS MAZARS — IFRS 9 PROFESSIONAL THEME
   Palette: Navy #1B1F6B · Cyan #009FE3 · White · Light gray
   ═══════════════════════════════════════════════════════════ */

/* ── Variables globales ───────────────────────────────────── */
:root {
  --fm-navy:   #1B1F6B;
  --fm-cyan:   #009FE3;
  --fm-dark:   #10144A;
  --fm-light:  #F0F4FA;
  --fm-border: #D8DFF0;
  --fm-text:   #1A1A2E;
  --fm-green:  #1A6B2E;
  --fm-red:    #A31919;
  --fm-gold:   #8B6914;
  --fm-radius: 8px;
  --fm-shadow: 0 2px 12px rgba(27,31,107,0.10);
}

/* ── Page principale ──────────────────────────────────────── */
.main .block-container {
  padding: 1rem 2.5rem 2rem 2.5rem !important;
  max-width: 1400px;
  background: #FAFBFE;
}

/* ── Sidebar ──────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, var(--fm-navy) 0%, var(--fm-dark) 100%) !important;
  border-right: 1px solid rgba(0,163,224,0.2);
}
[data-testid="stSidebarUserContent"] label {
  color: rgba(255,255,255,0.88) !important;
  font-weight: 600;
  font-size: 0.78rem;
  letter-spacing: 0.04em;
  text-transform: uppercase;
}
[data-testid="stSidebarUserContent"] .stSlider > div > div > div {
  background: var(--fm-cyan) !important;
}
[data-testid="stSidebarUserContent"] input,
[data-testid="stSidebarUserContent"] select {
  color: #1a1a1a !important;
  background: white !important;
  border-radius: 5px !important;
}
[data-testid="stSidebarUserContent"] .stMarkdown p {
  color: rgba(255,255,255,0.70);
  font-size: 0.76rem;
}
[data-testid="stSidebarUserContent"] hr {
  border-color: rgba(0,163,224,0.25) !important;
  margin: 10px 0;
}

/* ── Onglets principaux ───────────────────────────────────── */
[data-testid="stTabs"] > div:first-child {
  background: white;
  border-radius: var(--fm-radius) var(--fm-radius) 0 0;
  padding: 0 8px;
  border-bottom: 2px solid var(--fm-border);
  gap: 4px;
}
button[data-baseweb="tab"] {
  font-size: 0.72rem !important;
  font-weight: 700 !important;
  letter-spacing: 0.06em !important;
  text-transform: uppercase !important;
  color: #555 !important;
  padding: 10px 16px !important;
  border: none !important;
  background: transparent !important;
  border-radius: 6px 6px 0 0 !important;
  transition: all 0.2s ease;
}
button[data-baseweb="tab"]:hover {
  color: var(--fm-navy) !important;
  background: rgba(27,31,107,0.06) !important;
}
button[aria-selected="true"] {
  color: var(--fm-navy) !important;
  border-bottom: 3px solid var(--fm-cyan) !important;
  background: rgba(0,163,224,0.05) !important;
}

/* ── Métriques ────────────────────────────────────────────── */
[data-testid="stMetricValue"] {
  font-size: 1.45rem !important;
  font-weight: 800 !important;
  color: var(--fm-navy) !important;
  letter-spacing: -0.02em;
}
[data-testid="stMetricLabel"] {
  font-size: 0.68rem !important;
  font-weight: 700 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.06em !important;
  color: #666 !important;
}
[data-testid="stMetricDelta"] {
  font-size: 0.78rem !important;
  font-weight: 600 !important;
}

/* ── Boutons ──────────────────────────────────────────────── */
.stButton > button {
  background: var(--fm-navy) !important;
  color: white !important;
  font-weight: 700 !important;
  font-size: 0.78rem !important;
  letter-spacing: 0.05em !important;
  text-transform: uppercase !important;
  border: none !important;
  border-radius: 6px !important;
  padding: 8px 20px !important;
  transition: all 0.2s ease !important;
  box-shadow: 0 2px 6px rgba(27,31,107,0.25) !important;
}
.stButton > button:hover {
  background: var(--fm-cyan) !important;
  transform: translateY(-1px);
  box-shadow: 0 4px 12px rgba(0,163,224,0.35) !important;
}
.stButton > button[kind="primary"] {
  background: var(--fm-cyan) !important;
}
.stButton > button[kind="primary"]:hover {
  background: #007BC0 !important;
}

/* ── DataFrames ───────────────────────────────────────────── */
[data-testid="stDataFrame"] {
  border: 1px solid var(--fm-border) !important;
  border-radius: var(--fm-radius) !important;
  overflow: hidden;
}
[data-testid="stDataFrame"] th {
  background: var(--fm-navy) !important;
  color: white !important;
  font-size: 0.72rem !important;
  font-weight: 700 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.04em !important;
  padding: 10px 12px !important;
}
[data-testid="stDataFrame"] td {
  font-size: 0.78rem !important;
  padding: 8px 12px !important;
  border-bottom: 1px solid #EEF0F8 !important;
}
[data-testid="stDataFrame"] tr:nth-child(even) td {
  background: #F7F9FE !important;
}
[data-testid="stDataFrame"] tr:hover td {
  background: #EBF5FF !important;
}

/* ── Expanders ────────────────────────────────────────────── */
[data-testid="stExpander"] {
  border: 1px solid var(--fm-border) !important;
  border-radius: var(--fm-radius) !important;
  background: white !important;
  box-shadow: var(--fm-shadow);
  margin-bottom: 8px;
}
[data-testid="stExpander"] summary {
  font-weight: 700 !important;
  font-size: 0.82rem !important;
  color: var(--fm-navy) !important;
  letter-spacing: 0.02em;
  padding: 12px 16px !important;
}

/* ── Messages (info/warning/success/error) ────────────────── */
[data-testid="stAlert"] {
  border-radius: var(--fm-radius) !important;
  font-size: 0.82rem !important;
  border-left-width: 4px !important;
}

/* ── Selectbox / Inputs ───────────────────────────────────── */
[data-testid="stSelectbox"] select,
[data-testid="stNumberInput"] input {
  border-radius: 6px !important;
  border-color: var(--fm-border) !important;
  font-size: 0.82rem !important;
}
[data-testid="stSelectbox"] select:focus,
[data-testid="stNumberInput"] input:focus {
  border-color: var(--fm-cyan) !important;
  box-shadow: 0 0 0 2px rgba(0,163,224,0.2) !important;
}

/* ── Dividers / Separators ────────────────────────────────── */
hr { border-color: var(--fm-border) !important; }

/* ── Sub-headers dans les onglets ─────────────────────────── */
h2 { 
  color: var(--fm-navy) !important;
  font-size: 1.20rem !important;
  font-weight: 800 !important;
  letter-spacing: -0.01em;
  border-bottom: 2px solid var(--fm-cyan);
  padding-bottom: 6px;
  margin-bottom: 16px;
}
h3 {
  color: var(--fm-navy) !important;
  font-size: 0.95rem !important;
  font-weight: 700 !important;
}

/* ── Progress bars ────────────────────────────────────────── */
[data-testid="stProgressBar"] > div > div {
  background: linear-gradient(90deg, var(--fm-navy), var(--fm-cyan)) !important;
}

/* ── Download buttons ─────────────────────────────────────── */
[data-testid="stDownloadButton"] > button {
  background: white !important;
  color: var(--fm-navy) !important;
  border: 2px solid var(--fm-navy) !important;
  font-weight: 700 !important;
  font-size: 0.76rem !important;
  letter-spacing: 0.04em !important;
  border-radius: 6px !important;
}
[data-testid="stDownloadButton"] > button:hover {
  background: var(--fm-navy) !important;
  color: white !important;
}

/* ── Scrollbar (webkit) ───────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #F0F4FA; }
::-webkit-scrollbar-thumb { background: #B8C0E0; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--fm-navy); }

/* ── HDR sections (custom class) ─────────────────────────── */
.hdr {
  background: linear-gradient(135deg, var(--fm-navy) 0%, #2A3094 60%, var(--fm-cyan) 100%);
  padding: 14px 22px;
  border-radius: var(--fm-radius);
  margin-bottom: 14px;
}
.hdr h2 {
  color: white !important;
  border-bottom: 1px solid rgba(255,255,255,0.25) !important;
  font-size: 1.05rem !important;
  margin-bottom: 4px !important;
  padding-bottom: 4px !important;
}
.hdr p { color: rgba(255,255,255,0.80); font-size: 0.80rem; margin: 0; }

/* ── Plotly chart containers ──────────────────────────────── */
[data-testid="stPlotlyChart"] {
  border: 1px solid var(--fm-border);
  border-radius: var(--fm-radius);
  overflow: hidden;
  background: white;
  box-shadow: var(--fm-shadow);
}

/* ── File uploader ────────────────────────────────────────── */
[data-testid="stFileUploader"] {
  border: 2px dashed var(--fm-border) !important;
  border-radius: var(--fm-radius) !important;
  background: #FAFBFE !important;
}
[data-testid="stFileUploader"]:hover {
  border-color: var(--fm-cyan) !important;
}</style>""", unsafe_allow_html=True)
st.markdown(PFE_TITLE, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# CONSTANTES
# ════════════════════════════════════════════════════════════════
C = {"S1":"#1A6B2E","S2":"#856404","S3":"#A71E1E","navy":"#003087","blue":"#00A3E0","green":"#1A6B2E","amber":"#856404","red":"#A71E1E","teal":"#0F6E56","purple":"#5B2D8E","gray":"#6B7280"}

MVARS = ["inflation","croissance_pib","chomage",
         "taux_change","masse_monetaire","taux_directeur"]

PD_HIST = {
    1994:.0690,1995:.0,1996:.0714,1997:.0,1998:.0233,1999:.0152,
    2000:.0,2001:.0588,2002:.1408,2003:.0145,2004:.0476,2005:.0488,
    2006:.0233,2007:.0706,2008:.0204,2009:.0291,2010:.0367,2011:.0642,
    2012:.0273,2013:.0145,2014:.0192,2015:.0363,2016:.0305,2017:.0193,
    2018:.0143,2019:.0493,2020:.0865,2021:.0204,2022:.0068,2023:.0310,2024:.0199}
PD_MOY = float(np.mean([v for v in PD_HIST.values() if v>0]))

PD_SEC = {
    "Financials":.0384,"Industials":.0670,"Industrials":.0670,
    "Consumer discretionary":.0664,"Consumer Discretionary":.0664,
    "Consumer Staples":.0474,"Materials":.0400,"Energie":.0506,
    "Communication Services":.0350,"souverain":.0339}

SECTEURS_LISTE = ["Financials","Industrials","Consumer Discretionary",
                   "Consumer Staples","Materials","Energie",
                   "Communication Services","souverain"]

# Règle IFRS 9 §B5.5.28 — stade selon nb d'échéances manquées
def stade_from_echeances(n):
    if n == 0: return 1
    if n == 1: return 2
    return 3  # ≥ 2 → défaut probable

# ════════════════════════════════════════════════════════════════
# FONCTIONS CORE
# ════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def load_oblig(b):
    """Charge Data oblig VF.xlsx — colonnes exactes du fichier Forvis Mazars."""
    df = pd.read_excel(io.BytesIO(b))

    # Nettoyer les noms de colonnes (espaces, accents)
    df.columns = [str(c).strip() for c in df.columns]

    # ── Colonnes émetteur ─────────────────────────────────────────────────────
    # DESCRIPTION (court) = nom réel de l'émetteur (ex: TRESOR, WAFA SALAF, OCP)
    #   → utilisé pour grouper les émissions multiples du même émetteur
    # Description complète = libellé complet de chaque obligation (unique)
    if "DESCRIPTION" in df.columns:
        df["NOM_EMETTEUR"]  = df["DESCRIPTION"].astype(str).str.strip()
    else:
        df["NOM_EMETTEUR"]  = "Inconnu"

    if "Description complète" in df.columns:
        df["LIBELLE_OBLIG"] = df["Description complète"].astype(str)
    else:
        df["LIBELLE_OBLIG"] = df["NOM_EMETTEUR"].copy()

    # ── Secteur → SECTEUR_GICS ────────────────────────────────────────────────
    if "Secteur" in df.columns:
        df["SECTEUR_GICS"] = df["Secteur"].astype(str)
    elif "SECTEUR" in df.columns:
        df["SECTEUR_GICS"] = df["SECTEUR"].astype(str)
    else:
        df["SECTEUR_GICS"] = "Autre"

    df["SECTEUR_GICS"] = df["SECTEUR_GICS"].replace({
        "Industials":"Industrials",
        "Consumer Discretionary":"Consumer discretionary",
    })

    # ── Spread ────────────────────────────────────────────────────────────────
    if "Spread émis (bps)" in df.columns:
        df["SPREAD_EMISSION"] = pd.to_numeric(df["Spread émis (bps)"], errors="coerce")
    else:
        df["SPREAD_EMISSION"] = 0.0

    # ── Maturité ──────────────────────────────────────────────────────────────
    if "Maturité (ans)" in df.columns:
        df["MATURITE_RESIDUELLE"] = pd.to_numeric(df["Maturité (ans)"], errors="coerce")
    else:
        df["MATURITE_RESIDUELLE"] = 5.0

    # ── Coupon ────────────────────────────────────────────────────────────────
    if "Coupon (%)" in df.columns:
        df["COUPON"] = pd.to_numeric(df["Coupon (%)"], errors="coerce")

    # ── PD_TTC ────────────────────────────────────────────────────────────────
    # Dans Data oblig VF : colonne "PD" en valeur décimale (ex: 0.0384)
    if "PD" in df.columns:
        raw = pd.to_numeric(df["PD"], errors="coerce")
        # Si valeurs > 1 → en pourcentage, sinon déjà décimal
        df["PD_TTC"] = (raw / 100.0).where(raw > 1, raw).clip(1e-6, 1-1e-6)
    else:
        # PD sectorielle par défaut
        pd_sec_def = {"Financials":0.0384,"Industrials":0.0670,
                      "Consumer discretionary":0.0664,"Consumer Staples":0.0474,
                      "Materials":0.0400,"Energie":0.0506,
                      "Communication Services":0.0350,"souverain":0.0339}
        df["PD_TTC"] = df["SECTEUR_GICS"].map(pd_sec_def).fillna(0.04)

    # ── LGD ───────────────────────────────────────────────────────────────────
    if "LGD" in df.columns:
        df["LGD_TTC"] = pd.to_numeric(df["LGD"], errors="coerce").clip(0.005, 0.995)

    # ── Dates ─────────────────────────────────────────────────────────────────
    for src_col, dst_col in [("DATE_EMISSION","DATE_EMISSION"),
                               ("DATE_ECHEANCE","DATE_ECHEANCE"),
                               ("DATE_MAJ","DATE_MAJ")]:
        if src_col in df.columns:
            df[dst_col] = pd.to_datetime(df[src_col], errors="coerce")

    # ── Seuil de défaut Merton α_i = Φ⁻¹(PD_TTC) ────────────────────────────
    from scipy.stats import norm as _norm
    df["ALPHA_I"] = _norm.ppf(df["PD_TTC"])

    return df


def load_macro(b):
    """Charge Data facteur macro.xlsx — colonnes exactes du fichier BAM."""
    df = pd.read_excel(io.BytesIO(b))

    # Strip espaces (certaines colonnes ont des espaces en fin)
    df.columns = [str(c).strip() for c in df.columns]

    # Renommage vers noms normalisés utilisés dans le modèle
    df.rename(columns={
        "Année":                    "annee",
        "Inflation":                "inflation",
        "PIB":                      "pib",
        "Chomage":                  "chomage",
        "Croissance PIB":           "croissance_pib",
        "Taux de change":           "taux_change",
        "Masse monétaire":          "masse_monetaire",
        "Population active":        "pop_active",
        "Depense en consommation":  "depense_conso",
        "Balance exterieur B&S":    "balance_exterieure",
        "Taux directeur":           "taux_directeur",
    }, inplace=True, errors="ignore")

    # Convertir tout en numérique sauf annee
    for c in df.columns:
        if c != "annee":
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # S'assurer que annee est entier
    if "annee" in df.columns:
        df["annee"] = pd.to_numeric(df["annee"], errors="coerce").fillna(0).astype(int)

    return df.sort_values("annee").reset_index(drop=True) if "annee" in df.columns else df


def run_acp(df_m, mvars):
    vs = [v for v in mvars if v in df_m.columns]
    M  = df_m[vs].fillna(df_m[vs].mean()).values
    Ms = StandardScaler().fit_transform(M)
    pca = PCA(n_components=min(len(vs),3), random_state=42)
    pca.fit(Ms)
    pc1 = pca.transform(Ms)[:,0]
    if "croissance_pib" in vs:
        if np.corrcoef(pc1, df_m["croissance_pib"].values)[0,1] > 0:
            pc1 = -pc1
    pc1z   = zscore(pc1)
    stress = MinMaxScaler().fit_transform(pc1.reshape(-1,1)).flatten()
    return pc1, pc1z, stress, pca, vs


def ar1_proj(z, ma_w):
    if len(z) < 4:
        return z[-1], z[-ma_w:].mean(), z[-1], 0., 0., 0.
    Zl = z[:-1].reshape(-1,1); Zc = z[1:]
    c, *_ = lstsq(np.column_stack([np.ones(len(Zl)), Zl]), Zc)
    ph0, ph1 = float(c[0]), float(c[1])
    sig = float((Zc-(ph0+ph1*Zl.flatten())).std())
    zar = ph0+ph1*z[-1]; zma = float(z[-ma_w:].mean())
    return float(zar), zma, float(.5*zar+.5*zma), ph0, ph1, sig


def calib_seuils(df, sc, rc):
    sn = df[df[rc]=="Neutre"][sc].dropna()
    sa = df[df[rc]=="Adverse"][sc].dropna()
    s12 = float(sn.quantile(.75)) if len(sn)>0 else 70.
    s23 = (float(sa.median()) if len(sa)>=3
           else float(sn.quantile(.90)) if len(sn)>0 else 150.)
    if s23<=s12: s23 = s12*1.8
    return round(s12,1), round(s23,1)


def run_kmeans(spreads, kmin):
    X = spreads.dropna().values.reshape(-1,1)
    # Guard: need at least kmin samples for clustering
    if len(X) < max(kmin, 3):
        # Not enough data: assign all to cluster 0
        n_total = len(spreads)
        return np.zeros(n_total, dtype=int), kmin, {}
    sil = {}
    kmax = min(9, len(X))
    for k in range(3, kmax):
        try:
            km  = KMeans(n_clusters=k, random_state=42, n_init=20)
            lbl = km.fit_predict(X)
            if len(set(lbl)) > 1:
                sil[k] = silhouette_score(X, lbl)
        except Exception:
            continue
    kopt = max(sil, key=sil.get) if sil else kmin
    K    = max(kmin, min(kopt, len(X)))
    try:
        km   = KMeans(n_clusters=K, random_state=42, n_init=50)
        lbl  = km.fit_predict(X)
        ctrs = km.cluster_centers_.flatten()
        rank = {o:n for n,o in enumerate(np.argsort(ctrs))}
        return np.array([rank[l] for l in lbl]), K, sil
    except Exception:
        return np.zeros(len(X), dtype=int), K, sil


def compute_evol(df, sc):
    rows = {}
    for em, grp in df.groupby("NOM_EMETTEUR"):
        sub = grp.dropna(subset=[sc,"DATE_EMISSION"]).sort_values("DATE_EMISSION")
        n   = len(sub)
        r   = dict(n=n, s0=np.nan, sT=np.nan, delta=0.,
                   pente=0., pval=1., tend="—", stade=1)
        if n >= 2:
            r["s0"] = float(sub[sc].iloc[0])
            r["sT"] = float(sub[sc].iloc[-1])
            r["delta"] = r["sT"] - r["s0"]
            x = (sub["DATE_EMISSION"]-sub["DATE_EMISSION"].min()).dt.days.values.astype(float)
            y = sub[sc].values
            if x.std() > 0:
                sl,_,_,pv,_ = stats.linregress(x, y)
                r["pente"] = round(sl*365.25,2); r["pval"] = round(pv,4)
                r["tend"] = ("Dégradation" if sl>0 else " Amélioration") if pv<.10 else "Stable"
        d = r["delta"]
        r["stade"] = 3 if d>=50 else (2 if d>=0 else 1)
        rows[em] = r
    return pd.DataFrame(rows).T


def mc_spread(sobs, st0, st25, nsim):
    rng   = np.random.default_rng(42)
    betas = rng.normal(.35, .15, nsim)
    eps   = rng.normal(0, .15, nsim)
    dS    = (st25-st0) + eps
    return np.maximum(sobs*(1+betas*np.maximum(0,dS)), .1)


def sim_retard_hybride(base_jours, pd_ttc, delta_stress, nsim=3000, seed=42):
    rng        = np.random.default_rng(seed)
    pd_adj     = float(np.clip(pd_ttc*np.exp(1.5*max(0,delta_stress)), 1e-6, .999))
    stress_mult = 1. + 2.*max(0, delta_stress)
    if base_jours <= 0:
        has = rng.binomial(1, pd_adj, nsim).astype(bool)
        mag = rng.exponential(60.*stress_mult, nsim)
        ret = np.where(has, mag, 0.)
    else:
        ret = rng.gamma(2., base_jours/2., nsim) * stress_mult
    return np.maximum(ret, 0.), pd_adj


def export_excel(df_r, S12, S23, s_s2, s_s3, date_calc):
    wb = Workbook()
    def _f(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    def _h(sz=10): return Font(name="Calibri",bold=True,color="FFFFFF",size=sz)
    thin = Side(border_style="thin",color="D0D0D0")
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)
    CTR  = Alignment(horizontal="center",vertical="center")
    LFT  = Alignment(horizontal="left",vertical="center")
    SF   = {"S1":"D6EAD7","S2":"FFF3CD","S3":"F8D7DA"}

    ws = wb.active; ws.title = "Staging_IFRS9"
    cols = ["NOM_EMETTEUR","SECTEUR_GICS","SPREAD_EMISSION","DATE_EMISSION",
            "REGIME_EM","C1_STADE","SPREAD_ADJ","C2_STADE",
            "DELTA_SPREAD","C3_STADE","NB_ECHEANCES_MANQUEES","C5_STADE",
            "JOURS_RETARD","C4_STADE","SNUM","STADE","DEC",
            "PS1","PS2","PS3"]
    cols = [c for c in cols if c in df_r.columns]
    HDRS = {
        "NOM_EMETTEUR":"Émetteur","SECTEUR_GICS":"Secteur",
        "SPREAD_EMISSION":"Spread émis (bps)","DATE_EMISSION":"Date émission",
        "REGIME_EM":"Régime émission","C1_STADE":"→C1 K-Means",
        "SPREAD_ADJ":"Spread ajusté","C2_STADE":"→C2 Macro ACP",
        "DELTA_SPREAD":"ΔSpread (bps)","C3_STADE":"→C3 Évol.émetteur",
        "NB_ECHEANCES_MANQUEES":"Échéances manquées","C5_STADE":"→C5 Échéances",
        "JOURS_RETARD":"Retard Q75 (j)","C4_STADE":"→C4 Spread",
        "SNUM":"MAX(C1..C5)","STADE":" STADE FINAL","DEC":"Déclencheur",
        "PS1":"P(S1)%","PS2":"P(S2)%","PS3":"P(S3)%"}
    GCOL = {
        "NOM_EMETTEUR":"003087","SECTEUR_GICS":"003087",
        "SPREAD_EMISSION":"003087","DATE_EMISSION":"003087","REGIME_EM":"4A6741",
        "C1_STADE":"1A6B2E","SPREAD_ADJ":"2E5E1A","C2_STADE":"2E5E1A",
        "DELTA_SPREAD":"5C3317","C3_STADE":"5C3317",
        "NB_ECHEANCES_MANQUEES":"6B3B00","C5_STADE":"6B3B00",
        "JOURS_RETARD":"3B3B8A","C4_STADE":"3B3B8A",
        "SNUM":"A71E1E","STADE":"A71E1E","DEC":"A71E1E",
        "PS1":"1A6B2E","PS2":"856404","PS3":"A71E1E"}

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws["A1"]
    c.value = (f"FORVIS MAZARS — IFRS 9 Staging Dynamique | "
               f"S1/S2={S12}bps | S2/S3={S23}bps | "
               f"Retard S2≥{s_s2}j | Retard S3≥{s_s3}j | Date={date_calc}")
    c.font = _h(11); c.fill = _f("003087"); c.alignment = CTR
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{get_column_letter(len(cols))}2")
    c2 = ws["A2"]
    c2.value = ("C1:K-Means | C2:ACP+AR(1)/MA | C3:Régression émetteur | "
                "C4:Retards simulés | C5:Échéances manquées | STADE=MAX(C1..C5)")
    c2.font  = Font(name="Calibri",bold=True,size=9,color="003087")
    c2.fill  = _f("E8F0F7"); c2.alignment = CTR; ws.row_dimensions[2].height = 16

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=3, column=ci, value=HDRS.get(col, col))
        cell.font = _h(9); cell.fill = _f(GCOL.get(col,"003087"))
        cell.alignment = CTR; cell.border = bdr
    ws.row_dimensions[3].height = 36

    for ri, (_,row) in enumerate(df_r[cols].iterrows(), 4):
        stade = str(row.get("STADE","S1"))
        bg    = "F5F7FA" if ri%2==0 else "FFFFFF"
        for ci, col in enumerate(cols, 1):
            val  = row[col]; cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bdr
            if col == "STADE":
                cell.fill = _f(SF.get(stade,"FFFFFF"))
                cell.font = Font(name="Calibri",bold=True,size=11); cell.alignment = CTR
            elif col in ["C1_STADE","C2_STADE","C3_STADE","C4_STADE","C5_STADE","SNUM"]:
                sv = int(val) if pd.notna(val) else 1
                cell.fill = _f({1:"D6EAD7",2:"FFF3CD",3:"F8D7DA"}.get(sv,"FFF"))
                cell.font = Font(name="Calibri",bold=True,size=10); cell.alignment = CTR
            elif col == "REGIME_EM":
                cell.fill = _f({"Adverse":"F8D7DA","Neutre":"FFF3CD","Favorable":"D6EAD7"}.get(str(val),"FFF"))
                cell.font = Font(name="Calibri",bold=True,size=10); cell.alignment = CTR
            elif col == "NB_ECHEANCES_MANQUEES":
                nv = int(val) if pd.notna(val) else 0
                ec = "F8D7DA" if nv>=2 else ("FFF3CD" if nv==1 else "D6EAD7")
                cell.fill = _f(ec); cell.font = Font(name="Calibri",bold=True,size=10)
                cell.alignment = CTR
            else:
                cell.font = Font(name="Calibri",size=10)
                cell.alignment = LFT if col=="NOM_EMETTEUR" else CTR
                cell.fill = _f(bg)

    WW = {"NOM_EMETTEUR":26,"SECTEUR_GICS":22,"SPREAD_EMISSION":14,"DATE_EMISSION":14,
          "REGIME_EM":20,"DEC":28,"STADE":14,"NB_ECHEANCES_MANQUEES":18}
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WW.get(col, 13)

    # Onglet récap
    ws2 = wb.create_sheet("Récapitulatif")
    ws2["A1"] = "FORVIS MAZARS — Récapitulatif Staging IFRS 9"
    ws2["A1"].font = _h(13); ws2["A1"].fill = _f("003087")
    ws2["A1"].alignment = CTR; ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 22
    for ci,h in enumerate(["Paramètre","Valeur","Source","Justification"],1):
        c = ws2.cell(row=2,column=ci,value=h)
        c.font=_h(); c.fill=_f("2E4D7B"); c.alignment=CTR; c.border=bdr
    recap = [
        ["Seuil S1/S2",f"{S12} bps","Q75 spreads neutres","Calibré par crises (ACP Mann-Whitney)"],
        ["Seuil S2/S3",f"{S23} bps","Médiane spreads adverses","Niveau typique en période de crise"],
        ["Retard S2",f"≥{s_s2}j","Utilisateur","Pratique bancaire"],
        ["Retard S3",f"≥{s_s3}j","IFRS 9 §B5.5.2","Présomption réfutable de défaut"],
        ["Échéances S2","1 manquée","IFRS 9 §B5.5.28","Signal de difficultés de paiement"],
        ["Échéances S3","≥2 manquées","IFRS 9 §B5.5.28","Défaut probable"],
        ["S1",f"{(df_r['STADE']=='S1').sum()} titres","MAX(C1..C5)","ECL 12 mois"],
        ["S2",f"{(df_r['STADE']=='S2').sum()} titres","MAX(C1..C5)","ECL durée de vie"],
        ["S3",f"{(df_r['STADE']=='S3').sum()} titres","MAX(C1..C5)","ECL durée de vie — défaut"],
    ]
    for ri,(p,v,s,j) in enumerate(recap,3):
        bg = "F5F7FA" if ri%2==0 else "FFFFFF"
        for ci,val in enumerate([p,v,s,j],1):
            cell = ws2.cell(row=ri,column=ci,value=val)
            cell.font = Font(name="Calibri",bold=(ci==1),size=10)
            cell.alignment = CTR if ci<=2 else Alignment(horizontal="left",vertical="center",wrap_text=True)
            cell.border = bdr; cell.fill = _f(bg)
        ws2.row_dimensions[ri].height = 34
    for ci,w in enumerate([20,16,38,52],1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""<div style="background:#003087;border-radius:8px;
padding:10px;margin-bottom:10px;text-align:center">{LOGO_SVG}</div>""",
    unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### Fichiers de données")
    st.caption("Data oblig VF.xlsx  ·  Data facteur macro.xlsx")
    f_ob = st.file_uploader("Fichier 1 — Portefeuille obligataire\n(Data oblig VF.xlsx)", type=["xlsx","xls"], key="ob")
    f_ma = st.file_uploader("Fichier 2 — Données macro\n(Data facteur macro.xlsx)", type=["xlsx","xls"], key="ma")
    st.markdown("---")
    st.markdown("** K-Means**")
    kmin = st.slider("K minimum", 3, 8, 5, key="k_minimum")
    st.markdown("** Monte Carlo**")
    nsim = st.slider("N simulations", 500, 6000, 3000, 500, key="n_simulations")
    st.markdown("** Macro**")
    maw  = st.slider("Fenêtre MA", 2, 5, 3, key="fen_tre_ma")
    zadv = st.slider("Seuil Z Adverse",   .50, 1.50,  .75, .05, key="seuil_z_adverse")
    zfav = st.slider("Seuil Z Favorable",-1.50, -.50, -.75, .05, key="seuil_z_favorable")
    st.markdown("---")
    s_s2 = 30
    s_s3 = 90
    RETARDS_SEC = {sec: 0 for sec in SECTEURS_LISTE}
    st.markdown("---")
    date_calc = st.date_input("Date de reporting", value=pd.Timestamp("2025-04-25"), key="wk_da_0004")
    st.caption("© Forvis Mazars | IFRS 9 Staging Dynamique v2")


# ════════════════════════════════════════════════════════════════
# HEADER
# ════════════════════════════════════════════════════════════════
st.markdown(f"""<div class="hdr">
<div>{LOGO_SVG}</div>
<div class="hdr-txt">
<h2>IFRS 9 — Portefeuille Obligataire Marocain · 328 Obligations</h2>
<p>Staging Dynamique &nbsp;&middot;&nbsp; PD Merton-Vasicek &nbsp;&middot;&nbsp; PD CPV &nbsp;&middot;&nbsp; LGD Vasicek &amp; Frye-Jacobs &nbsp;&middot;&nbsp; EAD &nbsp;&middot;&nbsp; ECL IFRS 9 &nbsp;&middot;&nbsp; Analyse &amp; Interpretation</p>
</div></div>""", unsafe_allow_html=True)

if not f_ob or not f_ma:
    st.info("⬅️ Uploadez le portefeuille et les données macro dans la sidebar.")
    c1,c2,c3 = st.columns(3)
    c1.markdown("**Fichier 1 — Portefeuille**\n- `SPREAD_EMISSION` (bps)\n- `DATE_EMISSION`\n- `NOM_EMETTEUR`\n- `SECTEUR_GICS`")
    c2.markdown("**Fichier 2 — Macro**\n`Data facteur macro.xlsx`\n- Année, Inflation, PIB\n- Croissance PIB, Chômage\n- Taux directeur, Masse monétaire")
    c3.markdown("**Onglet Échéances**\n- Saisie manuelle par titre\n- 0 → S1, 1 → S2, ≥2 → S3\n- Réf. IFRS 9 §B5.5.28")
    st.info("**Fichiers attendus :**\n- `Data oblig VF.xlsx`\n- `Data facteur macro.xlsx`\n- (optionnel) `Historique GEMS LGD.xlsx`")
    st.stop()


# ════════════════════════════════════════════════════════════════
# CALCULS
# ════════════════════════════════════════════════════════════════
with st.spinner("Calculs en cours…"):
    df_raw = load_oblig(f_ob.read())
    df_mac = load_macro(f_ma.read())

    SC = "SPREAD_EMISSION"
    if SC not in df_raw.columns:
        st.error("Colonne SPREAD_EMISSION introuvable."); st.stop()

    DC = pd.Timestamp(date_calc)
    # Conserver TOUTES les 328 obligations (dont souverains avec spread=0)
    # Les obligations à spread=0 (TRESOR/souverain) seront classées S1 directement
    df = df_raw.copy().reset_index(drop=True)
    # Valeur par défaut pour les spreads nuls (évite division par zéro)
    df[SC] = pd.to_numeric(df[SC], errors="coerce").fillna(0)
    # ANC = ancienneté + ANNEE_EMISSION — gérer les dates manquantes
    df["DATE_EMISSION"] = pd.to_datetime(df.get("DATE_EMISSION", pd.NaT), errors="coerce")
    df["ANC"]           = (DC - df["DATE_EMISSION"]).dt.days.fillna(0) / 365.25
    df["ANNEE_EMISSION"] = df["DATE_EMISSION"].dt.year.fillna(DC.year).astype(int)

    # ACP + AR(1)/MA
    PC1, PC1z, stress_n, pca_m, mvs = run_acp(df_mac, MVARS)
    df_mac["PC1z"]   = PC1z
    df_mac["stress"] = stress_n
    df_mac["REGIME"] = df_mac["PC1z"].apply(
        lambda z: "Adverse" if z>zadv else ("Neutre" if z>zfav else "Favorable"))

    Z25ar, Z25ma, Z25, ph0, ph1, sig_ar = ar1_proj(PC1z, maw)
    Z25lo = Z25ar - 1.645*sig_ar
    Z25hi = Z25ar + 1.645*sig_ar
    st25  = float(MinMaxScaler().fit_transform(PC1.reshape(-1,1)).flatten()[-1])

    _annees_int = df_mac["annee"].astype("int64")
    rm = dict(zip(_annees_int, df_mac["REGIME"]))
    sm = dict(zip(_annees_int, df_mac["stress"]))
    zm = dict(zip(_annees_int, df_mac["PC1z"]))
    df["REGIME_EM"]  = df["ANNEE_EMISSION"].apply(lambda y: rm.get(int(y),"Neutre") if pd.notna(y) else "Neutre")
    df["STRESS_T0"]  = df["ANNEE_EMISSION"].apply(lambda y: sm.get(int(y),.5)     if pd.notna(y) else .5)
    df["Z_T0"]       = df["ANNEE_EMISSION"].apply(lambda y: zm.get(int(y),0.)     if pd.notna(y) else 0.)

    # Seuils calibrés
    S12, S23 = calib_seuils(df, SC, "REGIME_EM")

    # C1 — K-Means
    # K-Means uniquement sur les obligations avec spread > 0
    df_spread_pos = df[df[SC] > 0]
    if len(df_spread_pos) >= kmin:
        km_lbl_pos, K, sil_d = run_kmeans(df_spread_pos[SC], kmin)
        # Assigner les clusters (spread=0 → cluster 0 = risque le plus faible)
        df["CLU"] = 0
        spread_pos_idx = df[df[SC] > 0].index
        if len(km_lbl_pos) == len(spread_pos_idx):
            df.loc[spread_pos_idx, "CLU"] = km_lbl_pos.astype(int)
        km_lbl = km_lbl_pos; sil_d = sil_d
    else:
        km_lbl, K, sil_d = run_kmeans(df[SC], kmin)
    # CLU déjà assigné ci-dessus (spread>0 uniquement)
    NIV = ["Très faible","Faible","Modéré","Élevé","Très élevé"]
    df_int = pd.DataFrame([{
        "c":c, "niv":(NIV[c] if K==5 else f"N{c+1}"),
        "mn":round(float(df[df["CLU"]==c][SC].min()),1),
        "mx":round(float(df[df["CLU"]==c][SC].max()),1),
        "ct":round(float(df[df["CLU"]==c][SC].mean()),1),
        "n":int((df["CLU"]==c).sum())} for c in range(K)])

    def c2s(c):
        ct = float(df_int[df_int["c"]==int(c)]["ct"].values[0])
        return 3 if ct>=S23 else (2 if ct>=S12 else 1)

    df["C1_STADE"] = df["CLU"].apply(c2s)

    # C2 — Spread ajusté macro
    df["DELTA_STRESS"] = np.maximum(0., st25 - df["STRESS_T0"])
    # SPREAD_ADJ : pour spread=0 (souverain), garder 0 (pas d'ajustement cyclique)
    df["SPREAD_ADJ"]   = np.where(df[SC] == 0, 0.0,
                            df[SC] * (1 + df["DELTA_STRESS"] / df["STRESS_T0"].clip(lower=.01)))
    df["C2_STADE"]     = df["SPREAD_ADJ"].apply(lambda s: 3 if s>=S23 else (2 if s>=S12 else 1))

    # C3 — Évolution spread émetteur
    df_evol  = compute_evol(df, SC)
    df["C3_STADE"]    = df["NOM_EMETTEUR"].map(lambda e: int(df_evol.loc[e,"stade"]) if e in df_evol.index else 1)
    df["DELTA_SPREAD"]= df["NOM_EMETTEUR"].map(lambda e: float(df_evol.loc[e,"delta"]) if e in df_evol.index else 0.)

    # Monte Carlo spread
    mc_rows = []
    for _, r in df.iterrows():
        sims = mc_spread(r[SC], r["STRESS_T0"], st25, nsim)
        p1=(sims<S12).mean(); p2=((sims>=S12)&(sims<S23)).mean(); p3=(sims>=S23).mean()
        mc_rows.append(dict(MCQ50=round(float(np.quantile(sims,.50)),1),
                             MCQ75=round(float(np.quantile(sims,.75)),1),
                             MCQ95=round(float(np.quantile(sims,.95)),1),
                             PS1=round(p1*100,1), PS2=round(p2*100,1),
                             PS3=round(p3*100,1), FIAB=round(max(p1,p2,p3)*100,1)))
    df = df.join(pd.DataFrame(mc_rows, index=df.index))

    # C4 — Retards simulés
    def get_base_j(sec):
        sec = str(sec or "Financials")
        for k,v in RETARDS_SEC.items():
            if k.lower() in sec.lower() or sec.lower() in k.lower():
                return v
        return 0

    ret_rows = []
    for _, r in df.iterrows():
        sec   = str(r.get("SECTEUR_GICS","Financials") or "Financials")
        pd_t  = PD_SEC.get(sec, PD_MOY)
        dst   = float(max(0, st25-float(r["STRESS_T0"])))
        bj    = get_base_j(sec)
        sims_r, pd_adj = sim_retard_hybride(bj, pd_t, dst, nsim)
        q50,q75,q90    = [float(np.quantile(sims_r,q)) for q in (.5,.75,.9)]
        sr = 3 if q75>=s_s3 else (2 if q75>=s_s2 else 1)
        ret_rows.append(dict(RQ50=round(q50,1), RQ75=round(q75,1),
                              RQ90=round(q90,1), PDAJ=round(pd_adj*100,3),
                              C4_STADE=sr))
    df = df.join(pd.DataFrame(ret_rows, index=df.index))
    df["JOURS_RETARD"] = df["RQ75"]

    # C5 — Échéances manquées : initialiser à 0 (sera mis à jour via st.data_editor)
    if "NB_ECHEANCES_MANQUEES" not in df.columns:
        df["NB_ECHEANCES_MANQUEES"] = 0
    df["C5_STADE"] = df["NB_ECHEANCES_MANQUEES"].apply(stade_from_echeances)

    # STADE FINAL = MAX
    df["SNUM"]  = df[["C1_STADE","C2_STADE","C3_STADE","C4_STADE","C5_STADE"]].max(axis=1)
    df["STADE"] = df["SNUM"].map({1:"S1",2:"S2",3:"S3"})
    # Obligations souveraines (spread=0 : TRESOR, etc.) → S1 par definition IFRS 9
    mask_souv = df[SC] == 0
    if mask_souv.any():
        df.loc[mask_souv, "SNUM"]     = 1
        df.loc[mask_souv, "STADE"]    = "S1"
        df.loc[mask_souv, "C1_STADE"] = 1
        df.loc[mask_souv, "C2_STADE"] = 1
        df.loc[mask_souv, "C3_STADE"] = 1
        df.loc[mask_souv, "C4_STADE"] = 1
        df.loc[mask_souv, "C5_STADE"] = 1

    def dec(row):
        mx=row["SNUM"]; p=[]
        if row["C1_STADE"]==mx: p.append("K-Means")
        if row["C2_STADE"]==mx: p.append("Spread adj.")
        if row["C3_STADE"]==mx: p.append("Évol.émetteur")
        if row["C4_STADE"]==mx: p.append("Retard")
        if row["C5_STADE"]==mx: p.append("Échéances")
        return " + ".join(p)

    df["DEC"]  = df.apply(dec, axis=1)
    dist = df["STADE"].value_counts().sort_index()
    TOT  = len(df)

st.success(f" {TOT} obligations · {df['NOM_EMETTEUR'].nunique()} émetteurs · "
           f"S1/S2={S12} bps · S2/S3={S23} bps · Date : {date_calc}")


# ════════════════════════════════════════════════════════════════
# ONGLETS
# ════════════════════════════════════════════════════════════════
TABS = st.tabs([
    "Vue d\'ensemble",
    "Staging",
    "PD Vasicek",
    "PD CPV",
    "LGD PIT",
    "EAD avec PD PIT",
    "ECL IFRS 9",
    "Analyse & Interpretation",
    "Comparaison Scenarios",
    "Risque UL & VaR",
    "Rapport PDF",
])
tab_staging = 1
tab_pd      = 2
tab_cpv     = 3
tab_lgd     = 4
tab_ead     = 5
tab_comp    = 8
tab_risq    = 9
tab_rapp    = 10
tab_ecl     = 6

def _run_ecl_global(df_in, T_p=5, T_r=3, d_p=0.420, d_c=-0.252, d_o=-0.850,
                     w_p=0.20, w_c=0.70, w_o=0.10, r_flat=0.035):
    """
    ECL IFRS 9 — methode exacte notebook Actualisation_ECL__1_.ipynb
    Actualisation : TIE_i = r_ZC(T_mat_i) + Spread_i/10000  (individuel par obligation)
    r_ZC : courbe Nelson-Siegel-Svensson calibree sur BAM 08/05/2026
    PD reversion : interpolation PD annuelle
    LGD PIT : Frye-Jacobs avec Z observe
    """
    from scipy.stats import norm as _Ng
    from scipy.optimize import least_squares as _lsq
    import math as _mg
    from datetime import datetime as _dt, date as _date

    # ── Formules ─────────────────────────────────────────────────────────
    _ph  = lambda x: _Ng.cdf(np.clip(x,-30,30))
    _phi = lambda p: _Ng.ppf(np.clip(p,1e-8,1-1e-8))
    def _rho(p): w=(1-np.exp(-50*p))/(1-np.exp(-50)); return 0.12*w+0.24*(1-w)
    def _lpit(pt,lt,rho,Z):
        p=np.clip(pt*lt,1e-8,1-1e-8)
        pp=np.clip(_ph((_phi(pt)-np.sqrt(rho)*Z)/np.sqrt(1-rho)),1e-8,1-1e-8)
        return float(np.clip(_ph(_phi(pp)-(_phi(pt)-_phi(p))/np.sqrt(1-rho))/pp,0.001,0.999))
    def _pd_ann(pt,pp,t,Tp,Tr):
        if t<=Tp: return pp
        if t>Tp+Tr: return pt
        return pt+(pp-pt)*(Tp+Tr-t)/Tr
    def _pd_marg(pt,pp,t,Tp,Tr,cpd_prev):
        return _pd_ann(pt,pp,t,Tp,Tr)*(1-cpd_prev)
    def _ead_t(e0,meth,mr,pp,t):
        e=float(e0)
        for ti in range(1,int(t)+1):
            am=0.0 if str(meth).upper() in("ZC","TA") else 1.0/max(float(mr)-ti+1,1.0)
            e=max(0.0,e*(1-float(pp))*(1-am))
        return e

    # ── Courbe BAM / Nelson-Siegel-Svensson ───────────────────────────────
    def _build_nss():
        """Construit la courbe zero-coupon BAM via bootstrap + NSS."""
        DATE_VAL = _date(2026,5,7)
        _bam_raw = [
            (0.20, 0.0221),(0.28, 0.0226),(0.78, 0.0229),(1.37, 0.0234),
            (3.94, 0.0270),(5.45, 0.0282),(9.11, 0.0308),(13.20, 0.0343),
            (19.27, 0.0368),(28.95, 0.0401)
        ]
        T_obs=np.array([x[0] for x in _bam_raw])
        R_obs=np.array([x[1] for x in _bam_raw])
        # Bootstrap zero-coupon
        zs=np.zeros(len(T_obs))
        for i,(T,r) in enumerate(zip(T_obs,R_obs)):
            if T<=1.0: zs[i]=r
            else:
                n_f=int(np.floor(T)); tc=list(range(1,n_f+1)); tc[-1]=T
                pv=0.0
                for t_c in tc[:-1]:
                    zt=float(np.interp(t_c,T_obs[:i],zs[:i])) if i>0 else zs[0]
                    pv+=r/(1+zt)**t_c
                d=1-pv
                zs[i]=((1+r)/d)**(1/T)-1 if d>0 else r
        # NSS calibration
        def nss_f(T,b0,b1,b2,b3,t1,t2):
            T=np.atleast_1d(T).astype(float); t1=max(t1,0.01); t2=max(t2,0.01)
            x1=T/t1; x2=T/t2
            f1=np.where(x1>1e-6,(1-np.exp(-x1))/x1,1-x1/2)
            f2=np.where(x2>1e-6,(1-np.exp(-x2))/x2,1-x2/2)
            return b0+b1*f1+b2*(f1-np.exp(-x1))+b3*(f2-np.exp(-x2))
        try:
            res=_lsq(lambda p,T,z: nss_f(T,*p)-z,
                     x0=[zs[-1],zs[0]-zs[-1],0.01,0.0,2.0,4.0],
                     args=(T_obs,zs),
                     bounds=([0.001,-0.20,-0.20,-0.20,0.1,0.1],
                             [0.20, 0.20, 0.20, 0.20,30.,30.]),
                     method='trf',max_nfev=20000,ftol=1e-14)
            P=res.x
            def r_zc(T): return float(np.clip(nss_f(np.atleast_1d([T]),*P)[0],0.001,0.25))
        except:
            # Fallback: interpolation lineaire
            def r_zc(T):
                return float(np.clip(np.interp(float(T),T_obs,zs,left=zs[0],right=zs[-1]),0.001,0.25))
        return r_zc

    _r_zc = _build_nss()

    # ── Preparation donnees ───────────────────────────────────────────────
    d=df_in.copy()
    if "PD_TTC" not in d.columns: d["PD_TTC"]=pd.to_numeric(d.get("PD",0.05),errors="coerce")/100
    if "LGD_TTC" not in d.columns: d["LGD_TTC"]=pd.to_numeric(d.get("LGD",0.45),errors="coerce")
    d["PD_TTC"]=d["PD_TTC"].clip(1e-6,1-1e-6); d["LGD_TTC"]=d["LGD_TTC"].clip(0.005,0.995)
    d["NOMINAL"]=pd.to_numeric(d.get("NOMINAL",1e5),errors="coerce").fillna(1e5)
    mc=next((c for c in d.columns if "aturit" in c and "ans" in c.lower()),None)
    d["mat_res"]=pd.to_numeric(d[mc],errors="coerce").fillna(5).clip(lower=0.1) if mc else 5.0
    if "METHODE_VALO" not in d.columns: d["METHODE_VALO"]="ZC"
    d["METHODE_VALO"]=d["METHODE_VALO"].astype(str).str.strip().str.upper()
    for c2 in["DATE_JOUISSANCE","DATE_ECHEANCE"]:
        if c2 in d.columns:
            try: d[c2]=pd.to_datetime(d[c2],errors="coerce")
            except: pass
    if "DATE_JOUISSANCE" in d.columns and "DATE_ECHEANCE" in d.columns:
        mt2=(d["DATE_ECHEANCE"]-d["DATE_JOUISSANCE"]).dt.days/365.25
    else: mt2=None
    def _e0(r):
        m=str(r["METHODE_VALO"]).upper(); n=float(r["NOMINAL"]); mr=float(r["mat_res"])
        if m in("ZC","TA"): return n
        m2=float(mt2.iloc[r.name]) if mt2 is not None and pd.notna(mt2.iloc[r.name]) else None
        return n*(mr/m2) if m2 and m2>0 else n
    d["EAD_0"]=d.apply(_e0,axis=1)

    # STADE
    stc=next((c for c in["STADE","STADE_FINAL","SNUM"] if c in d.columns),None)
    if stc:
        ns={"1":"S1","2":"S2","3":"S3","s1":"S1","s2":"S2","s3":"S3"}
        d["STADE_ECL"]=[ns.get(str(v).strip().lower(),"S1") for v in d[stc].values]
    else: d["STADE_ECL"]="S1"

    # TIE individuel = r_ZC(mat_res_i) + spread_i/10000
    spread_col=next((c for c in["SPREAD_EMISSION","Spread émis (bps)","SPREAD"] if c in d.columns),None)
    secteur_col=next((c for c in["Secteur","SECTEUR_GICS","SECTEUR"] if c in d.columns),None)
    def _tie(row):
        T=max(float(row["mat_res"]),0.1)
        zc=_r_zc(T)
        # BDT souverains: spread=0
        is_bdt=(secteur_col and str(row.get(secteur_col,"")).lower()=="souverain")
        if is_bdt: return float(np.clip(zc,0.001,0.30))
        sp=float(row[spread_col])/10000 if spread_col and pd.notna(row.get(spread_col)) else 0.0
        return float(np.clip(zc+sp,0.001,0.30))
    d["TIE"]=d.apply(_tie,axis=1)

    # rho + LGD PIT Frye-Jacobs
    d["rho"]=d["PD_TTC"].apply(_rho)
    DELTA={"Pessimiste":d_p,"Central":d_c,"Optimiste":d_o}
    W={"Pessimiste":w_p,"Central":w_c,"Optimiste":w_o}
    pp_c=d["PD_TTC"].apply(lambda p: float(1/(1+np.exp(-(_phi(float(p))+d_c))))).values
    Z=(np.array([_phi(float(p)) for p in d["PD_TTC"].values])
       -np.sqrt(1-d["rho"].values)*np.array([_phi(float(p)) for p in pp_c])
      )/np.sqrt(d["rho"].values)
    Z=np.where(np.isfinite(Z),Z,0.0)
    d["LGD_PIT"]=np.array([_lpit(float(d["PD_TTC"].iloc[i]),float(d["LGD_TTC"].iloc[i]),
                                   float(d["rho"].iloc[i]),Z[i]) for i in range(len(d))])
    ne=next((c for c in["Description complète","NOM_EMETTEUR","DESCRIPTION"] if c in d.columns),"")
    sc_col2=next((c for c in["Secteur","SECTEUR_GICS"] if c in d.columns),"")

    # ── ECL par obligation ────────────────────────────────────────────────
    rows=[]; 
    for ix in range(len(d)):
        row=d.iloc[ix]; pt=float(row["PD_TTC"]); lp=float(row["LGD_PIT"])
        lt=float(row["LGD_TTC"]); e0=float(row["EAD_0"]); mr=float(row["mat_res"])
        meth=str(row["METHODE_VALO"]); stade=str(row["STADE_ECL"])
        tie_i=float(row["TIE"])  # TIE INDIVIDUEL
        tmax=1 if stade=="S1" else int(_mg.ceil(mr))
        ecl12={}; eclL={}; det_c=[]
        for sc_n,delta in DELTA.items():
            pp=float(1/(1+np.exp(-(_phi(pt)+delta))))
            e12=0.0; eL=0.0; cpd_prev=0.0; ann=[]
            for t in range(1,tmax+1):
                pm=_pd_marg(pt,pp,t,T_p,T_r,cpd_prev)
                pa=_pd_ann(pt,pp,t,T_p,T_r)
                cpd_prev=1-(1-cpd_prev)*(1-pa)
                et=_ead_t(e0,meth,mr,pp,t-1)
                dt=1.0/(1.0+tie_i)**t   # D_i(t) = 1/(1+TIE_i)^t
                el=pm*lp*et*dt
                if t==1: e12=el
                eL+=el
                if sc_n=="Central":
                    ann.append({"T":t,"PD_ann(%)":round(pa*100,4),
                                "PD_marg(%)":round(pm*100,4),
                                "EAD_t":round(et,0),"LGD_PIT(%)":round(lp*100,4),
                                "TIE(%)":round(tie_i*100,4),
                                "D_t":round(dt,6),"EL_t":round(el,2),"ECL_cumul":round(eL,2)})
            ecl12[sc_n]=e12; eclL[sc_n]=eL
            if sc_n=="Central": det_c=ann
        e12p=sum(W[s]*ecl12[s] for s in W); eLp=sum(W[s]*eclL[s] for s in W)
        rows.append({"CODE_ISIN":str(row.get("CODE_ISIN",row.get("CODE",""))),
            "Description":str(row.get(ne,""))[:55] if ne else "",
            "Secteur":str(row.get(sc_col2,"")) if sc_col2 else "",
            "METHODE":meth,"STADE":stade,"mat_res":round(mr,2),
            "NOMINAL":round(float(row["NOMINAL"]),0),"EAD_0":round(e0,0),
            "r_ZC":round(_r_zc(mr)*100,4),"TIE":round(tie_i*100,4),
            "PD_TTC":round(pt,6),"LGD_TTC":round(lt,4),"LGD_PIT":round(lp,4),
            "EL_TTC":round(pt*lt*e0,2),
            "ECL_12M_Pess":round(ecl12.get("Pessimiste",0),2),
            "ECL_12M_Cent":round(ecl12.get("Central",0),2),
            "ECL_12M_Opti":round(ecl12.get("Optimiste",0),2),
            "ECL_12M_Pond":round(e12p,2),
            "ECL_Life_Pess":round(eclL.get("Pessimiste",0),2),
            "ECL_Life_Cent":round(eclL.get("Central",0),2),
            "ECL_Life_Opti":round(eclL.get("Optimiste",0),2),
            "ECL_Life_Pond":round(eLp,2),
            "ECL_FINAL":round(e12p if stade=="S1" else eLp,2),
            "_det_c":det_c})
    df_res=pd.DataFrame(rows)
    return df_res


# ════════════════════════════════════════════════════════════════
# TAB 0 — VUE D'ENSEMBLE
# ════════════════════════════════════════════════════════════════
with TABS[0]:
    st.subheader(" Vue d'ensemble — Portefeuille Obligataire IFRS 9")
    # KPIs ligne 1 : staging
    k0,k1,k2,k3,k4,k5 = st.columns(6)
    k0.metric("Obligations", TOT)
    k1.metric("🏢 Émetteurs",   df["NOM_EMETTEUR"].nunique())
    for km,s,lb in [(k2,"S1"," Bucket 1"),(k3,"S2"," Bucket 2"),(k4,"S3","Bucket 3 (S3 — Défaut probable)")]:
        n = dist.get(s,0); km.metric(lb, n, f"{n/TOT*100:.0f}%")
    ech_pos=(df["NB_ECHEANCES_MANQUEES"]>0).sum() if "NB_ECHEANCES_MANQUEES" in df.columns else 0
    k5.metric("Échéances manquées", f"{ech_pos}")
    # KPIs ligne 2 : PD + LGD si disponibles
    has_pd_pit  = "PD_PIT_NEUTRE" in df.columns or "PD_PIT_CPV" in df.columns
    has_lgd_pit = "LGD_PIT_V_Pond" in df.columns or "LGD_PIT_FJ_Pond" in df.columns
    pd_col_home = "PD_TTC" if "PD_TTC" in df.columns else ("PD_decimal" if "PD_decimal" in df.columns else None)
    lgd_col_home= "LGD_TTC" if "LGD_TTC" in df.columns else ("LGD" if "LGD" in df.columns else None)
    if has_pd_pit or has_lgd_pit:
        st.divider()
        col_a,col_b,col_c,col_d=st.columns(4)
        if pd_col_home:
            col_a.metric("PD_TTC moy",f"{df[pd_col_home].mean()*100:.3f}%")
        if has_pd_pit:
            pd_pit_col="PD_PIT_NEUTRE" if "PD_PIT_NEUTRE" in df.columns else "PD_PIT_CPV"
            col_b.metric("PD_PIT moy",f"{df[pd_pit_col].mean()*100:.3f}%",
                delta=f"{(df[pd_pit_col].mean()-df[pd_col_home].mean())*100:+.3f}pp" if pd_col_home else None)
        if lgd_col_home:
            col_c.metric("💸 LGD_TTC moy",f"{pd.to_numeric(df[lgd_col_home],errors='coerce').mean()*100:.2f}%")
        if has_lgd_pit:
            lgd_pit_col="LGD_PIT_V_Pond" if "LGD_PIT_V_Pond" in df.columns else "LGD_PIT_FJ_Pond"
            lgd_ttc_v=pd.to_numeric(df.get(lgd_col_home,pd.Series([0.45]*len(df))),errors="coerce").mean()
            col_d.metric("💸 LGD_PIT moy",f"{df[lgd_pit_col].mean()*100:.2f}%",
                delta=f"{(df[lgd_pit_col].mean()-lgd_ttc_v)*100:+.2f}pp")

    st.divider()
    cl,cr = st.columns([1.1,1])
    with cl:
        fig = go.Figure(go.Pie(
            labels=["S1 — Sain","S2 — Surveillance","S3 — Défaut probable"],
            values=[dist.get("S1",0),dist.get("S2",0),dist.get("S3",0)],
            hole=.55,
            marker=dict(colors=[C["S1"],C["S2"],C["S3"]],line=dict(color="white",width=3)),
            textinfo="label+percent+value",textfont_size=13))
        fig.update_layout(title_text="Distribution des Stades — Règle du MAX",
                          height=400,paper_bgcolor="white",
                          legend=dict(orientation="h",y=-0.12))
        st.plotly_chart(fig, use_container_width=True, key="chart_001")
    with cr:
        crit_cols = ["C1_STADE","C2_STADE","C3_STADE","C5_STADE","SNUM"]
        crit_lbl  = ["C1 K-Means","C2 Macro","C3 Évol. DR","C4 Échéances","FINAL"]
        fig2 = go.Figure()
        for sn,sl,col in [(1,"S1",C["S1"]),(2,"S2",C["S2"]),(3,"S3",C["S3"])]:
            vals = [df[cc].value_counts().get(sn,0)/TOT*100 for cc in crit_cols]
            fig2.add_trace(go.Bar(name=sl,x=crit_lbl,y=vals,marker_color=col,
                                   text=[f"{v:.0f}%" for v in vals],textposition="inside",
                                   textfont_size=11))
        fig2.update_layout(barmode="stack",title_text="Répartition par Critère",
                           yaxis_title="%",height=400,paper_bgcolor="white",
                           plot_bgcolor="white",legend=dict(orientation="h",y=-0.15))
        st.plotly_chart(fig2, use_container_width=True, key="chart_002")

    if "SECTEUR_GICS" in df.columns:
        pv = df.pivot_table(index="SECTEUR_GICS",columns="STADE",aggfunc="size",fill_value=0)
        for s in ["S1","S2","S3"]:
            if s not in pv.columns: pv[s]=0
        pv_pct = pv[["S1","S2","S3"]].div(pv[["S1","S2","S3"]].sum(axis=1),axis=0)*100
        fig3 = px.imshow(pv_pct.round(1),color_continuous_scale=["#1A6B2E","#FFF3CD","#A71E1E"],
                         text_auto=True,aspect="auto",title="Stades par Secteur GICS (%)")
        fig3.update_layout(height=300,paper_bgcolor="white")
        st.plotly_chart(fig3, use_container_width=True, key="chart_003")

    st.info(f"Seuils calibrés : S1/S2 = **{S12} bps** (Q75 neutres) · "
            f"S2/S3 = **{S23} bps** (médiane adverses)")


# ════════════════════════════════════════════════════════════════
# TAB 1 — ANALYSE MACRO & SPREAD AJUSTÉ
# ════════════════════════════════════════════════════════════════
with TABS[tab_staging]:
    if df is None:
        st.warning(" Chargez vos données dans la sidebar."); st.stop()

    sub_stag = st.tabs([
        "Macro & Spread Ajusté",
        "K-Means & Seuils",
        " Évolution Émetteurs",
        "Monte Carlo",
        "Échéances Manquées",
    ])

    with sub_stag[0]:
            st.subheader("Analyse Macro — ACP + AR(1)/MA · Spread Émis vs Spread Ajusté")

            # ── 1A : Indice macro + projection ──────────────────────────────
            st.markdown("### Indice Macro Synthétique (ACP)")
            cl,cr = st.columns([3,1])
            with cl:
                ans  = df_mac["annee"].values.astype(int)
                Zv   = df_mac["PC1z"].values
                regs = df_mac["REGIME"].values
                fig  = go.Figure()
                CZ = {"Adverse":"rgba(167,30,30,.13)","Neutre":"rgba(133,100,4,.10)",
                      "Favorable":"rgba(26,107,46,.10)"}
                for i in range(len(ans)-1):
                    fig.add_shape(type="rect",x0=ans[i],x1=ans[i+1],
                        y0=float(Zv.min())-.5,y1=float(Zv.max())+.5,
                        fillcolor=CZ.get(regs[i],"rgba(0,0,0,0)"),line_width=0,layer="below")
                fig.add_trace(go.Scatter(x=ans,y=Zv,mode="lines+markers",
                    line=dict(color=C["navy"],width=2.5),
                    marker=dict(size=9,color=[C["S3"] if r=="Adverse" else
                                (C["S2"] if r=="Neutre" else C["S1"]) for r in regs]),
                    name="Z-score macro (ACP)"))
                fig.add_hline(y=zadv,line=dict(color="red",dash="dash",width=2),
                              annotation_text=f"Adverse ({zadv})",annotation_font_color="red")
                fig.add_hline(y=zfav,line=dict(color="green",dash="dash",width=2),
                              annotation_text=f"Favorable ({zfav})",annotation_font_color="green")
                fig.add_hline(y=0,line=dict(color="gray",dash="dot",width=1))
                for zv2,lbl,col,sym in [(Z25ar,f"AR(1) ({Z25ar:+.2f})","red","diamond"),
                                          (Z25ma,f"MA({maw}) ({Z25ma:+.2f})","#00A3E0","square"),
                                          (Z25,f"Combiné ({Z25:+.2f})","purple","star")]:
                    fig.add_trace(go.Scatter(x=[ans[-1],ans[-1]+1],y=[Zv[-1],zv2],
                        mode="lines+markers",line=dict(color=col,width=2,dash="dash"),
                        marker=dict(size=11,symbol=sym,color=col),name=lbl))
                fig.add_shape(type="line",x0=ans[-1]+1,x1=ans[-1]+1,
                              y0=Z25lo,y1=Z25hi,line=dict(color="red",width=4))
                fig.update_layout(title="Indice Macro (ACP) + Projections AR(1)/MA 2025",
                                  xaxis_title="Année",yaxis_title="Z-score",height=440,
                                  paper_bgcolor="white",plot_bgcolor="white",
                                  legend=dict(orientation="h",y=-0.25))
                st.plotly_chart(fig, use_container_width=True, key="chart_004")
            with cr:
                st.markdown("#### Projections 2025")
                for lbl,zv2 in [("AR(1)",Z25ar),(f"MA({maw})",Z25ma),("Combiné",Z25)]:
                    reg = "Adverse" if zv2>zadv else ("Favorable" if zv2<zfav else "Neutre")
                    bg  = {"Adverse":"#F8D7DA","Neutre":"#FFF3CD","Favorable":"#D6EAD7"}[reg]
                    st.markdown(f"""<div style="background:{bg};padding:10px;border-radius:8px;margin:5px 0">
        <b>{lbl}</b><br>Z = {zv2:+.4f}<br>Régime → <b>{reg}</b></div>""", unsafe_allow_html=True)
                st.divider()
                st.markdown(f"φ₀={ph0:+.4f} | φ₁={ph1:.4f}\nσ={sig_ar:.4f}")
                st.markdown(f"IC 90% : [{Z25lo:+.3f}, {Z25hi:+.3f}]")
                stat = " Stationnaire" if abs(ph1)<1 else " Non-stationnaire"
                st.caption(stat)

            # ── 1B : Spread émis vs Spread ajusté — CÔTE À CÔTE ─────────────
            st.divider()
            st.markdown("### Spread d'Émission vs Spread Ajusté Macro — Comparaison directe")

            st.markdown(f"""
        **Formule :** Spread_ajusté = Spread_émis × (1 + ΔStress / Stress_T0)
        **ΔStress** = max(0, Stress_2025 − Stress_T0) · Si ΔStress < 0 → pas d'ajustement (plancher prudentiel)
        **Stress_2025 projeté** = **{st25:.4f}** (via AR(1)+MA sur la PC1)
        """)

            col_a, col_b = st.columns(2)

            with col_a:
                # Scatter : spread émis vs spread ajusté, colorié par stade
                fig_sc = go.Figure()
                for stade, grp in df.groupby("STADE"):
                    fig_sc.add_trace(go.Scatter(
                        x=grp[SC], y=grp["SPREAD_ADJ"],
                        mode="markers",
                        marker=dict(size=8, color=C[stade], opacity=.7,
                                    line=dict(color="white",width=.8)),
                        name=stade,
                        hovertemplate=(
                            "<b>%{customdata[0]}</b><br>"
                            "Spread émis : %{x:.1f} bps<br>"
                            "Spread ajusté : %{y:.1f} bps<br>"
                            "ΔStress : %{customdata[1]:.4f}<extra></extra>"),
                        customdata=grp[["NOM_EMETTEUR","DELTA_STRESS"]].values
                    ))
                lmax = float(max(df[SC].max(), df["SPREAD_ADJ"].max())) + 15
                fig_sc.add_shape(type="line",x0=0,y0=0,x1=lmax,y1=lmax,
                                 line=dict(color="gray",dash="dash",width=1.5))
                fig_sc.add_annotation(x=lmax*.6,y=lmax*.6,text="Pas d'ajustement\n(ΔStress=0)",
                                       showarrow=False,font=dict(color="gray",size=10))
                fig_sc.add_hline(y=S12,line=dict(color=C["S2"],width=2,dash="dot"),
                                 annotation_text=f"S1/S2={S12}")
                fig_sc.add_hline(y=S23,line=dict(color=C["S3"],width=2,dash="dot"),
                                 annotation_text=f"S2/S3={S23}")
                fig_sc.add_vline(x=S12,line=dict(color=C["S2"],width=1.5,dash="dot"))
                fig_sc.add_vline(x=S23,line=dict(color=C["S3"],width=1.5,dash="dot"))
                # Zone d'impact : points au-dessus de la diagonale = spread a augmenté
                fig_sc.add_annotation(x=20,y=S23+20,
                                       text="▲ Spread ajusté > Spread émis<br>(conditions dégradées)",
                                       showarrow=False,font=dict(color=C["S3"],size=9))
                fig_sc.update_layout(
                    title="Spread Émis vs Spread Ajusté Macro<br>(chaque point = une obligation)",
                    xaxis_title="Spread d'émission original (bps)",
                    yaxis_title="Spread ajusté par stress macro (bps)",
                    height=460,paper_bgcolor="white",plot_bgcolor="white",
                    legend=dict(orientation="h",y=-0.15))
                st.plotly_chart(fig_sc, use_container_width=True, key="chart_005")

            with col_b:
                # Barres groupées : spread émis vs ajusté par année d'émission
                df_yr = df.groupby("ANNEE_EMISSION")[[SC,"SPREAD_ADJ"]].mean().reset_index()
                df_yr = df_yr.dropna()

                fig_yr = go.Figure()
                fig_yr.add_trace(go.Bar(
                    x=df_yr["ANNEE_EMISSION"].astype(str),
                    y=df_yr[SC],
                    name="Spread émis moyen",
                    marker_color=C["blue"],opacity=.8))
                fig_yr.add_trace(go.Bar(
                    x=df_yr["ANNEE_EMISSION"].astype(str),
                    y=df_yr["SPREAD_ADJ"],
                    name="Spread ajusté moyen",
                    marker_color=C["S3"],opacity=.8))
                fig_yr.add_hline(y=S12,line=dict(color=C["S2"],width=2,dash="dash"),
                                 annotation_text=f"S1/S2={S12}")
                fig_yr.add_hline(y=S23,line=dict(color=C["S3"],width=2,dash="dash"),
                                 annotation_text=f"S2/S3={S23}")
                fig_yr.update_layout(
                    barmode="group",
                    title="Spread Moyen par Année d'Émission<br>Émis (bleu) vs Ajusté macro (rouge)",
                    xaxis_title="Année d'émission",yaxis_title="Spread moyen (bps)",
                    height=460,paper_bgcolor="white",plot_bgcolor="white",
                    legend=dict(orientation="h",y=-0.15))
                st.plotly_chart(fig_yr, use_container_width=True, key="chart_006")

            # Distribution ΔStress par régime d'émission
            st.markdown("#### ΔStress par Régime d'Émission — Impact de l'ajustement macro")
            col_c, col_d = st.columns(2)
            with col_c:
                fig_ds = px.box(df, x="REGIME_EM", y="DELTA_STRESS", color="REGIME_EM",
                                color_discrete_map={"Favorable":C["S1"],"Neutre":C["S2"],"Adverse":C["S3"]},
                                points="all", title="Distribution ΔStress par Régime d'Émission",
                                labels={"REGIME_EM":"Régime à l'émission","DELTA_STRESS":"ΔStress"},
                                category_orders={"REGIME_EM":["Favorable","Neutre","Adverse"]})
                fig_ds.update_layout(height=360,paper_bgcolor="white",plot_bgcolor="white",showlegend=False)
                st.plotly_chart(fig_ds, use_container_width=True, key="chart_007")
            with col_d:
                # Tableau résumé impact
                df_impact = df.groupby("REGIME_EM").agg(
                    N=("SPREAD_EMISSION","count"),
                    Spread_moy=(SC,"mean"),
                    SpreadAdj_moy=("SPREAD_ADJ","mean"),
                    DeltaStress_moy=("DELTA_STRESS","mean"),
                ).round(2)
                df_impact["Impact_%"] = ((df_impact["SpreadAdj_moy"]/df_impact["Spread_moy"]-1)*100).round(1)
                st.markdown("**Impact de l'ajustement macro sur le spread moyen**")
                st.dataframe(df_impact, use_container_width=True)
                st.caption("""
                **Lecture :** Les obligations émises en période Favorable (faible stress) voient leur spread
                ajusté augmenter davantage — car le stress actuel est plus élevé que lors de leur émission.
                Les obligations émises en 2020 (Adverse, stress=1.0) ne sont pas pénalisées.
                """)

            # Chargements ACP
            st.divider()
            cl2, cr2 = st.columns(2)
            with cl2:
                ld = pd.DataFrame({"Var":mvs,"Load":pca_m.components_[0]}).sort_values("Load")
                fig_l = px.bar(ld,x="Load",y="Var",orientation="h",color="Load",
                               color_continuous_scale=["#1A6B2E","white","#A71E1E"],
                               color_continuous_midpoint=0,text="Load",title="Chargements PC1")
                fig_l.update_traces(texttemplate="%{text:.4f}",textposition="outside")
                fig_l.update_layout(height=320,paper_bgcolor="white",plot_bgcolor="white",
                                    showlegend=False,coloraxis_showscale=False)
                st.plotly_chart(fig_l, use_container_width=True, key="chart_008")
            with cr2:
                fig_st = px.bar(df_mac,x="annee",y="stress",color="REGIME",
                                color_discrete_map={"Favorable":C["S1"],"Neutre":C["S2"],"Adverse":C["S3"]},
                                title="Stress Normalisé [0,1] par Année")
                fig_st.update_layout(height=320,paper_bgcolor="white",plot_bgcolor="white")
                st.plotly_chart(fig_st, use_container_width=True, key="chart_009")


        # ════════════════════════════════════════════════════════════════
        # TAB 2 — K-MEANS
        # ════════════════════════════════════════════════════════════════

    with sub_stag[1]:
            st.subheader(f"K-Means — {K} clusters · Seuils calibrés par crises historiques")
            c1,c2 = st.columns(2)
            with c1:
                fig = px.bar(x=list(sil_d.keys()),y=list(sil_d.values()),
                             text=[f"{v:.3f}" for v in sil_d.values()],
                             color=list(sil_d.values()),color_continuous_scale="Viridis",
                             title="Score Silhouette par k",labels={"x":"k","y":"Silhouette"})
                fig.add_vline(x=K,line=dict(color="red",dash="dash",width=2.5),annotation_text=f"k={K}")
                fig.update_traces(textposition="outside")
                fig.update_layout(height=320,paper_bgcolor="white",plot_bgcolor="white",
                                  showlegend=False,coloraxis_showscale=False)
                st.plotly_chart(fig, use_container_width=True, key="chart_010")
            with c2:
                dfi = df_int.copy()
                dfi["Stade"]  = dfi["c"].apply(lambda c: f"S{c2s(c)}")
                dfi["Justif"] = dfi["ct"].apply(
                    lambda ct: f"<{S12}" if ct<S12 else (f"[{S12},{S23}]" if ct<S23 else f"≥{S23}"))
                st.dataframe(dfi.rename(columns={"c":"C","niv":"Niveau","mn":"Min","mx":"Max",
                                                  "ct":"Centre","n":"N"}
                                        )[["C","Niveau","Min","Max","Centre","N","Stade","Justif"]],
                             use_container_width=True,hide_index=True)

            CPAL = px.colors.qualitative.Set2[:K]
            fig2 = go.Figure()
            for c in range(K):
                sub = df[df["CLU"]==c][SC].dropna()
                r   = df_int[df_int["c"]==c].iloc[0]
                fig2.add_trace(go.Box(y=sub,name=f"C{c+1}—{r['niv']}<br>[{r['mn']},{r['mx']}]bps",
                                       marker_color=CPAL[c],boxpoints="all",jitter=.45,
                                       pointpos=0,marker=dict(size=5,opacity=.55)))
            fig2.add_hline(y=S12,line=dict(color=C["S2"],width=3,dash="dash"),
                           annotation_text=f"S1/S2={S12} bps")
            fig2.add_hline(y=S23,line=dict(color=C["S3"],width=3,dash="dash"),
                           annotation_text=f"S2/S3={S23} bps")
            fig2.add_hrect(y0=0,y1=S12,fillcolor="rgba(26,107,46,.05)",layer="below",line_width=0)
            fig2.add_hrect(y0=S12,y1=S23,fillcolor="rgba(133,100,4,.05)",layer="below",line_width=0)
            fig2.add_hrect(y0=S23,y1=float(df[SC].max())+50,fillcolor="rgba(167,30,30,.05)",layer="below",line_width=0)
            fig2.update_layout(title=f"K-Means (k={K}) — Calibrés par crises historiques",
                               yaxis_title="Spread (bps)",height=480,paper_bgcolor="white",
                               plot_bgcolor="white",legend=dict(orientation="h",y=-0.25))
            st.plotly_chart(fig2, use_container_width=True, key="chart_011")


        # ════════════════════════════════════════════════════════════════
        # TAB 3 — ÉVOLUTION ÉMETTEURS
        # ════════════════════════════════════════════════════════════════

    with sub_stag[2]:
            st.subheader("Évolution du Spread par Émetteur — Régression Linéaire")

            df_em = df_evol[df_evol["n"].astype(int)>=2].copy()
            df_em["delta"] = pd.to_numeric(df_em["delta"],errors="coerce")
            df_em["pente"] = pd.to_numeric(df_em["pente"],errors="coerce")
            df_em = df_em.dropna(subset=["delta"]).sort_values("delta")

            if len(df_em)==0:
                st.warning("Aucun émetteur avec plusieurs émissions détecté.")
            else:
                c1,c2 = st.columns(2)
                with c1:
                    cb = [C["S3"] if d>=50 else (C["S2"] if d>=0 else C["S1"]) for d in df_em["delta"]]
                    fig = go.Figure(go.Bar(
                        x=df_em["delta"].values,y=[str(e)[:26] for e in df_em.index],
                        orientation="h",marker=dict(color=cb,line=dict(color="white",width=.5)),
                        text=[f"{d:+.1f}" for d in df_em["delta"]],
                        textposition="outside",textfont_size=9))
                    fig.add_vline(x=0,line=dict(color="black",width=2))
                    fig.add_vline(x=50,line=dict(color=C["S3"],width=2,dash="dash"),annotation_text="S3")
                    fig.update_layout(title="ΔSpread par Émetteur",xaxis_title="ΔSpread (bps)",
                                      height=max(380,len(df_em)*22+80),
                                      paper_bgcolor="white",plot_bgcolor="white",showlegend=False)
                    st.plotly_chart(fig, use_container_width=True, key="chart_012")
                with c2:
                    df_sig = df_em[df_em["pval"]<.10].sort_values("pente")
                    if len(df_sig)>0:
                        cp = [C["S3"] if p>0 else C["S1"] for p in df_sig["pente"]]
                        fig2 = go.Figure(go.Bar(
                            x=df_sig["pente"].values,y=[str(e)[:26] for e in df_sig.index],
                            orientation="h",marker=dict(color=cp,line=dict(color="white",width=.5)),
                            text=[f"p={p:.3f}" for p in df_sig["pval"]],
                            textposition="outside",textfont_size=9))
                        fig2.add_vline(x=0,line=dict(color="black",width=2))
                        fig2.update_layout(title="Tendance (bps/an) — p<10%",
                                           xaxis_title="Pente",height=max(380,len(df_sig)*26+80),
                                           paper_bgcolor="white",plot_bgcolor="white",showlegend=False)
                        st.plotly_chart(fig2, use_container_width=True, key="chart_013")
                    else:
                        st.info("Aucune tendance significative.")

                st.divider()
                st.subheader(" Zoom — Chronologie d'un Émetteur")
                em_opts = sorted(df_evol[df_evol["n"].astype(int)>=2].index.tolist())
                em_sel  = st.selectbox("Sélectionner :", em_opts,
                    format_func=lambda e: f"{e}  |  n={int(df_evol.loc[e,'n'])}  |  "
                                           f"ΔSpread={df_evol.loc[e,'delta']:+.1f}bps  |  {df_evol.loc[e,'tend']}", key="wk_se_0003")

                sub_em = df[df["NOM_EMETTEUR"]==em_sel].dropna(
                    subset=[SC,"DATE_EMISSION"]).sort_values("DATE_EMISSION")

                if len(sub_em)>=1:
                    fig3 = go.Figure()
                    ann_cols = [C["S3"] if y>=2020 else (C["S2"] if y>=2015 else C["S1"])
                                for y in sub_em["ANNEE_EMISSION"].values]
                    fig3.add_trace(go.Scatter(
                        x=sub_em["DATE_EMISSION"],y=sub_em[SC],
                        mode="markers+lines+text",
                        marker=dict(size=13,color=ann_cols,line=dict(color="white",width=1.5)),
                        line=dict(color=C["navy"],width=2),
                        text=[f"{v:.1f}" for v in sub_em[SC]],
                        textposition="top center",textfont_size=10,
                        name="Spread émis (bps)"))

                    if len(sub_em)>=2:
                        xn = (sub_em["DATE_EMISSION"]-sub_em["DATE_EMISSION"].min()).dt.days.values.astype(float)
                        yn = sub_em[SC].values
                        if xn.std()>0:
                            sl,ic,_,pv,_ = stats.linregress(xn,yn)
                            yreg = ic+sl*xn; col_r = C["S3"] if sl>0 else C["S1"]
                            fig3.add_trace(go.Scatter(x=sub_em["DATE_EMISSION"],y=yreg,mode="lines",
                                line=dict(color=col_r,width=2.5,dash="dash"),
                                name=f"Tendance : {sl*365.25:+.1f}bps/an (p={pv:.3f})"))
                            x_ex = float((DC-sub_em["DATE_EMISSION"].min()).days)
                            y_ex = ic+sl*x_ex
                            fig3.add_trace(go.Scatter(
                                x=[sub_em["DATE_EMISSION"].iloc[-1],DC],y=[yreg[-1],y_ex],
                                mode="lines+markers",line=dict(color=col_r,width=1.5,dash="dot"),
                                marker=dict(size=11,symbol="diamond",color=col_r),
                                name=f"Projection {date_calc} : {y_ex:.1f}bps"))

                    fig3.add_hline(y=S12,line=dict(color=C["S2"],width=2,dash="dot"),annotation_text=f"S1/S2={S12}")
                    fig3.add_hline(y=S23,line=dict(color=C["S3"],width=2,dash="dot"),annotation_text=f"S2/S3={S23}")
                    fig3.add_hrect(y0=0,y1=S12,fillcolor="rgba(26,107,46,.05)",layer="below",line_width=0)
                    fig3.add_hrect(y0=S12,y1=S23,fillcolor="rgba(133,100,4,.05)",layer="below",line_width=0)
                    fig3.add_hrect(y0=S23,y1=float(sub_em[SC].max())+30,fillcolor="rgba(167,30,30,.05)",layer="below",line_width=0)
                    r_em = df_evol.loc[em_sel]
                    fig3.update_layout(title=f"<b>{em_sel}</b> · ΔSpread={r_em['delta']:+.1f}bps · S{int(r_em['stade'])} · {r_em['tend']}",
                                       xaxis_title="Date d'émission",yaxis_title="Spread (bps)",
                                       height=450,paper_bgcolor="white",plot_bgcolor="white",
                                       legend=dict(orientation="h",y=-0.2))
                    st.plotly_chart(fig3, use_container_width=True, key="chart_014")

                    ma1,ma2,ma3,ma4,ma5 = st.columns(5)
                    ma1.metric("N émissions",int(r_em["n"]))
                    ma2.metric("Spread initial",f"{r_em['s0']:.1f} bps")
                    ma3.metric("Spread récent",f"{r_em['sT']:.1f} bps")
                    ma4.metric("ΔSpread",f"{r_em['delta']:+.1f} bps",delta=f"S{int(r_em['stade'])}")
                    ma5.metric("Pente",f"{r_em['pente']:+.1f} bps/an",delta=r_em["tend"])

                st.divider()
                st.subheader("🗺️ Carte Globale — Tous les Titres")
                fig4 = px.scatter(df.dropna(subset=["DATE_EMISSION",SC]),
                                  x="DATE_EMISSION",y=SC,color="STADE",size=SC,size_max=22,
                                  hover_data=["NOM_EMETTEUR","SECTEUR_GICS","C1_STADE","C3_STADE","DEC"],
                                  color_discrete_map={"S1":C["S1"],"S2":C["S2"],"S3":C["S3"]},
                                  title="Spreads — Tous Titres · colorié par Stade Final",opacity=.75)
                fig4.add_hline(y=S12,line=dict(color=C["S2"],width=2,dash="dash"),annotation_text=f"S1/S2={S12}")
                fig4.add_hline(y=S23,line=dict(color=C["S3"],width=2,dash="dash"),annotation_text=f"S2/S3={S23}")
                fig4.update_layout(height=440,paper_bgcolor="white",plot_bgcolor="white",
                                   legend=dict(orientation="h",y=-0.15))
                st.plotly_chart(fig4, use_container_width=True, key="chart_015")


        # ════════════════════════════════════════════════════════════════
        # TAB 4 — MONTE CARLO
        # ════════════════════════════════════════════════════════════════

    with sub_stag[3]:
            st.subheader(f"Monte Carlo — Fiabilité du Staging ({nsim} simulations/titre)")

            st.markdown("""
        **Modèle :** `spread_sim = spread_obs × (1 + β_i × ΔStress_i)`
        **β_i ~ N(0.35, 0.15)** (incertitude sur la sensibilité) &nbsp;·&nbsp;
        **ε_i ~ N(0, 15%)** (incertitude sur la projection macro)

        **Lecture des résultats :**
        - **P(S3) > 50%** → classement S3 robuste
        - **Fiabilité > 80%** → stade attribué avec haute confiance
        - **Points au-dessus de la diagonale** → stress macro amplifie le risque
        """)

            c1,c2 = st.columns(2)
            with c1:
                fig = px.histogram(df,x="PS3",nbins=35,color_discrete_sequence=[C["S3"]],
                                   title="Distribution de P(S3) — par obligation",
                                   labels={"PS3":"P(Stade 3) (%)"})
                fig.add_vline(x=50,line=dict(color="black",width=2,dash="dash"),
                              annotation_text="Décision S3 (50%)")
                fig.update_layout(height=360,paper_bgcolor="white",plot_bgcolor="white")
                st.plotly_chart(fig, use_container_width=True, key="chart_016")
            with c2:
                fig2 = px.box(df,x="STADE",y="FIAB",color="STADE",points="all",
                              color_discrete_map={"S1":C["S1"],"S2":C["S2"],"S3":C["S3"]},
                              title="Fiabilité MC par Stade (%)",
                              category_orders={"STADE":["S1","S2","S3"]})
                fig2.update_layout(height=360,paper_bgcolor="white",plot_bgcolor="white",showlegend=False)
                st.plotly_chart(fig2, use_container_width=True, key="chart_017")

            # Spread émis vs MC Q50/Q75/Q95 — visualisation enrichie
            st.markdown("### Distribution des Spreads Simulés vs Spread Émis")

            fig3 = go.Figure()
            # Bande Q50-Q95
            df_sorted = df.sort_values(SC).reset_index(drop=True)
            x_idx = list(range(len(df_sorted)))
            fig3.add_trace(go.Scatter(
                x=x_idx+x_idx[::-1],
                y=list(df_sorted["MCQ95"])+list(df_sorted["MCQ50"])[::-1],
                fill="toself",fillcolor="rgba(0,163,224,.15)",line=dict(color="rgba(0,163,224,0)"),
                name="Bande Q50–Q95 MC",showlegend=True))
            fig3.add_trace(go.Scatter(
                x=x_idx,y=df_sorted[SC],mode="lines",
                line=dict(color=C["navy"],width=2),name="Spread émis (trié)"))
            fig3.add_trace(go.Scatter(
                x=x_idx,y=df_sorted["MCQ75"],mode="lines",
                line=dict(color=C["S2"],width=2,dash="dash"),name="MC Q75"))
            fig3.add_trace(go.Scatter(
                x=x_idx,y=df_sorted["MCQ95"],mode="lines",
                line=dict(color=C["S3"],width=1.5,dash="dot"),name="MC Q95"))
            fig3.add_hline(y=S12,line=dict(color=C["S2"],width=2,dash="dash"),annotation_text=f"S1/S2={S12}")
            fig3.add_hline(y=S23,line=dict(color=C["S3"],width=2,dash="dash"),annotation_text=f"S2/S3={S23}")
            fig3.update_layout(
                title="Enveloppe Monte Carlo des Spreads (obligations triées par spread croissant)",
                xaxis_title="Obligations (triées par spread croissant)",
                yaxis_title="Spread (bps)",height=440,
                paper_bgcolor="white",plot_bgcolor="white",
                legend=dict(orientation="h",y=-0.2))
            st.plotly_chart(fig3, use_container_width=True, key="chart_018")

            # Scatter avec IC
            fig4 = go.Figure()
            for stade,grp in df.groupby("STADE"):
                fig4.add_trace(go.Scatter(
                    x=grp[SC],y=grp["MCQ75"],mode="markers",
                    marker=dict(size=8,color=C[stade],opacity=.7,line=dict(color="white",width=.8)),
                    name=stade,
                    error_y=dict(type="data",array=(grp["MCQ95"]-grp["MCQ75"]).values,
                                 arrayminus=(grp["MCQ75"]-grp["MCQ50"]).values,
                                 color=C[stade],thickness=1,width=3),
                    hovertemplate="<b>%{customdata}</b><br>Émis:%{x:.1f} · MC Q75:%{y:.1f}<extra></extra>",
                    customdata=grp["NOM_EMETTEUR"].values))
            lmax = float(max(df[SC].max(),df["MCQ95"].max()))+10
            fig4.add_shape(type="line",x0=0,y0=0,x1=lmax,y1=lmax,
                           line=dict(color="gray",dash="dash",width=1.5))
            fig4.add_hline(y=S12,line=dict(color=C["S2"],width=1.5,dash="dot"))
            fig4.add_hline(y=S23,line=dict(color=C["S3"],width=1.5,dash="dot"))
            fig4.update_layout(
                title="Spread Émis vs Spread MC Q75 (avec intervalle Q50–Q95)",
                xaxis_title="Spread émis (bps)",yaxis_title="Spread MC (bps)",
                height=440,paper_bgcolor="white",plot_bgcolor="white")
            st.plotly_chart(fig4, use_container_width=True, key="chart_019")

            # Zoom individuel
            st.divider()
            st.subheader("Zoom Monte Carlo — Titre individuel")
            tsel = st.selectbox("Obligation :",df.index.tolist(),
                format_func=lambda i:f"{df.loc[i,'NOM_EMETTEUR']} · {df.loc[i,SC]:.1f}bps · {df.loc[i,'STADE']}", key="wk_se_0002")
            rs   = df.loc[tsel]
            sims = mc_spread(rs[SC],rs["STRESS_T0"],st25,nsim)
            p1t  = (sims<S12).mean(); p2t=((sims>=S12)&(sims<S23)).mean(); p3t=(sims>=S23).mean()

            fig5 = go.Figure()
            fig5.add_trace(go.Histogram(x=sims,nbinsx=60,marker_color=C["blue"],opacity=.72,
                                         name=f"N={nsim} simulations"))
            fig5.add_vrect(x0=0,x1=S12,fillcolor="rgba(26,107,46,.07)",layer="below",line_width=0)
            fig5.add_vrect(x0=S12,x1=S23,fillcolor="rgba(133,100,4,.07)",layer="below",line_width=0)
            fig5.add_vrect(x0=S23,x1=float(sims.max())+10,fillcolor="rgba(167,30,30,.07)",layer="below",line_width=0)
            for q,col,lbl in [(.50,C["navy"],"Q50"),(.75,"orange","Q75"),(.95,C["S3"],"Q95")]:
                qv = float(np.quantile(sims,q))
                fig5.add_vline(x=qv,line=dict(color=col,width=2.5,dash="dash"),annotation_text=f"{lbl}={qv:.1f}")
            fig5.add_vline(x=rs[SC],line=dict(color="black",width=2),annotation_text=f"Émis={rs[SC]:.1f}")
            fig5.add_vline(x=S12,line=dict(color=C["S2"],width=2,dash="dot"))
            fig5.add_vline(x=S23,line=dict(color=C["S3"],width=2,dash="dot"))
            fig5.update_layout(
                title=f"{rs['NOM_EMETTEUR']} · P(S1)={p1t*100:.1f}% · P(S2)={p2t*100:.1f}% · P(S3)={p3t*100:.1f}%",
                xaxis_title="Spread simulé (bps)",height=420,
                paper_bgcolor="white",plot_bgcolor="white")
            st.plotly_chart(fig5, use_container_width=True, key="chart_020")
            ca,cb_,cc_ = st.columns(3)
            ca.metric("P(S1)",f"{p1t*100:.1f}%")
            cb_.metric("P(S2)",f"{p2t*100:.1f}%")
            cc_.metric("P(S3)",f"{p3t*100:.1f}%",
                       delta=" Risque élevé" if p3t>.30 else None,delta_color="inverse")


        # ════════════════════════════════════════════════════════════════
        # TAB 5 — JOURS DE RETARD
        # ════════════════════════════════════════════════════════════════

    with sub_stag[4]:
            st.subheader("Échéances Manquées — Saisie et Analyse")

            st.markdown(f"""
        **Référence IFRS 9 §B5.5.28 :** Le nombre d'échéances manquées constitue un
        indicateur présumé de dégradation significative du risque de crédit.

        | Échéances manquées | Stade | Interprétation |
        |-------------------|-------|---------------|
        | **0** | S1 | Aucune difficulté de paiement |
        | **1** | S2 | Signal de difficultés — surveillance renforcée |
        | **≥ 2** | S3 | Défaut probable — présomption réfutable |

        **Mode opératoire :** Modifiez le tableau ci-dessous. Les colonnes grisées sont calculées automatiquement.
        """)

            # ─ Tableau de saisie interactif ─────────────────────────────
            st.markdown("### Saisir les Échéances Manquées par Obligation")

            cols_ech = ["NOM_EMETTEUR","SECTEUR_GICS","SPREAD_EMISSION","DATE_EMISSION",
                        "NB_ECHEANCES_MANQUEES"]
            cols_ech = [c for c in cols_ech if c in df.columns]

            df_edit = df[cols_ech].copy()
            df_edit["NB_ECHEANCES_MANQUEES"] = df_edit["NB_ECHEANCES_MANQUEES"].astype(int)

            # Colonnes configurables dans le data_editor
            col_config = {
                "NOM_EMETTEUR":         st.column_config.TextColumn("Émetteur", disabled=True),
                "SECTEUR_GICS":         st.column_config.TextColumn("Secteur",  disabled=True),
                "SPREAD_EMISSION":      st.column_config.NumberColumn("Spread (bps)", disabled=True, format="%.1f"),
                "DATE_EMISSION":        st.column_config.DateColumn("Date émission", disabled=True),
                "NB_ECHEANCES_MANQUEES":st.column_config.NumberColumn(
                    "Échéances manquées",
                    min_value=0, max_value=24, step=1,
                    help="0=S1 · 1=S2 · ≥2=S3 (IFRS 9 §B5.5.28)"),
            }

            df_edited = st.data_editor(
                df_edit, column_config=col_config,
                use_container_width=True, num_rows="fixed",
                hide_index=True, height=420, key="editor_ech")

            # Bouton pour appliquer
            if st.button(" Appliquer les échéances au staging", type="primary", use_container_width=True):
                df["NB_ECHEANCES_MANQUEES"] = df_edited["NB_ECHEANCES_MANQUEES"].values
                df["C5_STADE"]  = df["NB_ECHEANCES_MANQUEES"].apply(stade_from_echeances)
                df["SNUM"]      = df[["C1_STADE","C2_STADE","C3_STADE","C4_STADE","C5_STADE"]].max(axis=1)
                df["STADE"]     = df["SNUM"].map({1:"S1",2:"S2",3:"S3"})
                df["DEC"]       = df.apply(dec, axis=1)
                dist            = df["STADE"].value_counts().sort_index()
                st.success(" Staging recalculé avec les échéances manquées.")
                st.rerun()

            st.divider()

            # ─ Analyse des échéances ────────────────────────────────────
            st.markdown("### Analyse des Échéances Manquées")

            df["NB_ECHEANCES_MANQUEES"] = df_edited["NB_ECHEANCES_MANQUEES"].values
            df["C5_STADE"] = df["NB_ECHEANCES_MANQUEES"].apply(stade_from_echeances)

            n_ech_0 = (df["NB_ECHEANCES_MANQUEES"]==0).sum()
            n_ech_1 = (df["NB_ECHEANCES_MANQUEES"]==1).sum()
            n_ech_2p= (df["NB_ECHEANCES_MANQUEES"]>=2).sum()

            k1,k2,k3,k4 = st.columns(4)
            k1.metric("0 échéance (S1)", n_ech_0)
            k2.metric("1 échéance (S2)", n_ech_1)
            k3.metric("≥2 échéances (S3)", n_ech_2p)
            k4.metric("Total affectés", n_ech_1+n_ech_2p)

            c1,c2 = st.columns(2)
            with c1:
                fig = go.Figure()
                cats = ["0 (S1)","1 (S2)","≥2 (S3)"]
                vals_e = [n_ech_0, n_ech_1, n_ech_2p]
                fig.add_trace(go.Bar(
                    x=cats, y=vals_e,
                    marker=dict(color=[C["S1"],C["S2"],C["S3"]],line=dict(color="white",width=1)),
                    text=[f"{v}\n({v/TOT*100:.0f}%)" for v in vals_e],
                    textposition="outside",textfont_size=12))
                fig.update_layout(title="Distribution des Échéances Manquées",
                                  xaxis_title="Nombre d'échéances manquées",
                                  yaxis_title="Nombre d'obligations",
                                  height=380,paper_bgcolor="white",plot_bgcolor="white",showlegend=False)
                st.plotly_chart(fig, use_container_width=True, key="chart_021")

            with c2:
                # Donut C5 uniquement
                fig2 = go.Figure(go.Pie(
                    labels=["S1 (0 éch.)","S2 (1 éch.)","S3 (≥2 éch.)"],
                    values=[n_ech_0,n_ech_1,n_ech_2p],
                    hole=.55,
                    marker=dict(colors=[C["S1"],C["S2"],C["S3"]],line=dict(color="white",width=3)),
                    textinfo="label+percent+value",textfont_size=12))
                fig2.update_layout(title="C5 — Stade Échéances Seul",
                                   height=380,paper_bgcolor="white",
                                   legend=dict(orientation="h",y=-0.1))
                st.plotly_chart(fig2, use_container_width=True, key="chart_022")

            # Heatmap : C5 vs stade final
            if n_ech_1+n_ech_2p > 0:
                st.markdown("#### Impact des Échéances sur le Stade Final")
                pv_e = df.pivot_table(index="C5_STADE",columns="STADE",aggfunc="size",fill_value=0)
                pv_e.index = [f"C5=S{i}" for i in pv_e.index]
                for s in ["S1","S2","S3"]:
                    if s not in pv_e.columns: pv_e[s]=0
                pv_e = pv_e[["S1","S2","S3"]]
                fig3 = px.imshow(pv_e,color_continuous_scale=["#1A6B2E","#FFF3CD","#A71E1E"],
                                 text_auto=True,aspect="auto",
                                 title="C5 Échéances vs Stade Final — Nombre d'obligations")
                fig3.update_layout(height=260,paper_bgcolor="white")
                st.plotly_chart(fig3, use_container_width=True, key="chart_023")

                # Titres avec échéances manquées
                df_ech = df[df["NB_ECHEANCES_MANQUEES"]>0].sort_values("NB_ECHEANCES_MANQUEES",ascending=False)
                if len(df_ech)>0:
                    st.markdown("#### Obligations avec Échéances Manquées")
                    cols_show = ["NOM_EMETTEUR","SECTEUR_GICS","SPREAD_EMISSION",
                                 "NB_ECHEANCES_MANQUEES","C5_STADE","SNUM","STADE","DEC"]
                    cols_show = [c for c in cols_show if c in df_ech.columns]

                    def hl_ech(row):
                        s = str(row.get("STADE",""))
                        bg = {"S1":"#D6EAD7","S2":"#FFF3CD","S3":"#F8D7DA"}.get(s,"white")
                        return [f"background-color:{bg}" if col=="STADE" else "" for col in row.index]

                    st.dataframe(df_ech[cols_show].style.apply(hl_ech,axis=1),
                                 use_container_width=True, hide_index=True)
            else:
                st.info("Aucune échéance manquée saisie pour l'instant. "
                        "Modifiez le tableau ci-dessus pour entrer les données.")



        # ════════════════════════════════════════════════════════════════
        # TAB PD — PD MERTON-VASICEK COMPLET
        # ════════════════════════════════════════════════════════════════


with TABS[tab_pd]:
    from scipy.stats import norm, jarque_bera, kstest, chi2
    from sklearn.decomposition import PCA

    st.markdown("""<div class='hdr'>
      <div class='hdr-txt'>
        <h2>Modele PD Merton-Vasicek — Forward Looking IFRS 9</h2>
        <p>Seuils de défaut · Calibration ρ (MOM) · Z_t systémique · ACP macro · PD_PIT par scénario</p>
      </div></div>""", unsafe_allow_html=True)

    if df is None:
        st.warning(" Veuillez charger vos données dans la barre latérale.")
        st.stop()

    # ── DONNÉES ───────────────────────────────────────────────────────────────
    PD_HIST_VALS = {
        1994:.0690,1995:.0,1996:.0714,1997:.0,1998:.0233,1999:.0152,
        2000:.0,2001:.0588,2002:.1408,2003:.0145,2004:.0476,2005:.0488,
        2006:.0233,2007:.0706,2008:.0204,2009:.0291,2010:.0367,2011:.0642,
        2012:.0273,2013:.0145,2014:.0192,2015:.0363,2016:.0305,2017:.0193,
        2018:.0143,2019:.0493,2020:.0865,2021:.0204,2022:.0068,2023:.0310,2024:.0199}
    ANNEES = np.array(sorted(PD_HIST_VALS.keys()),dtype=int)
    PD_GLOBAL = np.array([PD_HIST_VALS[a] for a in ANNEES])
    T_HIST = len(ANNEES)

    mvs = ["inflation","croissance_pib","chomage","taux_change",
           "masse_monetaire","taux_directeur"]

    df_w = df.copy()
    df_w["PD_TTC"] = pd.to_numeric(df_w.get("PD", df_w.get("PD_TTC",None)),errors="coerce")/100
    df_w["PD_TTC"] = df_w["PD_TTC"].clip(1e-6,1-1e-6)
    df_w["SECTEUR"] = df_w.get("SECTEUR_GICS",df_w.get("Secteur",df_w.get("SECTEUR","Autre")))

    # ── PARAMÈTRES LATÉRAUX ────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### Paramètres Vasicek")
        n_boot_pd = st.number_input("Bootstrap B",500,20000,10000,500,key="pd_n_boot")
        t_max_pd  = st.slider("Horizon max T",3,15,10,key="pd_t_max")
        w_cent    = st.slider("Poids Central %",50,90,70,5,key="pd_w_cent")
        w_pess    = st.slider("Poids Pessimiste %",5,40,20,5,key="pd_w_pess")
        w_opti    = 100-w_cent-w_pess
        st.info(f"Optimiste = {w_opti}%")

    # ══════════════════════════════════════════════════════════════════════════
    # CALCULS CORE
    # ══════════════════════════════════════════════════════════════════════════
    @st.cache_data(show_spinner=False)
    def compute_vasicek(pd_hist_vals, n_boot, t_max):
        annees = np.array(sorted(pd_hist_vals.keys()),dtype=int)
        pd_gl  = np.array([pd_hist_vals[a] for a in annees])
        t_hist = len(annees)
        eps    = 1.0/(t_hist+1)
        pd_adj = pd_gl.copy(); pd_adj[pd_adj==0]=eps
        y_t    = norm.ppf(pd_adj)

        lam0   = float(np.mean(y_t))
        lam1sq = float(np.mean((y_t-lam0)**2))
        lam1   = np.sqrt(lam1sq)
        rho    = lam1sq/(1+lam1sq)

        # Bootstrap IC
        np.random.seed(42)
        rb=[]; 
        for _ in range(n_boot):
            idx=np.random.choice(t_hist,t_hist,replace=True)
            yt_=y_t[idx]; l0_=yt_.mean(); l1sq_=np.mean((yt_-l0_)**2)
            rb.append(l1sq_/(1+l1sq_))
        rb=np.array(rb)
        rho_lo=float(np.percentile(rb,2.5)); rho_hi=float(np.percentile(rb,97.5))

        z_hist = (y_t-lam0)/(lam1+1e-10)
        regimes= np.where(z_hist>0.75,"Adverse",np.where(z_hist<-0.75,"Favorable","Neutre"))

        # Tests
        jb_s,jb_p = jarque_bera(y_t)
        ks_s,ks_p = kstest(y_t,"norm",args=(y_t.mean(),y_t.std()))

        return {"annees":annees,"pd_global":pd_gl,"pd_adj":pd_adj,"y_t":y_t,
                "lam0":lam0,"lam1":lam1,"rho":rho,"rho_lo":rho_lo,"rho_hi":rho_hi,
                "z_hist":z_hist,"regimes":regimes,"rb":rb,"jb_p":jb_p,"ks_p":ks_p}

    with st.spinner("Calibration du modèle Vasicek…"):
        V = compute_vasicek(PD_HIST_VALS, n_boot_pd, t_max_pd)

    RHO   = V["rho"]; SQRT_RHO=np.sqrt(RHO); SQRT_1RHO=np.sqrt(1-RHO)
    LAM0  = V["lam0"]; LAM1=V["lam1"]
    Z_HIST= V["z_hist"]; ANNEES=V["annees"]

    # Seuils α_i
    df_w["ALPHA_I"] = norm.ppf(df_w["PD_TTC"])

    # ── MACRO + ACP ───────────────────────────────────────────────────────────
    has_macro = df_mac is not None
    Z_RECON = None; r_acp = 0; best_k=1; ANNEES_C=None; Z_C=None

    if has_macro:
        @st.cache_data(show_spinner=False)
        def compute_acp_macro(mac_data_json, z_hist_arr, annees_all):
            df_m = pd.read_json(io.StringIO(mac_data_json))
            mvars_use = [c for c in ["inflation","croissance_pib","chomage",
                                      "taux_change","masse_monetaire","taux_directeur"]
                         if c in df_m.columns]
            if len(mvars_use)<2: return None,None,None,None,None,None,None,None
            ann_mac = df_m["annee"].values.astype(int) if "annee" in df_m.columns else                       df_m["Année"].values.astype(int)
            mask = np.isin(annees_all, ann_mac)
            ann_c= annees_all[mask]; z_c=z_hist_arr[mask]
            M_raw= df_m[mvars_use].fillna(df_m[mvars_use].mean()).values
            Ms   = StandardScaler().fit_transform(M_raw)
            pca  = PCA(n_components=min(3,len(mvars_use)),random_state=42)
            pca.fit(Ms); PC_ALL=pca.transform(Ms)
            PC1=PC_ALL[:,0]
            if len(PC1)==len(df_m) and "croissance_pib" in df_m.columns:
                if np.corrcoef(PC1,df_m["croissance_pib"].fillna(0).values)[0,1]>0:
                    PC1=-PC1
            # Régression Z_t ~ PC
            def aic_lr(X,y,k):
                n=len(y); Xc=np.column_stack([np.ones(n),X])
                b,*_=lstsq(Xc,y); yh=Xc@b; sse=np.sum((y-yh)**2)
                return n*np.log(sse/n+1e-12)+2*(k+1),b
            aic1,b1=aic_lr(PC1.reshape(-1,1),z_c,1)
            aic2,b2=(aic_lr(PC_ALL[:,:2],z_c,2) if PC_ALL.shape[1]>=2 else (1e9,b1))
            aic3,b3=(aic_lr(PC_ALL[:,:3],z_c,3) if PC_ALL.shape[1]>=3 else (1e9,b1))
            bk=np.argmin([aic1,aic2,aic3])+1
            B_acp=(b1 if bk==1 else b2 if bk==2 else b3)
            X_acp=(PC1.reshape(-1,1) if bk==1 else PC_ALL[:,:bk])
            Z_rc=np.column_stack([np.ones(len(X_acp)),X_acp])@B_acp
            from scipy.stats import pearsonr
            r,_=pearsonr(Z_rc,z_c)
            # AR(1) sur PC1
            xl=PC1[:-1].reshape(-1,1); xc=PC1[1:]
            ph,*_=lstsq(np.column_stack([np.ones(len(xl)),xl]),xc)
            ph0,ph1=float(ph[0]),float(ph[1])
            sig_ar=float((xc-(ph0+ph1*xl.flatten())).std())
            return ann_c,z_c,Z_rc,r,bk,B_acp,PC1,{"ph0":ph0,"ph1":ph1,"sig":sig_ar,
                    "ev":pca.explained_variance_ratio_.tolist(),
                    "mvars":mvars_use,"ann_mac":ann_mac.tolist()}

        mac_json = df_mac.to_json()
        result_acp = compute_acp_macro(mac_json, Z_HIST, ANNEES)
        if result_acp[0] is not None:
            ANNEES_C,Z_C,Z_RECON,r_acp,best_k,B_ACP,PC1_MACRO,AR1_INFO = result_acp
            has_macro = True

    # ── SCÉNARIOS Z_2025 ──────────────────────────────────────────────────────
    def pd_pit_model(alpha, Z, rho=RHO):
        return float(norm.cdf((alpha - np.sqrt(rho)*Z)/np.sqrt(1-rho)))

    if has_macro and Z_RECON is not None:
        ph0_c=AR1_INFO["ph0"]; ph1_c=AR1_INFO["ph1"]; sig_c=AR1_INFO["sig"]
        pc1_last=float(PC1_MACRO[-1])
        pc1_2025_proj=ph0_c+ph1_c*pc1_last
        pc1_2025_lo=pc1_2025_proj-1.645*sig_c
        pc1_2025_hi=pc1_2025_proj+1.645*sig_c

        def pc1_to_z(pc1_val):
            x_new=np.zeros(len(B_ACP)); x_new[0]=1.0; x_new[1]=pc1_val
            return float(x_new@B_ACP)

        Z_SCEN = {
            "Favorable (P25)": pc1_to_z(np.percentile(PC1_MACRO,25)),
            "Neutre (AR1)":    pc1_to_z(pc1_2025_proj),
            "Adverse (P75)":   pc1_to_z(np.percentile(PC1_MACRO,75)),
            "Crise (P90)":     pc1_to_z(np.percentile(PC1_MACRO,90)),
        }
        Z_2025 = Z_SCEN["Neutre (AR1)"]
    else:
        Z_SCEN = {"Favorable":-0.75,"Neutre":0.0,"Adverse":0.75,"Crise":1.5}
        Z_2025 = 0.0

    # PD_PIT par scénario et par obligation
    for sc,z_sc in Z_SCEN.items():
        col=f"PD_PIT_{sc.split()[0]}"
        df_w[col]=df_w["ALPHA_I"].apply(lambda a: pd_pit_model(float(a),z_sc))
    df_w["PD_PIT_NEUTRE"]=df_w["ALPHA_I"].apply(lambda a: pd_pit_model(float(a),Z_2025))
    df_w["PD_PIT_NEUTRE"]=df_w["ALPHA_I"].apply(lambda a: pd_pit_model(float(a),Z_2025))
    # Propager PD_PIT Vasicek (neutre 2025) vers df global
    if "PD_PIT_NEUTRE" in df_w.columns and "CODE_ISIN" in df.columns and "CODE_ISIN" in df_w.columns:
        _pit_v=df_w.set_index("CODE_ISIN")["PD_PIT_NEUTRE"]
        df["PD_PIT_NEUTRE"]=df["CODE_ISIN"].map(_pit_v).fillna(df["PD_TTC"])
    elif "PD_PIT_NEUTRE" in df_w.columns:
        df["PD_PIT_NEUTRE"]=df_w["PD_PIT_NEUTRE"].values[:len(df)]
    df_w["PD_PIT_IC_HI"]=df_w["ALPHA_I"].apply(lambda a: pd_pit_model(float(a),Z_2025,rho=V["rho_lo"]))

    # ══════════════════════════════════════════════════════════════════════════
    # LAYOUT — 3 ONGLETS INTERNES
    # ══════════════════════════════════════════════════════════════════════════
    sub_pd = st.tabs(["Seuils & ρ","📡 Facteur Z_t & ACP","🔮 PD_PIT Scénarios"])

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 0 — SEUILS ET ρ
    # ─────────────────────────────────────────────────────────────────────────
    with sub_pd[0]:
        st.subheader("Seuils de défaut α_i et Calibration ρ (MOM + Bootstrap)")

        # KPIs
        k1,k2,k3,k4=st.columns(4)
        k1.metric("ρ̂ (MOM)",f"{RHO*100:.2f}%",
                   delta=f"IC95 [{V['rho_lo']*100:.1f}%,{V['rho_hi']*100:.1f}%]")
        k2.metric("λ̂₀",f"{LAM0:.4f}",
                   delta=f"PD_TTC moy={norm.cdf(LAM0)*100:.2f}%")
        k3.metric("λ̂₁",f"{LAM1:.4f}",
                   delta=f"λ̂₁² = {LAM1**2:.4f}")
        k4.metric("Normalité",
                   " OK" if V["jb_p"]>0.05 else " Attention",
                   delta=f"JB p={V['jb_p']:.3f} · KS p={V['ks_p']:.3f}")

        col_s1,col_s2=st.columns([1.4,1])

        with col_s1:
            # Distribution des seuils α_i par secteur
            fig_a=go.Figure()
            secs=sorted(df_w["SECTEUR"].unique())
            for sec in secs:
                g=df_w[df_w["SECTEUR"]==sec]["ALPHA_I"].dropna()
                if len(g)==0: continue
                fig_a.add_trace(go.Violin(y=g,name=sec[:14],box_visible=True,
                                           meanline_visible=True,points="all",
                                           pointpos=-1.2,jitter=0.3))
            fig_a.add_hline(y=df_w["ALPHA_I"].mean(),line_dash="dash",
                             line_color=C["navy"],annotation_text=f"Moy={df_w['ALPHA_I'].mean():.3f}")
            fig_a.update_layout(title="Distribution des Seuils α_i = Φ⁻¹(PD_TTC) par Secteur",
                                 height=400,showlegend=False,
                                 paper_bgcolor="white",plot_bgcolor="#F8F9FC")
            st.plotly_chart(fig_a,use_container_width=True, key="chart_024")

        with col_s2:
            # Bootstrap ρ
            fig_b=go.Figure()
            fig_b.add_trace(go.Histogram(x=V["rb"]*100,nbinsx=80,
                                          marker_color=C["blue"],opacity=0.72,
                                          name="Bootstrap ρ̂"))
            fig_b.add_vline(x=RHO*100,line_color=C["S3"],line_width=3,
                             annotation_text=f"ρ̂={RHO*100:.2f}%")
            fig_b.add_vrect(x0=V["rho_lo"]*100,x1=V["rho_hi"]*100,
                             fillcolor="rgba(0,163,224,0.12)",layer="below",
                             annotation_text="IC 95%")
            fig_b.update_layout(title=f"Bootstrap ρ (B={n_boot_pd:,})<br>IC 95%=[{V['rho_lo']*100:.2f}%,{V['rho_hi']*100:.2f}%]",
                                 height=400,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                 showlegend=False,xaxis_title="ρ (%)")
            st.plotly_chart(fig_b,use_container_width=True, key="chart_025")

        # Seuils par maturité — CPD_TTC
        st.subheader("Courbe CPD_TTC par Secteur (probabilités cumulées TTC)")
        T_ax=list(range(1,t_max_pd+1))
        fig_cpd=go.Figure()
        clrs=px.colors.qualitative.Plotly
        for i,sec in enumerate(secs):
            g=df_w[df_w["SECTEUR"]==sec]
            pd1=float(g["PD_TTC"].mean())
            cpd=[1-(1-pd1)**t for t in T_ax]
            fig_cpd.add_trace(go.Scatter(x=T_ax,y=[c*100 for c in cpd],
                                          mode="lines+markers",name=sec[:16],
                                          line=dict(color=clrs[i%len(clrs)],width=2),
                                          marker=dict(size=6)))
        fig_cpd.update_layout(title="CPD_TTC(T) = 1−(1−PD_TTC)^T par Secteur",
                               xaxis_title="Horizon T (ans)",yaxis_title="CPD_TTC (%)",
                               height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                               legend=dict(orientation="h",y=-0.25))
        st.plotly_chart(fig_cpd,use_container_width=True, key="chart_026")

        # Tableau PD TTC par secteur
        tbl=df_w.groupby("SECTEUR").agg(
            N=("PD_TTC","count"),
            PD_min=("PD_TTC","min"),PD_moy=("PD_TTC","mean"),PD_max=("PD_TTC","max"),
            alpha_moy=("ALPHA_I","mean")
        ).reset_index()
        for col in ["PD_min","PD_moy","PD_max"]:
            tbl[col]=(tbl[col]*100).round(3).astype(str)+"%"
        tbl["alpha_moy"]=tbl["alpha_moy"].round(4)
        st.dataframe(tbl,use_container_width=True,hide_index=True)

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 1 — Z_t ET ACP MACRO
    # ─────────────────────────────────────────────────────────────────────────
    with sub_pd[1]:
        st.subheader("Facteur Systémique Z_t — Extraction et Lien Macro (ACP)")

        # Graphique Z_t historique
        col_z1,col_z2=st.columns([2,1])
        with col_z1:
            fig_z=go.Figure()
            # Zones régimes
            for i in range(len(ANNEES)-1):
                r=V["regimes"][i]
                col_z=("rgba(167,30,30,0.08)" if r=="Adverse" else
                        "rgba(26,107,46,0.08)" if r=="Favorable" else "rgba(133,100,4,0.04)")
                fig_z.add_vrect(x0=int(ANNEES[i]),x1=int(ANNEES[i+1]),
                                 fillcolor=col_z,layer="below",line_width=0)
            # Z_t
            fig_z.add_trace(go.Scatter(x=ANNEES.tolist(),y=Z_HIST.tolist(),
                                        mode="lines+markers",name="Z_t historique",
                                        line=dict(color=C["navy"],width=2.5),
                                        marker=dict(size=7,
                                            color=["#A71E1E" if r=="Adverse" else
                                                   "#1A6B2E" if r=="Favorable" else "#856404"
                                                   for r in V["regimes"]],
                                            line=dict(color="white",width=1))))
            # Z_t reconstruit macro
            if Z_RECON is not None:
                fig_z.add_trace(go.Scatter(x=ANNEES_C.tolist(),y=Z_RECON.tolist(),
                                            mode="lines",name=f"Z_t reconstitué (ACP, r={r_acp:.3f})",
                                            line=dict(color=C["blue"],width=2,dash="dot")))
            fig_z.add_hline(y=0,line_dash="dash",line_color="gray",line_width=1)
            fig_z.add_hline(y=0.75,line_color=C["S3"],line_dash="dot",line_width=1,
                             annotation_text="Seuil Adverse")
            fig_z.add_hline(y=-0.75,line_color=C["S1"],line_dash="dot",line_width=1,
                             annotation_text="Seuil Favorable")
            fig_z.update_layout(title="Facteur Systémique Z_t — Régimes macro-économiques",
                                 height=420,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                 xaxis_title="Année",yaxis_title="Z_t",
                                 legend=dict(orientation="h",y=-0.22))
            st.plotly_chart(fig_z,use_container_width=True, key="chart_027")

        with col_z2:
            # PD historique vs PD reconstruite
            pd_recon=norm.cdf(LAM0+LAM1*Z_HIST)*100
            fig_pd=go.Figure()
            fig_pd.add_trace(go.Bar(x=ANNEES.tolist(),y=(V["pd_global"]*100).tolist(),
                                     name="PD observée",marker_color=C["blue"],opacity=0.7))
            fig_pd.add_trace(go.Scatter(x=ANNEES.tolist(),y=pd_recon.tolist(),
                                         mode="lines+markers",name="PD reconstruite",
                                         line=dict(color=C["S3"],width=2)))
            fig_pd.update_layout(title="PD Historique vs Modèle",
                                   height=420,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                   legend=dict(orientation="h",y=-0.22),
                                   xaxis_title="Année",yaxis_title="PD (%)")
            st.plotly_chart(fig_pd,use_container_width=True, key="chart_028")

        # ACP macro (si disponible)
        if has_macro and Z_RECON is not None:
            st.divider()
            col_a1,col_a2=st.columns([1.2,1])
            with col_a1:
                # Z_t obs vs Z_t ACP
                fig_acp=go.Figure()
                fig_acp.add_trace(go.Scatter(x=ANNEES_C.tolist(),y=Z_C.tolist(),
                                              mode="lines+markers",name="Z_t observé",
                                              line=dict(color=C["navy"],width=2.5)))
                fig_acp.add_trace(go.Scatter(x=ANNEES_C.tolist(),y=Z_RECON.tolist(),
                                              mode="lines+markers",name=f"Z_t ACP (r={r_acp:.3f})",
                                              line=dict(color=C["blue"],width=2,dash="dash")))
                fig_acp.update_layout(title=f"Z_t Observé vs Reconstitué par ACP macro (PC1..PC{best_k})",
                                       height=340,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                       legend=dict(orientation="h",y=-0.28),
                                       xaxis_title="Année",yaxis_title="Z_t")
                st.plotly_chart(fig_acp,use_container_width=True, key="chart_029")

            with col_a2:
                # Variance expliquée ACP
                ev=AR1_INFO["ev"]
                fig_ev=go.Figure(go.Bar(
                    x=[f"PC{i+1}" for i in range(len(ev))],
                    y=[e*100 for e in ev],
                    marker_color=[C["navy"],C["blue"],C["S2"]][:len(ev)],
                    text=[f"{e*100:.1f}%" for e in ev],textposition="outside"))
                fig_ev.update_layout(title="Variance Expliquée par Composante ACP",
                                      height=280,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                      yaxis_title="%",showlegend=False)
                st.plotly_chart(fig_ev,use_container_width=True, key="chart_030")

                # AR(1) info
                st.info(f"""**AR(1) sur PC1 :**  
φ₀ = {AR1_INFO['ph0']:+.4f}  ·  φ₁ = {AR1_INFO['ph1']:.4f}  ·  σ = {AR1_INFO['sig']:.4f}""")

            # Scatter Z_obs vs Z_ACP
            from scipy.stats import pearsonr
            r_val,_=pearsonr(Z_C,Z_RECON)
            fig_sc=go.Figure()
            fig_sc.add_trace(go.Scatter(x=Z_C.tolist(),y=Z_RECON.tolist(),
                                         mode="markers+text",
                                         text=[str(int(a)) for a in ANNEES_C],
                                         textposition="top center",
                                         marker=dict(size=10,color=C["blue"],
                                                     line=dict(color="white",width=1))))
            lim=max(abs(Z_C).max(),abs(Z_RECON).max())+0.2
            fig_sc.add_shape(type="line",x0=-lim,x1=lim,y0=-lim,y1=lim,
                              line=dict(dash="dash",color="gray"))
            fig_sc.update_layout(title=f"Scatter Z_obs vs Z_ACP — r={r_val:.4f}",
                                   height=350,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                   xaxis_title="Z_t observé",yaxis_title="Z_t ACP")
            st.plotly_chart(fig_sc,use_container_width=True, key="chart_031")

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 2 — PD_PIT SCÉNARIOS + BACKTESTING
    # ─────────────────────────────────────────────────────────────────────────
    with sub_pd[2]:
        st.subheader("PD_PIT Forward-Looking par Scénario — Modèle Merton-Vasicek")

        # KPIs scénarios
        cols_sc=st.columns(len(Z_SCEN))
        for i,(sc_n,z_sc) in enumerate(Z_SCEN.items()):
            pd_moy=df_w["ALPHA_I"].apply(lambda a: pd_pit_model(float(a),z_sc)).mean()
            delta=pd_moy-df_w["PD_TTC"].mean()
            cols_sc[i].metric(sc_n,f"{pd_moy*100:.3f}%",
                               delta=f"Δ={delta*100:+.3f}pp vs TTC",
                               delta_color=("inverse" if z_sc<0 else "normal"))

        col_p1,col_p2=st.columns([1.6,1])
        with col_p1:
            # PD_TTC vs PD_PIT par scénario — scatter
            fig_pit=go.Figure()
            sc_cols={"Favorable":"#1A6B2E","Neutre":"#2E4D7B",
                     "Adverse":"#A71E1E","Crise":"#5B2D8E"}
            for sc_n,z_sc in Z_SCEN.items():
                col_key=sc_n.split()[0]
                col_pit=f"PD_PIT_{col_key}"
                lbl=sc_n; cc=list(sc_cols.values())[list(Z_SCEN.keys()).index(sc_n)%4]
                if col_pit in df_w.columns:
                    fig_pit.add_trace(go.Scatter(
                        x=df_w["PD_TTC"]*100,y=df_w[col_pit]*100,
                        mode="markers",name=lbl,
                        marker=dict(size=6,opacity=0.65,color=cc,
                                    line=dict(color="white",width=0.5))))
            # Identité
            lm=max(df_w["PD_TTC"].max(),df_w[[c for c in df_w.columns if "PD_PIT" in c]].max().max())*105
            fig_pit.add_shape(type="line",x0=0,x1=lm,y0=0,y1=lm,
                               line=dict(dash="dash",color="gray"))
            fig_pit.update_layout(title="PD_TTC vs PD_PIT par Scénario",
                                   height=420,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                   xaxis_title="PD_TTC (%)",yaxis_title="PD_PIT (%)",
                                   legend=dict(orientation="h",y=-0.25))
            st.plotly_chart(fig_pit,use_container_width=True, key="chart_032")

        with col_p2:
            # Heatmap PD_PIT par secteur × scénario
            hm_data=[]
            for sec in sorted(df_w["SECTEUR"].unique()):
                row={"Secteur":sec[:16]}
                g=df_w[df_w["SECTEUR"]==sec]
                for sc_n,z_sc in Z_SCEN.items():
                    pd_val=g["ALPHA_I"].apply(lambda a: pd_pit_model(float(a),z_sc)).mean()
                    row[sc_n.split()[0]]=round(pd_val*100,3)
                hm_data.append(row)
            df_hm=pd.DataFrame(hm_data).set_index("Secteur")
            fig_hm=go.Figure(go.Heatmap(
                z=df_hm.values,x=df_hm.columns.tolist(),y=df_hm.index.tolist(),
                colorscale="YlOrRd",text=df_hm.values.round(2),
                texttemplate="%{text}%",textfont=dict(size=9),
                colorbar=dict(title="PD_PIT %")))
            fig_hm.update_layout(title="Heatmap PD_PIT% — Secteur × Scénario",
                                   height=380,paper_bgcolor="white")
            st.plotly_chart(fig_hm,use_container_width=True, key="chart_033")

        # Sensibilité à ρ
        st.subheader("Analyse de Sensibilité — Impact de ρ sur PD_PIT")
        rho_range=np.linspace(0.02,0.35,100)
        alpha_med=float(df_w["ALPHA_I"].median())
        fig_sens=go.Figure()
        for sc_n,z_sc in Z_SCEN.items():
            cc=list(sc_cols.values())[list(Z_SCEN.keys()).index(sc_n)%4]
            pd_rho=[float(norm.cdf((alpha_med-np.sqrt(r)*z_sc)/np.sqrt(1-r)))*100
                    for r in rho_range]
            fig_sens.add_trace(go.Scatter(x=rho_range*100,y=pd_rho,
                                           mode="lines",name=sc_n,
                                           line=dict(color=cc,width=2)))
        fig_sens.add_vline(x=RHO*100,line_dash="dash",line_color=C["navy"],
                            annotation_text=f"ρ̂={RHO*100:.2f}%")
        fig_sens.add_hline(y=df_w["PD_TTC"].median()*100,line_dash="dot",
                            line_color="gray",annotation_text="PD_TTC médiane")
        fig_sens.update_layout(title=f"Sensibilité PD_PIT à ρ (α médian={alpha_med:.3f})",
                                height=360,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                                xaxis_title="ρ (%)",yaxis_title="PD_PIT (%)",
                                legend=dict(orientation="h",y=-0.25))
        st.plotly_chart(fig_sens,use_container_width=True, key="chart_034")

        # Backtesting
        st.subheader("Backtesting & Validation Statistique")
        n_train=int(T_HIST*2//3)
        Y_T=V["y_t"]; annees_all=V["annees"]
        y_train=Y_T[:n_train]; y_test=Y_T[n_train:]
        l0t=y_train.mean(); l1sq_t=np.mean((y_train-l0t)**2)
        rho_bt=l1sq_t/(1+l1sq_t); l1t=np.sqrt(l1sq_t)
        z_test=(y_test-l0t)/(l1t+1e-10)
        pd_test_obs=V["pd_adj"][n_train:]
        pd_test_pred=norm.cdf(l0t+l1t*z_test)
        rmse_bt=float(np.sqrt(np.mean((pd_test_obs-pd_test_pred)**2)))
        mae_bt=float(np.mean(np.abs(pd_test_obs-pd_test_pred)))

        c_bt1,c_bt2,c_bt3,c_bt4=st.columns(4)
        c_bt1.metric("Train",f"{annees_all[0]}–{annees_all[n_train-1]}",delta=f"{n_train} ans")
        c_bt2.metric("Test",f"{annees_all[n_train]}–{annees_all[-1]}",delta=f"{T_HIST-n_train} ans")
        c_bt3.metric("RMSE (test)",f"{rmse_bt*100:.4f}%")
        c_bt4.metric("MAE (test)",f"{mae_bt*100:.4f}%")

        fig_bt=go.Figure()
        ann_tr=annees_all[:n_train].tolist(); ann_te=annees_all[n_train:].tolist()
        fig_bt.add_trace(go.Scatter(x=ann_tr,y=(V["pd_adj"][:n_train]*100).tolist(),
                                     name="PD observée (train)",mode="lines+markers",
                                     line=dict(color=C["navy"],width=2)))
        fig_bt.add_trace(go.Scatter(x=ann_te,y=(pd_test_obs*100).tolist(),
                                     name="PD observée (test)",mode="lines+markers",
                                     line=dict(color=C["S1"],width=2)))
        fig_bt.add_trace(go.Scatter(x=ann_te,y=(pd_test_pred*100).tolist(),
                                     name="PD prédite (modèle)",mode="lines+markers",
                                     line=dict(color=C["blue"],width=2,dash="dash")))
        fig_bt.add_vrect(x0=annees_all[n_train]-0.5,x1=annees_all[-1]+0.5,
                          fillcolor="rgba(167,30,30,0.05)",layer="below",
                          annotation_text="Période test")
        fig_bt.update_layout(title="Backtesting — PD Observée vs Prédite par le Modèle",
                              height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                              xaxis_title="Année",yaxis_title="PD (%)",
                              legend=dict(orientation="h",y=-0.25))
        st.plotly_chart(fig_bt,use_container_width=True, key="chart_035")

        # Tableau résultats PD_PIT
        st.subheader("Résultats PD_PIT par Obligation")
        cols_show=["DESCRIPTION","SECTEUR","PD_TTC","ALPHA_I","PD_PIT_NEUTRE",
                   "PD_PIT_IC_LO","PD_PIT_IC_HI"]
        cols_show=[c for c in cols_show if c in df_w.columns]
        df_show=df_w[cols_show].copy()
        for c in ["PD_TTC","PD_PIT_NEUTRE","PD_PIT_IC_LO","PD_PIT_IC_HI"]:
            if c in df_show.columns:
                df_show[c]=(df_show[c]*100).round(4).astype(str)+"%"
        if "ALPHA_I" in df_show.columns:
            df_show["ALPHA_I"]=df_show["ALPHA_I"].round(4)
        st.dataframe(df_show,use_container_width=True,hide_index=True)

        # Export
        xlsx_pd=io.BytesIO()
        with pd.ExcelWriter(xlsx_pd,engine="openpyxl") as wr:
            df_w[[c for c in df_w.columns if c not in
                   ["STADE","C2_STADE","C3_STADE","C4_STADE","C5_STADE"]]
                  ].to_excel(wr,sheet_name="PD_PIT",index=False)
        st.download_button("Exporter PD_PIT (Excel)",
                            data=xlsx_pd.getvalue(),
                            file_name="IFRS9_PD_PIT_Vasicek.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key="dl_007")

# ════════════════════════════════════════════════════════════════
# TAB 7 — RÉSULTATS & EXPORT
# ════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════════
# TAB LGD — LGD PIT : VASICEK + FRYE-JACOBS
# ════════════════════════════════════════════════════════════════════════════════
with TABS[tab_cpv]:
    import statsmodels.api as sm
    from statsmodels.tsa.arima.model import ARIMA
    from statsmodels.tsa.stattools import adfuller, kpss
    from statsmodels.tsa.vector_ar.var_model import VAR
    from statsmodels.stats.outliers_influence import variance_inflation_factor
    from sklearn.linear_model import LassoCV, LinearRegression
    from sklearn.ensemble import RandomForestRegressor
    from sklearn.preprocessing import StandardScaler
    from sklearn.model_selection import cross_val_score

    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>PD CPV — Credit Portfolio View</h2>
      <p>Transformations · ADF/KPSS · VAR/ARIMA · 5 méthodes sélection · Modèle CPV · PD_PIT · Projections 5 ans · Scénarios</p>
    </div></div>""", unsafe_allow_html=True)

    if df is None:
        st.warning(" Chargez vos données dans la sidebar."); st.stop()
    if df_mac is None:
        st.warning(" Chargez Data facteur macro.xlsx dans la sidebar."); st.stop()

    with st.sidebar:
        st.markdown("### Paramètres CPV")
        cpv_seuil_adf  = st.slider("Seuil ADF/KPSS",0.05,0.25,0.20,0.01, key="seuil_adf_kpss")
        cpv_seuil_corr = st.slider("Seuil |ρ| corrélation",0.10,0.60,0.30,0.05,key="cpv_seuil_corr_1")
        cpv_vote_min   = st.slider("Vote minimum (méthodes)",2,5,3,1, key="wk_sl_0001")
        cpv_n_sim      = st.number_input("N simulations",1000,10000,5000,500, key="cpv_nsim_main")
        cpv_horizon    = st.slider("Horizon projection (ans)",3,10,5,key="cpv_horizon_1")

    SEED_CPV=42
    def _sig(x): return 1.0/(1.0+np.exp(-np.clip(x,-30,30)))
    def _logit(p): return float(np.log(np.clip(p,1e-8,1-1e-8)/(1-np.clip(p,1e-8,1-1e-8))))

    # ══ ÉTAPE 1 : CHARGEMENT & CONSTRUCTION logit_DR ═════════════════════════
    @st.cache_data(show_spinner=False)
    def cpv_etape1(mac_json, oblig_json, annee_debut=2007):
        dm = pd.read_json(io.StringIO(mac_json))
        do = pd.read_json(io.StringIO(oblig_json))
        dm = dm[dm["annee"]>=annee_debut].copy().reset_index(drop=True)
        # Variables macro disponibles
        VARS_M=["inflation","croissance_pib","chomage","taux_change",
                "masse_monetaire","taux_directeur","balance_exterieure","depense_conso","pib"]
        vars_ok=[v for v in VARS_M if v in dm.columns]
        # Construire DR synthétique depuis PD portefeuille via ACP
        pd_col="PD_TTC" if "PD_TTC" in do.columns else "PD_decimal"
        if pd_col not in do.columns:
            do["PD_TTC"]=pd.to_numeric(do.get("PD",0),errors="coerce")/100
            pd_col="PD_TTC"
        dr_base=float(do[pd_col].mean())
        # ACP sur macro → DR synthétique sensible au cycle
        M_raw=dm[vars_ok].fillna(dm[vars_ok].mean()).values
        if M_raw.shape[0]>=5 and M_raw.shape[1]>=2:
            from sklearn.preprocessing import StandardScaler as _SS
            from sklearn.decomposition import PCA as _PCA
            Ms=_SS().fit_transform(M_raw)
            pc1=_PCA(n_components=1,random_state=42).fit_transform(Ms).flatten()
            if "chomage" in vars_ok:
                idx_c=vars_ok.index("chomage")
                if float(np.corrcoef(pc1,Ms[:,idx_c])[0,1])<0: pc1=-pc1
            from sklearn.preprocessing import MinMaxScaler as _MMS
            pc_01=_MMS().fit_transform(pc1.reshape(-1,1)).flatten()
            dr_series=(dr_base*0.5 + pc_01*(dr_base*1.5)).clip(1e-6,1-1e-6)
        else:
            dr_series=np.full(len(dm),dr_base)
        dm["DR"]=dr_series
        dm["logit_DR"]=[_logit(v) for v in dr_series]
        return dm.to_json(), do.to_json(), vars_ok

    with st.spinner("Étape 1 — Chargement et construction logit_DR…"):
        dm_json, do_json, VARS_CPV = cpv_etape1(df_mac.to_json(), df.to_json())

    dm_cpv = pd.read_json(io.StringIO(dm_json)); do_cpv = pd.read_json(io.StringIO(do_json))
    pd_col_cpv = "PD_TTC" if "PD_TTC" in do_cpv.columns else "PD_decimal"
    if pd_col_cpv not in do_cpv.columns:
        do_cpv["PD_TTC"]=pd.to_numeric(do_cpv.get("PD",0),errors="coerce")/100; pd_col_cpv="PD_TTC"

    # ══ ÉTAPE 2 : TRANSFORMATIONS + STATIONNARITÉ ═════════════════════════════
    @st.cache_data(show_spinner=False)
    def cpv_etape2(dm_json, vars_ok, seuil):
        dm=pd.read_json(io.StringIO(dm_json))
        results=[]
        df_t=pd.DataFrame({"annee":dm["annee"].values,"logit_DR":dm["logit_DR"].values})
        for v in vars_ok:
            serie=dm[v].fillna(dm[v].mean())
            # Tester 4 transformations
            transfo_best=None; pval_best=1.0; col_best=v
            for tf,fn in [("niveau",lambda s:s),
                          ("diff",  lambda s:s.diff()),
                          ("logdiff",lambda s:np.log(s.clip(1e-8)).diff()),
                          ("logit", lambda s: np.log(s.clip(1e-7)/( 1-s.clip(1e-7,1-1e-7)) ) if s.min()>0 and s.max()<1 else s.diff())]:
                try:
                    s_t=fn(serie).dropna()
                    if len(s_t)<5: continue
                    adf_p=adfuller(s_t,autolag="AIC")[1]
                    if adf_p<pval_best:
                        pval_best=adf_p; transfo_best=tf; serie_best=fn(serie)
                except: pass
            serie_final=serie_best if transfo_best else serie
            col_name=f"{v}_{transfo_best}" if transfo_best and transfo_best!="niveau" else v
            df_t[col_name]=serie_final.values[:len(dm)]
            # Ajouter lags 0,1,2,3
            for lag in range(4):
                lag_col=f"{col_name}_lag{lag}"
                df_t[lag_col]=serie_final.shift(lag).values[:len(dm)]
            stationnaire=pval_best<=seuil
            results.append({"Variable":v,"Transformation":transfo_best,"ADF_p":round(pval_best,4),
                             "Stationnaire":"" if stationnaire else "","Col":col_name})
        return pd.DataFrame(results).to_json(), df_t.to_json()

    with st.spinner("Étape 2 — Transformations et tests ADF…"):
        df_rapport_json, df_t_json = cpv_etape2(dm_json, VARS_CPV, cpv_seuil_adf)

    df_rapport=pd.read_json(io.StringIO(df_rapport_json))
    df_t_cpv=pd.read_json(io.StringIO(df_t_json))

    # ══ ÉTAPE 3 : VAR + ARIMA ══════════════════════════════════════════════════
    @st.cache_data(show_spinner=False)
    def cpv_etape3(dt_json, rapport_json):
        dt=pd.read_json(io.StringIO(dt_json)); rapp=pd.read_json(io.StringIO(rapport_json))
        feat_cols=[c for c in dt.columns if c not in ["annee","logit_DR"] and dt[c].notna().sum()>5]
        # VAR(p) sur les 4 meilleures variables (stationnaires)
        vars_stat=[r["Col"] for _,r in rapp.iterrows() if r["Stationnaire"]==""][:4]
        vars_stat=[v for v in vars_stat if v in dt.columns]
        var_results={"p":1,"aic":None,"stable":False,"vars":vars_stat}
        if len(vars_stat)>=2:
            try:
                dt_var=dt[vars_stat].dropna()
                best_aic=np.inf; best_p=1
                for p in range(1,min(4,len(dt_var)//4)):
                    try:
                        m=VAR(dt_var).fit(maxlags=p,ic=None)
                        if m.aic<best_aic: best_aic=m.aic; best_p=p
                    except: pass
                m_best=VAR(dt_var).fit(maxlags=best_p,ic=None)
                is_stable=all(np.abs(e)<=1.0 for e in m_best.roots)
                var_results={"p":best_p,"aic":round(float(best_aic),3),"stable":is_stable,"vars":vars_stat}
            except: pass
        # ARIMA par variable
        arima_results={}
        for v in (rapp["Col"].tolist() if "Col" in rapp.columns else []):
            if v not in dt.columns: continue
            s=dt[v].dropna()
            if len(s)<6: continue
            best_aic=np.inf; best_ord=(1,0,0)
            for p in range(3):
                for q in range(3):
                    try:
                        m=ARIMA(s,order=(p,0,q)).fit()
                        if m.aic<best_aic: best_aic=m.aic; best_ord=(p,0,q)
                    except: pass
            arima_results[v]={"ordre":best_ord,"aic":round(float(best_aic),3)}
        return var_results, arima_results

    with st.spinner("Étape 3 — VAR et ARIMA…"):
        var_res_dict, arima_res_dict = cpv_etape3(df_t_json, df_rapport_json)

    # ══ ÉTAPE 4 : SÉLECTION 5 MÉTHODES + MODÈLE CPV ══════════════════════════
    @st.cache_data(show_spinner=False)
    def cpv_etape4(dt_json, rapport_json, seuil_corr, vote_min):
        dt=pd.read_json(io.StringIO(dt_json)); rapp=pd.read_json(io.StringIO(rapport_json))
        y=dt["logit_DR"].dropna()
        feat_cols=[c for c in dt.columns if c.startswith(tuple(rapp["Col"].tolist() if "Col" in rapp.columns else []))
                    and c in dt.columns and dt[c].notna().sum()>=len(y)-2]
        common_idx=dt[feat_cols+["logit_DR"]].dropna().index
        Xc=dt.loc[common_idx,feat_cols]; yc=y.loc[common_idx]; ann_c=dt.loc[common_idx,"annee"]
        if len(yc)<5: return None,None,None,None,None,None,None
        # M1 Corrélation
        corrs={c:abs(Xc[c].corr(yc)) for c in Xc.columns if Xc[c].notna().all()}
        m1=[c for c,v in corrs.items() if v>=seuil_corr]
        # M2 OLS univarié (p<=0.15)
        m2=[]
        for c in Xc.columns:
            try:
                r_=sm.OLS(yc,sm.add_constant(Xc[[c]])).fit()
                if c in r_.pvalues and r_.pvalues[c]<=0.15: m2.append(c)
            except: pass
        # M3 Forward stepwise
        m3=[]; rem=list(Xc.columns); best_aic=np.inf
        for _ in range(min(6,len(rem))):
            best_c=None
            for c in rem:
                try:
                    r_=sm.OLS(yc,sm.add_constant(Xc[m3+[c]])).fit()
                    if r_.aic<best_aic: best_aic=r_.aic; best_c=c
                except: pass
            if best_c: m3.append(best_c); rem.remove(best_c)
            else: break
        # M4 LASSO
        m4=[]
        try:
            Xs=StandardScaler().fit_transform(Xc.fillna(0))
            la=LassoCV(cv=min(5,len(yc)-1),max_iter=5000,random_state=SEED_CPV).fit(Xs,yc)
            m4=[c for c,w in zip(Xc.columns,la.coef_) if w!=0]
        except: pass
        # M5 Random Forest
        m5=[]
        try:
            rf=RandomForestRegressor(n_estimators=200,random_state=SEED_CPV,n_jobs=-1)
            rf.fit(Xc.fillna(0),yc)
            imp=pd.Series(rf.feature_importances_,index=Xc.columns)
            m5=list(imp[imp>=0.05].sort_values(ascending=False).index)
        except: pass
        # Vote consensus
        all_v=set(m1+m2+m3+m4+m5)
        votes={v:sum(1 for m_ in [m1,m2,m3,m4,m5] if v in m_) for v in all_v}
        consensus=[v for v,sc in votes.items() if sc>=vote_min]
        # Modèle final OLS
        vars_final=consensus if consensus else m1[:3] if m1 else list(Xc.columns[:2])
        vars_final=[v for v in vars_final if v in Xc.columns]
        if not vars_final: vars_final=list(Xc.columns[:2])
        try:
            res_best=sm.OLS(yc,sm.add_constant(Xc[vars_final])).fit()
        except: return None,None,None,None,None,None,None
        # VIF
        X_vif=sm.add_constant(Xc[vars_final].fillna(0))
        try:
            vif=[{"Variable":c,"VIF":round(variance_inflation_factor(X_vif.values,i),2)}
                  for i,c in enumerate(X_vif.columns) if c!="const"]
        except: vif=[]
        return (res_best,Xc[vars_final],yc,ann_c,
                pd.DataFrame({"Variable":list(votes.keys()),"Votes":list(votes.values()),
                               "Retenu":["" if k in consensus else "—" for k in votes.keys()]}).to_json(),
                pd.DataFrame(vif).to_json() if vif else None,
                {"m1":len(m1),"m2":len(m2),"m3":len(m3),"m4":len(m4),"m5":len(m5)})

    with st.spinner("Étape 4 — Sélection 5 méthodes + modèle CPV…"):
        cpv_result = cpv_etape4(df_t_json, df_rapport_json, cpv_seuil_corr, cpv_vote_min)
        res_cpv4, Xc_cpv4, yc_cpv4, ann_cpv4, votes_json4, vif_json4, meth_counts = cpv_result

    # ══ ÉTAPE 5 : PD_PIT PAR OBLIGATION ═══════════════════════════════════════
    if res_cpv4 is not None:
        Y_hist_moy = float(yc_cpv4.mean())
        Y_last     = float(res_cpv4.fittedvalues.iloc[-1])
        delta_cpv  = Y_last - Y_hist_moy
        # Appliquer delta à chaque obligation
        df["logit_PD_TTC"] = df[pd_col_cpv].apply(lambda p: _logit(float(p)))
        df["PD_PIT_CPV"]   = df["logit_PD_TTC"].apply(lambda y_: _sig(y_+delta_cpv))
        df["delta_PD_CPV"] = df["PD_PIT_CPV"] - df[pd_col_cpv]
        df["logit_PIT"]    = df["logit_PD_TTC"] + delta_cpv

    # ══ ÉTAPE 6+7 : PROJECTIONS ARIMA + SCÉNARIOS ════════════════════════════
    @st.cache_data(show_spinner=False)
    def cpv_projections(dt_json, res_params, res_vars, yc_arr, n_sim, horizon, seed):
        np.random.seed(seed); dt=pd.read_json(io.StringIO(dt_json))
        beta=pd.Series(res_params); Y_moy=float(np.mean(yc_arr)); sig=float(np.std(yc_arr))*0.6
        # Projeter logit_DR via Monte Carlo (bootstrap résidus)
        Y_sim=np.full(n_sim,float(beta.get("const",0.0)))
        for c in res_vars:
            if c in dt.columns:
                mu_c=float(dt[c].mean()); sg_c=float(dt[c].std()+1e-6)
                Y_sim+=float(beta.get(c,0.0))*np.random.normal(mu_c,sg_c,n_sim)
        Y_sim+=np.random.normal(0,max(sig,0.01),n_sim)
        # Projections par horizon (ARIMA simple AR(1))
        phi1=0.6; traj=np.zeros((horizon,n_sim))
        Y_now=float(Y_sim.mean())
        for t in range(horizon):
            noise=np.random.normal(0,sig,n_sim)
            Y_now=phi1*Y_now+(1-phi1)*Y_moy+noise*0.3
            traj[t]=Y_sim*((horizon-t)/horizon)+Y_now*(t/horizon)+noise*0.1
        PD_proj=1/(1+np.exp(-np.clip(traj,-30,30)))
        return {"traj":traj.tolist(),"PD_proj":PD_proj.tolist(),
                "Y_sim":Y_sim[:2000].tolist(),"Y_moy":Y_moy,"sig":sig}

    with st.spinner("Étapes 6-7 — Projections et scénarios…"):
        if res_cpv4 is not None:
            proj_res=cpv_projections(df_t_json,res_cpv4.params.to_dict(),
                                      list(Xc_cpv4.columns),yc_cpv4.tolist(),
                                      cpv_n_sim,cpv_horizon,SEED_CPV)
        else:
            proj_res=None

    # ══ VISUALISATIONS — SOUS-ONGLETS ══════════════════════════════════════════
    sub_cpv=st.tabs(["EDA & Stationnarité","🔧 Sélection & Modèle",
                      "PD_PIT par Obligation","🎲 MC & Stress","Projections"])

    # ─── SUB 0 : EDA & Stationnarité ─────────────────────────────────────────
    with sub_cpv[0]:
        st.subheader("Étape 1-2 : Données Macro + Tests Stationnarité")
        k1,k2,k3,k4=st.columns(4)
        k1.metric("N observations",str(len(dm_cpv)))
        k2.metric("Variables macro",str(len(VARS_CPV)))
        k3.metric("logit_DR moyen",f"{dm_cpv['logit_DR'].mean():.4f}")
        k4.metric("DR moyen",f"{dm_cpv['DR'].mean()*100:.2f}%")
        col_e1,col_e2=st.columns(2)
        with col_e1:
            fig_dr=go.Figure()
            fig_dr.add_trace(go.Bar(x=dm_cpv["annee"].tolist(),y=(dm_cpv["DR"]*100).tolist(),
                marker_color=[C["red"] if v>0.05 else(C["amber"] if v>0.03 else C["green"])
                               for v in dm_cpv["DR"].values],opacity=0.82,name="DR(%)"))
            fig_dr.add_trace(go.Scatter(x=dm_cpv["annee"].tolist(),y=dm_cpv["logit_DR"].tolist(),
                mode="lines+markers",yaxis="y2",name="logit(DR)",
                line=dict(color=C["navy"],width=2.5),marker=dict(size=6)))
            fig_dr.update_layout(title="Taux de Défaut Synthétique + logit(DR)",
                height=360,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                yaxis=dict(title="DR (%)"),
                yaxis2=dict(title="logit(DR)",overlaying="y",side="right"),
                legend=dict(orientation="h",y=-0.25))
            st.plotly_chart(fig_dr,use_container_width=True, key="chart_036")
        with col_e2:
            fig_mac2=go.Figure()
            for i,v in enumerate(VARS_CPV[:6]):
                s=dm_cpv[v].dropna()
                fig_mac2.add_trace(go.Scatter(x=dm_cpv["annee"].tolist()[:len(s)],y=s.tolist(),
                    mode="lines+markers",name=v[:14],
                    line=dict(width=2),marker=dict(size=4)))
            fig_mac2.update_layout(title="Séries Macroéconomiques 2007-2024",
                height=360,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                legend=dict(orientation="h",y=-0.35,font=dict(size=8)))
            st.plotly_chart(fig_mac2,use_container_width=True, key="chart_037")
        # Tableau ADF
        st.subheader("Tests de Stationnarité (ADF)")
        st.dataframe(df_rapport[["Variable","Transformation","ADF_p","Stationnaire"]],
                      use_container_width=True,hide_index=True)
        # Corrélations avec logit_DR
        corr_vals={v:float(dm_cpv[v].corr(dm_cpv["logit_DR"])) for v in VARS_CPV if v in dm_cpv.columns}
        cv_s=pd.Series(corr_vals).sort_values()
        fig_cv=go.Figure(go.Bar(y=cv_s.index.tolist(),x=cv_s.values.tolist(),
            orientation="h",marker_color=[C["red"] if v>0 else C["green"] for v in cv_s.values],
            opacity=0.82))
        fig_cv.add_vline(x=cpv_seuil_corr,line_dash="dash",line_color=C["green"])
        fig_cv.add_vline(x=-cpv_seuil_corr,line_dash="dash",line_color=C["green"])
        fig_cv.update_layout(title="Corrélation Variables × logit(DR)",
            height=320,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
            xaxis_title="ρ",showlegend=False)
        st.plotly_chart(fig_cv,use_container_width=True, key="chart_038")

    # ─── SUB 1 : Sélection & Modèle ──────────────────────────────────────────
    with sub_cpv[1]:
        st.subheader("Étapes 3-4 : VAR/ARIMA + Sélection 5 Méthodes + OLS")
        # VAR info
        col_v1,col_v2=st.columns(2)
        with col_v1:
            st.info(f"**VAR(p={var_res_dict['p']})** | AIC={var_res_dict['aic']} | "
                     f"{' Stable' if var_res_dict['stable'] else ' Instable'} | "
                     f"Variables: {', '.join(var_res_dict['vars'][:4])}")
            # ARIMA summary
            arima_tbl=[{"Variable":k,"Ordre":str(v["ordre"]),"AIC":v["aic"]}
                        for k,v in arima_res_dict.items()]
            if arima_tbl:
                st.dataframe(pd.DataFrame(arima_tbl),use_container_width=True,hide_index=True)
        with col_v2:
            # Vote consensus
            if votes_json4:
                df_votes=pd.read_json(io.StringIO(votes_json4)).sort_values("Votes",ascending=False)
                st.dataframe(df_votes,use_container_width=True,hide_index=True)
            if meth_counts:
                fig_mc2=go.Figure(go.Bar(
                    x=["M1 Corr","M2 OLS","M3 Forward","M4 LASSO","M5 RF"],
                    y=[meth_counts["m1"],meth_counts["m2"],meth_counts["m3"],
                        meth_counts["m4"],meth_counts["m5"]],
                    marker_color=[C["blue"],C["green"],C["amber"],C["red"],C["purple"]],
                    opacity=0.82))
                fig_mc2.update_layout(title="Variables sélectionnées par méthode",
                    height=280,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                    yaxis_title="N variables",showlegend=False)
                st.plotly_chart(fig_mc2,use_container_width=True, key="chart_039")

        if res_cpv4 is not None:
            k1,k2,k3,k4=st.columns(4)
            k1.metric("R²",f"{res_cpv4.rsquared:.4f}")
            k2.metric("R² ajusté",f"{res_cpv4.rsquared_adj:.4f}")
            k3.metric("AIC",f"{res_cpv4.aic:.2f}")
            k4.metric("Δ cycle",f"{delta_cpv:+.4f}",
                       delta="Adverse ↑PD" if delta_cpv>0 else "Favorable ↓PD")
            col_m1,col_m2=st.columns(2)
            with col_m1:
                coefs=res_cpv4.params.drop("const",errors="ignore")
                pvals=res_cpv4.pvalues.drop("const",errors="ignore")
                conf=res_cpv4.conf_int().drop("const",errors="ignore")
                fig_coef=go.Figure(go.Bar(
                    y=coefs.index.tolist(),x=coefs.values.tolist(),orientation="h",
                    marker_color=[C["red"] if v>0 else C["green"] for v in coefs.values],
                    opacity=0.82,error_x=dict(type="data",symmetric=False,
                    array=(conf[1]-coefs).values.tolist(),
                    arrayminus=(coefs-conf[0]).values.tolist())))
                fig_coef.add_vline(x=0,line_color="black",line_width=1)
                fig_coef.update_layout(title=f"Coefficients β CPV (IC 95%) R²={res_cpv4.rsquared:.4f}",
                    height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",showlegend=False)
                st.plotly_chart(fig_coef,use_container_width=True, key="chart_040")
            with col_m2:
                y_pred=res_cpv4.fittedvalues
                fig_fit=go.Figure()
                fig_fit.add_trace(go.Scatter(x=ann_cpv4.tolist(),y=yc_cpv4.tolist(),
                    mode="lines+markers",name="logit(DR) observé",
                    line=dict(color=C["navy"],width=2.5),marker=dict(size=7)))
                fig_fit.add_trace(go.Scatter(x=ann_cpv4.tolist(),y=y_pred.tolist(),
                    mode="lines+markers",name="CPV prédit",
                    line=dict(color=C["blue"],width=2,dash="dash"),marker=dict(size=6,symbol="square")))
                fig_fit.update_layout(title="logit(DR) Observé vs CPV Prédit",
                    height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                    legend=dict(orientation="h",y=-0.25))
                st.plotly_chart(fig_fit,use_container_width=True, key="chart_041")
            # VIF
            if vif_json4:
                st.subheader("Multicolinéarité (VIF)")
                df_vif=pd.read_json(io.StringIO(vif_json4))
                fig_vif=go.Figure(go.Bar(x=df_vif["Variable"].tolist(),y=df_vif["VIF"].tolist(),
                    marker_color=[C["red"] if v>10 else(C["amber"] if v>5 else C["green"])
                                   for v in df_vif["VIF"].values],opacity=0.82))
                fig_vif.add_hline(y=5,line_dash="dash",line_color=C["amber"],annotation_text="VIF=5")
                fig_vif.add_hline(y=10,line_dash="dash",line_color=C["red"],annotation_text="VIF=10")
                fig_vif.update_layout(title="VIF par variable (vert<5  orange 5-10  rouge>10 ❌)",
                    height=300,paper_bgcolor="white",plot_bgcolor="#F8F9FC",showlegend=False)
                st.plotly_chart(fig_vif,use_container_width=True, key="chart_042")

    # ─── SUB 2 : PD_PIT par Obligation ───────────────────────────────────────
    with sub_cpv[2]:
        st.subheader("Étape 5 — PD_PIT par Obligation (Ajustement Cyclique CPV)")
        if res_cpv4 is None:
            st.error("Modèle CPV requis"); st.stop()
        k1,k2,k3,k4=st.columns(4)
        k1.metric("Δ cycle (logit)",f"{delta_cpv:+.4f}")
        k2.metric("Régime"," Adverse" if delta_cpv>0.5 else(" Favorable" if delta_cpv<-0.5 else " Neutre"))
        k3.metric("PD_TTC moy",f"{df[pd_col_cpv].mean()*100:.3f}%")
        k4.metric("PD_PIT_CPV moy",f"{df['PD_PIT_CPV'].mean()*100:.3f}%",
                   delta=f"{(df['PD_PIT_CPV'].mean()-df[pd_col_cpv].mean())*100:+.3f}pp")
        # Scatter
        col_c1,col_c2=st.columns([1.2,1])
        with col_c1:
            fig_sc=go.Figure()
            sect_col_="SECTEUR_GICS" if "SECTEUR_GICS" in df.columns else "SECTEUR"
            clrs_=px.colors.qualitative.Plotly
            if sect_col_ in df.columns:
                for i_s,sec in enumerate(df[sect_col_].unique()):
                    g=df[df[sect_col_]==sec]
                    fig_sc.add_trace(go.Scatter(x=g[pd_col_cpv]*100,y=g["PD_PIT_CPV"]*100,
                        mode="markers",name=sec[:14],
                        marker=dict(size=7,opacity=0.75,color=clrs_[i_s%len(clrs_)],
                        line=dict(color="white",width=0.5))))
            lm_=max(df[pd_col_cpv].max(),df["PD_PIT_CPV"].max())*105
            fig_sc.add_shape(type="line",x0=0,x1=lm_,y0=0,y1=lm_,
                line=dict(dash="dash",color="gray"))
            fig_sc.update_layout(title=f"PD_TTC vs PD_PIT_CPV<br>Δ={delta_cpv:+.4f}",
                height=420,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="PD_TTC (%)",yaxis_title="PD_PIT_CPV (%)",
                legend=dict(orientation="h",y=-0.28))
            st.plotly_chart(fig_sc,use_container_width=True, key="chart_043")
        with col_c2:
            # Heatmap PD par secteur
            hm_=[]; 
            if sect_col_ in df.columns:
                for sec in df[sect_col_].unique():
                    g=df[df[sect_col_]==sec]
                    hm_.append({"Secteur":sec[:16],"PD_TTC":round(g[pd_col_cpv].mean()*100,3),
                                  "PD_PIT":round(g["PD_PIT_CPV"].mean()*100,3),
                                  "Δ(pp)":round(g["delta_PD_CPV"].mean()*100,3)})
                df_hm_=pd.DataFrame(hm_).set_index("Secteur")
                fig_hm_=go.Figure(go.Heatmap(z=df_hm_.values,x=df_hm_.columns.tolist(),
                    y=df_hm_.index.tolist(),colorscale="YlOrRd",
                    text=df_hm_.round(3).values,texttemplate="%{text}",
                    textfont=dict(size=9),colorbar=dict(title="PD%")))
                fig_hm_.update_layout(title="Heatmap PD CPV par Secteur",
                    height=380,paper_bgcolor="white")
                st.plotly_chart(fig_hm_,use_container_width=True, key="chart_044")
        # PD Cumulative convergence
        st.subheader("CPD_PIT CPV par Secteur — Convergence dans le Temps")
        T_conv=[1,2,3,4,5,6,7,8,9,10]
        fig_conv=go.Figure()
        clrs_cv=px.colors.qualitative.Plotly
        if sect_col_ in df.columns:
            for i_s,sec in enumerate(df[sect_col_].unique()):
                g=df[df[sect_col_]==sec]
                if len(g)==0: continue
                pd_ttc_s=float(g[pd_col_cpv].mean()); pd_pit_s=float(g["PD_PIT_CPV"].mean())
                cpd_ttc=[(1-(1-pd_ttc_s)**t)*100 for t in T_conv]
                cpd_pit=[(1-(1-pd_pit_s)**t)*100 for t in T_conv]
                col_s=clrs_cv[i_s%len(clrs_cv)]
                fig_conv.add_trace(go.Scatter(x=T_conv,y=cpd_ttc,mode="lines",
                    name=f"{sec[:12]} TTC",line=dict(color=col_s,width=1.2,dash="dot")))
                fig_conv.add_trace(go.Scatter(x=T_conv,y=cpd_pit,mode="lines+markers",
                    name=f"{sec[:12]} CPV",line=dict(color=col_s,width=2.5),marker=dict(size=5)))
        fig_conv.update_layout(title=f"Convergence CPD (TTC pointillé | CPV plein) Δ={delta_cpv:+.4f}",
            height=480,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
            xaxis_title="Horizon T (ans)",yaxis_title="CPD (%)",
            legend=dict(orientation="h",y=-0.35,font=dict(size=8)),
            xaxis=dict(tickvals=T_conv))
        st.plotly_chart(fig_conv,use_container_width=True, key="chart_045")
        # Tableau par obligation
        st.subheader("PD_PIT CPV par Obligation")
        desc_col_=next((c for c in ["NOM_EMETTEUR","DESCRIPTION","LIBELLE_OBLIG"] if c in df.columns),"")
        cols_t=[c for c in [desc_col_,sect_col_,pd_col_cpv,"PD_PIT_CPV","delta_PD_CPV"] if c and c in df.columns]
        df_tbl_=df[cols_t].copy()
        for c in [pd_col_cpv,"PD_PIT_CPV","delta_PD_CPV"]:
            if c in df_tbl_.columns: df_tbl_[c]=(df_tbl_[c]*100).round(4).astype(str)+"%"
        st.dataframe(df_tbl_,use_container_width=True,hide_index=True,height=450)

    # ─── SUB 3 : MC & Stress ─────────────────────────────────────────────────
    with sub_cpv[3]:
        st.subheader("Monte Carlo & Stress Tests CPV — PD uniquement")
        if proj_res:
            Y_sim_arr=np.array(proj_res["Y_sim"])
            PD_sim=1/(1+np.exp(-np.clip(Y_sim_arr,-30,30)))*100
            k1,k2,k3,k4=st.columns(4)
            k1.metric("PD_macro moy MC",f"{PD_sim.mean():.3f}%")
            k2.metric("PD_macro P50",f"{np.percentile(PD_sim,50):.3f}%")
            k3.metric("PD_macro P5",f"{np.percentile(PD_sim,5):.3f}%",delta_color="inverse")
            k4.metric("PD_macro P95",f"{np.percentile(PD_sim,95):.3f}%")
            col_mc1,col_mc2=st.columns(2)
            with col_mc1:
                fig_mc_=go.Figure(go.Histogram(x=PD_sim.tolist(),nbinsx=70,
                    marker=dict(color=C["blue"],line=dict(color="white",width=0.3)),opacity=0.80))
                for pct,col_v,lbl in [(5,C["green"],"P5"),(50,C["amber"],"P50"),(95,C["red"],"P95")]:
                    fig_mc_.add_vline(x=float(np.percentile(PD_sim,pct)),line_color=col_v,line_width=2,
                        annotation_text=f"{lbl}={np.percentile(PD_sim,pct):.2f}%",annotation_font_size=8)
                fig_mc_.update_layout(title=f"Distribution PD_macro CPV ({cpv_n_sim:,} sim.)",
                    height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                    xaxis_title="PD_macro (%)",showlegend=False)
                st.plotly_chart(fig_mc_,use_container_width=True, key="chart_046")
            with col_mc2:
                Y_m_cpv=proj_res["Y_moy"]; sig_y_cpv=proj_res["sig"]
                st_rows_=[]; 
                for nom,choc,col_st in [(" Optimiste",+1.5,C["green"]),
                                          (" Baseline",0,C["blue"]),
                                          ("🟠 Adverse",-1.5,C["amber"]),
                                          (" Crise",-3.0,C["red"])]:
                    y_s=Y_m_cpv+choc*sig_y_cpv
                    d_s=y_s-Y_m_cpv
                    pd_s=float(df[pd_col_cpv].apply(lambda p: _sig(_logit(float(p))+d_s)).mean())
                    st_rows_.append({"Scénario":nom,"ΔY":f"{choc:+.1f}σ",
                                      "PD_macro%":round(pd_s*100,4),
                                      "Δ vs TTC pp":round((pd_s-float(df[pd_col_cpv].mean()))*100,4)})
                st.dataframe(pd.DataFrame(st_rows_),use_container_width=True,hide_index=True)
                fig_st_=go.Figure(go.Bar(x=[r["Scénario"] for r in st_rows_],
                    y=[r["PD_macro%"] for r in st_rows_],
                    marker_color=[C["green"],C["blue"],C["amber"],C["red"]],
                    opacity=0.82,text=[f"{r['PD_macro%']:.3f}%" for r in st_rows_],
                    textposition="outside"))
                fig_st_.update_layout(title="PD_macro CPV par Scénario",
                    height=280,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                    yaxis_title="PD_macro (%)",showlegend=False)
                st.plotly_chart(fig_st_,use_container_width=True, key="chart_047")

    # ─── SUB 4 : Résumé CPV ──────────────────────────────────────────────────
    with sub_cpv[4]:
        st.subheader("Résumé CPV & Comparaison PD_TTC vs PD_PIT")
        st.info(" Les projections EAD/ECL détaillées sont disponibles dans l'onglet **💰 EAD & ECL** avec les 3 scénarios Vasicek/Frye-Jacobs.")
        # Tableau synthèse PD par secteur
        pd_col_sum = "PD_TTC" if "PD_TTC" in df.columns else "PD_decimal"
        sect_col_s = "SECTEUR_GICS" if "SECTEUR_GICS" in df.columns else "SECTEUR"
        if pd_col_sum in df.columns and "PD_PIT_CPV" in df.columns:
            summ=[{"Secteur":sec,
                   "N":len(g),
                   "PD_TTC%":f"{g[pd_col_sum].mean()*100:.3f}%",
                   "PD_PIT_CPV%":f"{g['PD_PIT_CPV'].mean()*100:.3f}%",
                   "Δ(pp)":f"{(g['PD_PIT_CPV'].mean()-g[pd_col_sum].mean())*100:+.3f}pp",
                   "CPD(5a)TTC":f"{(1-(1-g[pd_col_sum].mean())**5)*100:.2f}%",
                   "CPD(5a)CPV":f"{(1-(1-g['PD_PIT_CPV'].mean())**5)*100:.2f}%"}
                  for sec,g in df.groupby(sect_col_s) if len(g)>0]
            st.dataframe(pd.DataFrame(summ),use_container_width=True,hide_index=True)
        xlsx_cpv_s=io.BytesIO()
        T_h=list(range(1,cpv_horizon+1))
        if False:  # removed projections - see EAD tab
            pass
            SCEN_PROJ={" Optimiste":{"pct_lo":75,"pct_hi":95,"color":C["green"]},
                        " Central":  {"pct_lo":40,"pct_hi":60,"color":C["blue"]},
                        " Pessimiste":{"pct_lo":5,"pct_hi":25,"color":C["red"]}}
            fig_fan=go.Figure()
            # Fan chart
            for sc_n,sc_d in SCEN_PROJ.items():
                lo=np.percentile(PD_proj*100,sc_d["pct_lo"],axis=1)
                hi=np.percentile(PD_proj*100,sc_d["pct_hi"],axis=1)
                med=np.percentile(PD_proj*100,50,axis=1)
                fig_fan.add_trace(go.Scatter(x=T_h,y=hi.tolist(),mode="lines",
                    line=dict(color=sc_d["color"],width=0),showlegend=False))
                fig_fan.add_trace(go.Scatter(x=T_h,y=lo.tolist(),mode="lines",
                    fill="tonexty",fillcolor=f"rgba({int(sc_d['color'][1:3],16)},"
                    f"{int(sc_d['color'][3:5],16)},{int(sc_d['color'][5:7],16)},0.15)",
                    line=dict(color=sc_d["color"],width=0),name=sc_n))
                fig_fan.add_trace(go.Scatter(x=T_h,y=med.tolist(),mode="lines+markers",
                    line=dict(color=sc_d["color"],width=2.5),marker=dict(size=6),
                    showlegend=False))
            fig_fan.add_hline(y=float(df[pd_col_cpv].mean()*100),
                line_dash="dash",line_color="gray",annotation_text="PD_TTC moy")
            fig_fan.update_layout(title=f"Fan Chart PD_PIT CPV — {cpv_horizon} ans<br>"
                f"P5-25=Pessimiste | P40-60=Central | P75-95=Optimiste",
                height=480,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="Horizon T (ans)",yaxis_title="PD_PIT (%)",
                xaxis=dict(tickvals=T_h),legend=dict(orientation="h",y=-0.2))
            st.plotly_chart(fig_fan,use_container_width=True, key="chart_048")
            # Heatmap PD par secteur × horizon (médiane)
            if "SECTEUR_GICS" in df.columns or "SECTEUR" in df.columns:
                sc2_=sect_col_ if sect_col_ in df.columns else "SECTEUR"
                hm_h=[]; 
                for sec in sorted(df[sc2_].unique()):
                    g=df[df[sc2_]==sec]
                    if len(g)==0: continue
                    pd_ttc_s=float(g[pd_col_cpv].mean())
                    pd_pit_s=float(g["PD_PIT_CPV"].mean()) if "PD_PIT_CPV" in df.columns else pd_ttc_s
                    row_h={"Secteur":sec[:18]}
                    for t_i,t in enumerate(T_h):
                        med_macro=float(np.median(PD_proj[t_i]*100))
                        adj=pd_pit_s*100+(med_macro-float(df[pd_col_cpv].mean()*100))
                        row_h[f"T={t}a"]=round(adj,3)
                    hm_h.append(row_h)
                if hm_h:
                    df_hm_h=pd.DataFrame(hm_h).set_index("Secteur")
                    fig_hm_h=go.Figure(go.Heatmap(z=df_hm_h.values,x=df_hm_h.columns.tolist(),
                        y=df_hm_h.index.tolist(),colorscale="YlOrRd",
                        text=df_hm_h.round(2).values,texttemplate="%{text}%",
                        textfont=dict(size=8),colorbar=dict(title="PD%")))
                    fig_hm_h.update_layout(title="Heatmap PD_PIT CPV (%) — Secteur × Horizon (médiane)",
                        height=max(300,len(hm_h)*40+80),paper_bgcolor="white")
                    st.plotly_chart(fig_hm_h,use_container_width=True, key="chart_049")
            # Export
            xlsx_cpv_=io.BytesIO()
            df_exp_=df[[c for c in [next((c for c in ["NOM_EMETTEUR","DESCRIPTION"] if c in df.columns),""),
                sect_col_ if sect_col_ in df.columns else "",pd_col_cpv,
                "PD_PIT_CPV","delta_PD_CPV"] if c and c in df.columns]].copy()
            for c in [pd_col_cpv,"PD_PIT_CPV","delta_PD_CPV"]:
                if c in df_exp_.columns: df_exp_[c]=(df_exp_[c]*100).round(4)
            with pd.ExcelWriter(xlsx_cpv_,engine="openpyxl") as wr:
                df_exp_.to_excel(wr,sheet_name="PD_CPV",index=False)
            st.download_button("Exporter PD_CPV",data=xlsx_cpv_.getvalue(),
                file_name="PD_PIT_CPV.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_006")

with TABS[tab_lgd]:
    from scipy.stats import norm as _norm, beta as _beta_dist, kstest, shapiro
    from scipy.special import expit as _sigmoid_lgd
    from scipy.optimize import minimize as _minimize_lgd

    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>💸 LGD PIT — Vasicek & Frye-Jacobs</h2>
      <p>Calibration Beta MLE · ρ_LGD Bootstrap · X_t^LGD · LGD_PIT Vasicek · Frye-Jacobs LGD · Scénarios · Monte Carlo</p>
    </div></div>""", unsafe_allow_html=True)

    if df is None:
        st.warning(" Chargez vos données dans la sidebar."); st.stop()

    # Sidebar params LGD
    with st.sidebar:
        st.markdown("### Paramètres LGD")
        lgd_rho_method = st.radio("ρ_LGD méthode", ["MOM Historique","Bâle II"], index=0, key="lgd_rho_method_2")
        lgd_n_gh       = st.slider("Points Gauss-Hermite", 20, 200, 100, 10, key="lgd_n_gh_2")
        lgd_n_boot     = st.number_input("Bootstrap B (ρ_LGD)", 1000, 20000, 10000, 1000, key="lgd_n_boot_2")
        lgd_rho_bale   = st.slider("ρ Bâle II (si choisi)", 0.05, 0.30, 0.15, 0.01, key="lgd_rho_bale_2")

    # ── CALCULS CORE LGD ──────────────────────────────────────────────────────
    @st.cache_data(show_spinner=False)
    def compute_lgd_core(oblig_json, n_boot, rho_method, rho_bale_val, n_gh, _v="v4_obs_z"):
        df_o = pd.read_json(io.StringIO(oblig_json))
        pd_col = "PD_TTC" if "PD_TTC" in df_o.columns else "PD_decimal"
        if pd_col not in df_o.columns:
            df_o["PD_TTC"] = pd.to_numeric(df_o.get("PD",0),errors="coerce")/100
            pd_col = "PD_TTC"
        lgd_col = "LGD_TTC" if "LGD_TTC" in df_o.columns else "LGD"
        if lgd_col == "LGD" and "LGD" in df_o.columns:
            df_o["LGD_TTC"] = pd.to_numeric(df_o["LGD"],errors="coerce").clip(0.005,0.995)
            lgd_col = "LGD_TTC"
        elif lgd_col not in df_o.columns:
            df_o["LGD_TTC"] = 0.45; lgd_col = "LGD_TTC"

        df_o[pd_col]  = pd.to_numeric(df_o[pd_col], errors="coerce").clip(1e-6,1-1e-6)
        df_o[lgd_col] = pd.to_numeric(df_o[lgd_col],errors="coerce").clip(0.005,0.995)
        df_o["EAD avec PD PIT"]   = pd.to_numeric(df_o.get("NOMINAL",pd.Series([1e6]*len(df_o))),errors="coerce").fillna(1e6)

        # Statistiques LGD TTC
        lgd_vals = df_o[lgd_col].values
        mu_lgd   = float(lgd_vals.mean())
        s2_lgd   = float(np.var(lgd_vals,ddof=1))
        cv_lgd   = float(lgd_vals.std()/mu_lgd) if mu_lgd>0 else 0.5

        # Beta MLE sur LGD_TTC
        d_mom = max(mu_lgd*(1-mu_lgd)/max(s2_lgd,1e-8)-1, 0.05)
        a_mom, b_mom = mu_lgd*d_mom, (1-mu_lgd)*d_mom

        def neg_ll(params):
            a,b = params
            if a<=0 or b<=0: return 1e10
            return -float(np.sum(_beta_dist.logpdf(lgd_vals.clip(1e-6,1-1e-6),a,b)))

        try:
            res_mle = _minimize_lgd(neg_ll,[a_mom,b_mom],method="Nelder-Mead",
                                     options={"xatol":1e-9,"fatol":1e-9,"maxiter":10000})
            A_MLE,B_MLE = float(res_mle.x[0]),float(res_mle.x[1])
        except:
            A_MLE,B_MLE = a_mom,b_mom

        ll_mom = float(np.sum(_beta_dist.logpdf(lgd_vals.clip(1e-6,1-1e-6),a_mom,b_mom)))
        ll_mle = float(np.sum(_beta_dist.logpdf(lgd_vals.clip(1e-6,1-1e-6),A_MLE,B_MLE)))
        aic_mom = -2*ll_mom+4; aic_mle = -2*ll_mle+4
        ks_s,ks_p = kstest(lgd_vals, lambda x: _beta_dist.cdf(x,A_MLE,B_MLE))
        sw_s,sw_p = shapiro(lgd_vals[:min(50,len(lgd_vals))])

        # ρ_LGD
        if rho_method == "MOM Historique":
            y_lgd = np.array([float(_norm.ppf(_beta_dist.cdf(v,A_MLE,B_MLE).clip(1e-6,1-1e-6)))
                               for v in lgd_vals])
            l0 = float(y_lgd.mean()); l1sq = float(np.mean((y_lgd-l0)**2))
            rho_lgd = l1sq/(1+l1sq)
        else:
            def rho_bale(pd_v):
                w=(1-np.exp(-50*pd_v))/(1-np.exp(-50))
                return float(0.12*w+0.24*(1-w))
            rho_lgd = float(np.mean([rho_bale(p) for p in df_o[pd_col].values]))

        # Bootstrap IC ρ_LGD
        np.random.seed(42)
        y_boot = np.array([float(_norm.ppf(_beta_dist.cdf(v,A_MLE,B_MLE).clip(1e-6,1-1e-6)))
                            for v in lgd_vals])
        rb_lgd=[]
        for _ in range(n_boot):
            idx=np.random.choice(len(y_boot),len(y_boot),replace=True)
            y_b=y_boot[idx]; l0b=y_b.mean(); l1sq_b=np.mean((y_b-l0b)**2)
            rb_lgd.append(l1sq_b/(1+l1sq_b))
        rb_lgd=np.array(rb_lgd)
        rho_lo=float(np.percentile(rb_lgd,2.5)); rho_hi=float(np.percentile(rb_lgd,97.5))

        # Gauss-Hermite
        from scipy.special import roots_hermite
        z_gh,w_gh=roots_hermite(n_gh)
        w_norm=w_gh/np.sqrt(np.pi); w_nodes=np.sqrt(2)*z_gh

        def calib_beta_ind(mu_i, cv=cv_lgd):
            """Beta individuelle : E[Beta_i] = LGD_TTC_i  σ_i = CV × LGD_TTC_i"""
            mu_i = np.clip(mu_i, 0.005, 0.995)
            sig  = np.clip(cv * mu_i, 0.02, 0.40)
            d    = max(mu_i*(1-mu_i)/max(sig**2,1e-8) - 1, 0.05)
            return max(0.05, mu_i*d), max(0.05, (1-mu_i)*d)

        def lgd_pit_gh(lgd_i, X_t, rho=rho_lgd):
            """E[LGD_i|X_t] via Beta INDIVIDUELLE ancrée sur LGD_TTC_i"""
            a_i, b_i = calib_beta_ind(float(lgd_i))
            Y = -np.sqrt(rho)*X_t + np.sqrt(1-rho)*w_nodes
            u = _norm.cdf(Y).clip(1e-6, 1-1e-6)
            return float(np.clip(np.dot(w_norm, _beta_dist.ppf(u, a_i, b_i)), 0.005, 0.995))

        # X_t^LGD depuis LGD_TTC via Beta MLE
        lam0 = float(np.mean(y_boot))
        lam1 = float(np.sqrt(np.mean((y_boot-lam0)**2)))
        X_t_lgd = float((np.mean(y_boot)-lam0)/(lam1+1e-10))

        # ── Frye-Jacobs (notebook LGD_CPV__1_.ipynb) ─────────────────────────
        # Formule exacte du notebook:
        # 1. Extraire Z observe par obligation depuis PD_PIT (etape2_extraire_Z)
        #    Z = (phi_inv(PD_TTC) - sqrt(1-rho)*phi_inv(PD_PIT)) / sqrt(rho)
        # 2. LGD_PIT = (1/PD_PIT(Z)) * phi( phi_inv(PD_PIT) - delta )
        #    delta = (phi_inv(PD_TTC) - phi_inv(p)) / sqrt(1-rho)  avec p=PD_TTC*LGD_TTC
        #
        # PD_PIT: utiliser la colonne CPV si disponible, sinon Vasicek neutre

        def _phi_lgd(x): return _norm.cdf(np.clip(x,-30,30))
        def _phi_inv_lgd(p): return _norm.ppf(np.clip(p,1e-8,1-1e-8))

        def pd_pit_z_lgd(pd_ttc, rho, Z):
            return float(_phi_lgd((_phi_inv_lgd(float(pd_ttc))-np.sqrt(rho)*Z)/np.sqrt(1-rho)))

        def lgd_pit_fj_vec(pd_ttc_arr, lgd_ttc_arr, rho_arr, Z_arr):
            """Vectorise: LGD_PIT Frye-Jacobs exact per notebook."""
            pd_ttc = np.clip(pd_ttc_arr.astype(float), 1e-8, 1-1e-8)
            lgd_ttc= np.clip(lgd_ttc_arr.astype(float), 1e-8, 1-1e-8)
            rho    = np.clip(rho_arr.astype(float), 0.01, 0.40)
            Z      = Z_arr.astype(float)
            p      = np.clip(pd_ttc * lgd_ttc, 1e-8, 1-1e-8)
            phi_pd = _norm.ppf(pd_ttc)
            phi_p  = _norm.ppf(p)
            delta  = (phi_pd - phi_p) / np.sqrt(1-rho)
            pd_pit = np.clip(
                _norm.cdf((phi_pd - np.sqrt(rho)*Z) / np.sqrt(1-rho)),
                1e-8, 1-1e-8)
            lgd_pit= np.clip(
                _norm.cdf(_norm.ppf(pd_pit) - delta) / pd_pit,
                0.005, 0.999)
            return lgd_pit

        # Recuperer PD_PIT observe par obligation (CPV ou Vasicek neutre)
        pd_pit_col = next((c for c in ["PD_PIT_CPV","PD_PIT_NEUTRE","PD_PIT"] if c in df_o.columns), None)
        pd_ttc_arr = df_o[pd_col].values.astype(float)
        lgd_ttc_arr= df_o[lgd_col].values.astype(float)
        rho_per_ob = np.array([
            float(0.12*((1-np.exp(-50*p))/(1-np.exp(-50)))+0.24*(1-((1-np.exp(-50*p))/(1-np.exp(-50)))))
            for p in pd_ttc_arr])

        if pd_pit_col is not None:
            pd_pit_obs = np.clip(df_o[pd_pit_col].values.astype(float), 1e-8, 1-1e-8)
        else:
            # Fallback: utiliser PD_TTC comme proxy (Z_obs ≈ 0)
            pd_pit_obs = pd_ttc_arr.copy()

        # Z observe par obligation (etape2_extraire_Z du notebook)
        Z_obs = (np.array([_phi_inv_lgd(p) for p in pd_ttc_arr])
                 - np.sqrt(1-rho_per_ob) * np.array([_phi_inv_lgd(p) for p in pd_pit_obs])
                ) / np.sqrt(rho_per_ob)
        Z_obs = np.where(np.isfinite(Z_obs), Z_obs, 0.0)
        df_o["Z_obs"] = Z_obs

        # LGD_PIT principal (Z observe = resultat du tableau notebook)
        df_o["LGD_PIT_FJ"] = lgd_pit_fj_vec(pd_ttc_arr, lgd_ttc_arr, rho_per_ob, Z_obs)

        # Scenarios (Monte Carlo comme notebook etape7)
        np.random.seed(42)
        _Z_sim_lgd = np.random.standard_normal(10000)
        SCEN_LGD = {
            "Pessimiste": {"Z": float(np.percentile(_Z_sim_lgd,  5)), "w": 0.20, "color": "#A71E1E"},
            "Central":    {"Z": float(np.percentile(_Z_sim_lgd, 50)), "w": 0.70, "color": "#2E4D7B"},
            "Optimiste":  {"Z": float(np.percentile(_Z_sim_lgd, 95)), "w": 0.10, "color": "#1A6B2E"},
        }

        for sc, d in SCEN_LGD.items():
            Z_sc_arr = np.full(len(df_o), d["Z"])
            # Vasicek Gauss-Hermite
            df_o[f"LGD_PIT_V_{sc}"] = df_o[lgd_col].apply(
                lambda v: lgd_pit_gh(float(v), X_t_lgd + d["Z"]*0.5))
            # Frye-Jacobs scenario
            df_o[f"LGD_PIT_FJ_{sc}"] = lgd_pit_fj_vec(pd_ttc_arr, lgd_ttc_arr, rho_per_ob, Z_sc_arr)

        # Ponderation scenarios (weighted average for capital computation)
        df_o["LGD_PIT_V_Pond"] = sum(d["w"]*df_o[f"LGD_PIT_V_{sc}"] for sc,d in SCEN_LGD.items())
        df_o["LGD_PIT_FJ_Pond"]= sum(d["w"]*df_o[f"LGD_PIT_FJ_{sc}"] for sc,d in SCEN_LGD.items())
        df_o["delta_LGD_V"]    = df_o["LGD_PIT_V_Pond"]  - df_o[lgd_col]
        df_o["delta_LGD_FJ"]   = df_o["LGD_PIT_FJ_Pond"] - df_o[lgd_col]

        return {"df":df_o.to_json(),"A_MLE":A_MLE,"B_MLE":B_MLE,"a_mom":a_mom,"b_mom":b_mom,
                "aic_mom":aic_mom,"aic_mle":aic_mle,"ll_mle":ll_mle,"ll_mom":ll_mom,
                "rho_lgd":rho_lgd,"rho_lo":rho_lo,"rho_hi":rho_hi,"rb_lgd":rb_lgd.tolist(),
                "lam0":lam0,"lam1":lam1,"X_t_lgd":X_t_lgd,"ks_p":ks_p,"sw_p":sw_p,
                "mu_lgd":mu_lgd,"cv_lgd":cv_lgd,"SCEN":list(SCEN_LGD.keys()),
                "SCEN_CFG":{k:{"Z":v["Z"],"color":v["color"]} for k,v in SCEN_LGD.items()},
                "pd_col":pd_col,"lgd_col":lgd_col}

    with st.spinner("Calibration LGD (Beta MLE + Gauss-Hermite)…"):
        lgd_res = compute_lgd_core(df.to_json(), lgd_n_boot, lgd_rho_method, lgd_rho_bale, lgd_n_gh, "v4_obs_z")

    df_lgd   = pd.read_json(io.StringIO(lgd_res["df"]))
    A_MLE    = lgd_res["A_MLE"]; B_MLE = lgd_res["B_MLE"]
    RHO_LGD  = lgd_res["rho_lgd"]; X_T_LGD = lgd_res["X_t_lgd"]
    pd_col   = lgd_res["pd_col"]; lgd_col = lgd_res["lgd_col"]
    SCEN_CFG = lgd_res["SCEN_CFG"]

    sub_lgd = st.tabs([
        "📐 Densité Beta LGD",
        "ρ_LGD & X_t",
        "LGD_PIT Vasicek",
        "🔗 Frye-Jacobs",
        "Tableau par Obligation",
    ])

    x_rng = np.linspace(0.001,0.999,400)

    # ── SUB 0 : Densité Beta ─────────────────────────────────────────────────
    with sub_lgd[0]:
        st.subheader("Estimation Densité LGD — Beta MLE vs Moments vs KDE")
        k1,k2,k3,k4 = st.columns(4)
        k1.metric("α Beta (MLE)",f"{A_MLE:.4f}")
        k2.metric("β Beta (MLE)",f"{B_MLE:.4f}")
        k3.metric("E[Beta]",f"{A_MLE/(A_MLE+B_MLE)*100:.2f}%")
        k4.metric("AIC MLE",f"{lgd_res['aic_mle']:.2f}",
                   delta=f"vs MOM={lgd_res['aic_mom']:.2f}",delta_color="inverse")

        col_d1,col_d2 = st.columns(2)
        with col_d1:
            # Histogramme + 3 courbes densité
            from scipy.stats import gaussian_kde as _kde
            kde_fn = _kde(df_lgd[lgd_col].values, bw_method="silverman")
            fig_dns=go.Figure()
            fig_dns.add_trace(go.Histogram(x=df_lgd[lgd_col].tolist(),nbinsx=25,
                histnorm="probability density",
                marker=dict(color="#E8EFF8",line=dict(color="#9EB3D4",width=0.8)),
                name="Empirique",opacity=0.75))
            fig_dns.add_trace(go.Scatter(x=x_rng.tolist(),
                y=_beta_dist.pdf(x_rng,lgd_res["a_mom"],lgd_res["b_mom"]).tolist(),
                mode="lines",name=f"Beta Moments (AIC={lgd_res['aic_mom']:.2f})",
                line=dict(color=C["teal"],width=2,dash="dash")))
            fig_dns.add_trace(go.Scatter(x=x_rng.tolist(),
                y=_beta_dist.pdf(x_rng,A_MLE,B_MLE).tolist(),
                mode="lines",name=f"Beta MLE  (AIC={lgd_res['aic_mle']:.2f})",
                line=dict(color=C["red"],width=3)))
            fig_dns.add_trace(go.Scatter(x=x_rng.tolist(),
                y=kde_fn.evaluate(x_rng).tolist(),
                mode="lines",name="KDE Silverman",
                line=dict(color=C["purple"],width=2,dash="dot")))
            fig_dns.add_vline(x=lgd_res["mu_lgd"],line_color=C["navy"],line_width=2,
                               line_dash="dash",annotation_text=f"μ={lgd_res['mu_lgd']:.3f}")
            fig_dns.update_layout(title="Densité LGD — 3 estimateurs",
                height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="LGD",yaxis_title="Densité",
                legend=dict(orientation="h",y=-0.28))
            st.plotly_chart(fig_dns,use_container_width=True, key="chart_050")

        with col_d2:
            # Q-Q plot + résidus
            lgd_s = np.sort(df_lgd[lgd_col].values)
            n_q   = len(lgd_s)
            p_emp = (np.arange(1,n_q+1)-0.375)/(n_q+0.25)
            q_th  = _beta_dist.ppf(p_emp,A_MLE,B_MLE)
            fig_qq=go.Figure()
            fig_qq.add_trace(go.Scatter(x=q_th.tolist(),y=lgd_s.tolist(),
                mode="markers",marker=dict(color=C["blue"],size=7,
                line=dict(color="white",width=0.5)),name="Quantiles obs."))
            lm=max(q_th.max(),lgd_s.max())*1.05
            fig_qq.add_shape(type="line",x0=0,x1=lm,y0=0,y1=lm,
                line=dict(color=C["red"],dash="dash",width=2))
            fig_qq.update_layout(title=f"Q-Q Plot — Beta MLE<br>KS p={lgd_res['ks_p']:.4f}  SW p={lgd_res['sw_p']:.4f}",
                height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="Quantiles théoriques Beta MLE",yaxis_title="Quantiles empiriques")
            st.plotly_chart(fig_qq,use_container_width=True, key="chart_051")

        # CDF comparaison
        cdf_emp=np.arange(1,n_q+1)/n_q
        fig_cdf=go.Figure()
        fig_cdf.add_trace(go.Scatter(x=lgd_s.tolist(),y=cdf_emp.tolist(),
            mode="lines",name="CDF empirique",line=dict(color=C["navy"],width=2.5)))
        fig_cdf.add_trace(go.Scatter(x=x_rng.tolist(),
            y=_beta_dist.cdf(x_rng,A_MLE,B_MLE).tolist(),
            mode="lines",name="CDF Beta MLE",line=dict(color=C["red"],width=2.5)))
        fig_cdf.add_trace(go.Scatter(x=x_rng.tolist(),
            y=_beta_dist.cdf(x_rng,lgd_res["a_mom"],lgd_res["b_mom"]).tolist(),
            mode="lines",name="CDF Beta Moments",line=dict(color=C["teal"],width=1.8,dash="dash")))
        fig_cdf.update_layout(title="CDF Empirique vs Beta MLE",
            height=300,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
            xaxis_title="LGD",yaxis_title="Probabilité cumulée",
            legend=dict(orientation="h",y=-0.35))
        st.plotly_chart(fig_cdf,use_container_width=True, key="chart_052")

        # Tableau comparatif
        tbl_dns=pd.DataFrame([
            {"Modèle":"Beta Moments","α/μ":f"{lgd_res['a_mom']:.4f}","β/σ":f"{lgd_res['b_mom']:.4f}",
             "Log-L":f"{lgd_res['ll_mom']:.4f}","AIC":f"{lgd_res['aic_mom']:.4f}","Retenu?":"Non"},
            {"Modèle":"Beta MLE ","α/μ":f"{A_MLE:.4f}","β/σ":f"{B_MLE:.4f}",
             "Log-L":f"{lgd_res['ll_mle']:.4f}","AIC":f"{lgd_res['aic_mle']:.4f}","Retenu?":" OUI"},
        ])
        st.dataframe(tbl_dns,use_container_width=True,hide_index=True)

    # ── SUB 1 : ρ_LGD & X_t ─────────────────────────────────────────────────
    with sub_lgd[1]:
        st.subheader("Calibration ρ_LGD (MOM + Bootstrap) & Facteur X_t^{LGD}")
        k1,k2,k3,k4 = st.columns(4)
        k1.metric("ρ_LGD",f"{RHO_LGD*100:.2f}%",
                   delta=f"IC [{lgd_res['rho_lo']*100:.1f}%,{lgd_res['rho_hi']*100:.1f}%]")
        k2.metric("λ̂₀_LGD",f"{lgd_res['lam0']:+.4f}")
        k3.metric("λ̂₁_LGD",f"{lgd_res['lam1']:.4f}")
        k4.metric("X_t^LGD",f"{X_T_LGD:+.4f}",
                   delta="Adverse ↑" if X_T_LGD>0.5 else("Favorable ↓" if X_T_LGD<-0.5 else "Neutre"))

        col_r1,col_r2 = st.columns(2)
        with col_r1:
            rb=np.array(lgd_res["rb_lgd"])
            fig_rb=go.Figure()
            fig_rb.add_trace(go.Histogram(x=(rb*100).tolist(),nbinsx=70,
                marker=dict(color=C["blue"],line=dict(color="white",width=0.3)),opacity=0.75))
            fig_rb.add_vline(x=RHO_LGD*100,line_color=C["red"],line_width=3,
                annotation_text=f"ρ̂={RHO_LGD*100:.2f}%")
            fig_rb.add_vrect(x0=lgd_res["rho_lo"]*100,x1=lgd_res["rho_hi"]*100,
                fillcolor="rgba(0,163,224,0.10)",layer="below",annotation_text="IC 95%")
            fig_rb.update_layout(title=f"Bootstrap ρ_LGD (B={lgd_n_boot:,})",
                height=360,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="ρ_LGD (%)",showlegend=False)
            st.plotly_chart(fig_rb,use_container_width=True, key="chart_053")

        with col_r2:
            # Sensibilité LGD_PIT à ρ_LGD
            rho_range=np.linspace(0.05,0.70,100)
            fig_sens=go.Figure()
            mu_lgd_v=lgd_res["mu_lgd"]
            for sc_n,sc_d in SCEN_CFG.items():
                X_s=X_T_LGD+sc_d["Z"]*0.5
                lgd_curve=[]
                for r in rho_range:
                    from scipy.special import roots_hermite as _rh
                    z_g,w_g=_rh(50); w_n=w_g/np.sqrt(np.pi); z_n=np.sqrt(2)*z_g
                    Y=-np.sqrt(r)*X_s+np.sqrt(1-r)*z_n
                    u=_norm.cdf(Y).clip(1e-6,1-1e-6)
                    lgd_curve.append(float(np.dot(w_n,_beta_dist.ppf(u,A_MLE,B_MLE))))
                fig_sens.add_trace(go.Scatter(x=(rho_range*100).tolist(),y=[v*100 for v in lgd_curve],
                    mode="lines",name=sc_n,line=dict(color=sc_d["color"],width=2)))
            fig_sens.add_vline(x=RHO_LGD*100,line_color=C["navy"],line_dash="dash",line_width=2.5,
                annotation_text=f"ρ retenu={RHO_LGD:.3f}")
            fig_sens.add_hline(y=mu_lgd_v*100,line_color="gray",line_dash="dot",
                annotation_text=f"LGD_TTC moy={mu_lgd_v*100:.1f}%")
            fig_sens.update_layout(title="Sensibilité LGD_PIT à ρ_LGD",
                height=360,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="ρ_LGD (%)",yaxis_title="LGD_PIT (%)",
                legend=dict(orientation="h",y=-0.25))
            st.plotly_chart(fig_sens,use_container_width=True, key="chart_054")

    # ── SUB 2 : LGD_PIT Vasicek ──────────────────────────────────────────────
    with sub_lgd[2]:
        st.subheader("LGD_PIT Vasicek — Intégration Gauss-Hermite")
        k1,k2,k3,k4 = st.columns(4)
        k1.metric("LGD_TTC moy",  f"{df_lgd[lgd_col].mean()*100:.3f}%")
        k2.metric("LGD_PIT (Pond)",f"{df_lgd['LGD_PIT_V_Pond'].mean()*100:.3f}%",
                   delta=f"{(df_lgd['LGD_PIT_V_Pond'].mean()-df_lgd[lgd_col].mean())*100:+.3f}pp")
        k3.metric("LGD_PIT min",  f"{df_lgd['LGD_PIT_V_Pond'].min()*100:.2f}%")
        k4.metric("LGD_PIT max",  f"{df_lgd['LGD_PIT_V_Pond'].max()*100:.2f}%")

        col_v1,col_v2=st.columns(2)
        with col_v1:
            # Scatter LGD_TTC vs LGD_PIT_Vasicek
            fig_sv=go.Figure()
            sect_col_lgd="SECTEUR_GICS" if "SECTEUR_GICS" in df_lgd.columns else "SECTEUR"
            clrs_lgd=px.colors.qualitative.Plotly
            if sect_col_lgd in df_lgd.columns:
                for i_s,sec in enumerate(df_lgd[sect_col_lgd].unique()):
                    g=df_lgd[df_lgd[sect_col_lgd]==sec]
                    fig_sv.add_trace(go.Scatter(x=(g[lgd_col]*100).tolist(),
                        y=(g["LGD_PIT_V_Pond"]*100).tolist(),mode="markers",name=sec[:14],
                        marker=dict(size=7,opacity=0.75,color=clrs_lgd[i_s%len(clrs_lgd)],
                        line=dict(color="white",width=0.5))))
            else:
                fig_sv.add_trace(go.Scatter(x=(df_lgd[lgd_col]*100).tolist(),
                    y=(df_lgd["LGD_PIT_V_Pond"]*100).tolist(),mode="markers",
                    marker=dict(size=7,color=C["blue"])))
            lm_v=max(df_lgd[lgd_col].max(),df_lgd["LGD_PIT_V_Pond"].max())*105
            fig_sv.add_shape(type="line",x0=0,x1=lm_v,y0=0,y1=lm_v,
                line=dict(dash="dash",color="gray"))
            fig_sv.update_layout(title="LGD_TTC vs LGD_PIT Vasicek",
                height=400,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="LGD_TTC (%)",yaxis_title="LGD_PIT Vasicek (%)",
                legend=dict(orientation="h",y=-0.28))
            st.plotly_chart(fig_sv,use_container_width=True, key="chart_055")

        with col_v2:
            # Heatmap LGD par secteur × scénario
            hm_v=[]
            if sect_col_lgd in df_lgd.columns:
                for sec in df_lgd[sect_col_lgd].unique():
                    g=df_lgd[df_lgd[sect_col_lgd]==sec]
                    row_h={"Secteur":sec[:16],"LGD_TTC":round(g[lgd_col].mean()*100,3)}
                    for sc in lgd_res["SCEN"]:
                        row_h[f"V_{sc}"]=round(g[f"LGD_PIT_V_{sc}"].mean()*100,3)
                    row_h["V_Pond"]=round(g["LGD_PIT_V_Pond"].mean()*100,3)
                    hm_v.append(row_h)
                df_hmv=pd.DataFrame(hm_v).set_index("Secteur")
                fig_hmv=go.Figure(go.Heatmap(z=df_hmv.values,x=df_hmv.columns.tolist(),
                    y=df_hmv.index.tolist(),colorscale="YlOrRd",
                    text=df_hmv.round(2).values,texttemplate="%{text}%",
                    textfont=dict(size=9),colorbar=dict(title="LGD%")))
                fig_hmv.update_layout(title="LGD_PIT Vasicek par Secteur × Scénario",
                    height=380,paper_bgcolor="white")
                st.plotly_chart(fig_hmv,use_container_width=True, key="chart_056")

        # Profil toutes obligations
        df_lgd_s=df_lgd.sort_values(lgd_col).reset_index(drop=True)
        fig_prof=go.Figure()
        fig_prof.add_trace(go.Scatter(x=list(range(len(df_lgd_s))),
            y=(df_lgd_s[lgd_col]*100).tolist(),mode="lines",name="LGD_TTC",
            line=dict(color="black",width=1.8,dash="dash")))
        for sc_n,sc_d in SCEN_CFG.items():
            fig_prof.add_trace(go.Scatter(x=list(range(len(df_lgd_s))),
                y=(df_lgd_s[f"LGD_PIT_V_{sc_n}"]*100).tolist(),mode="lines",name=f"Vasicek {sc_n}",
                line=dict(color=sc_d["color"],width=1.8)))
        fig_prof.add_trace(go.Scatter(x=list(range(len(df_lgd_s))),
            y=(df_lgd_s["LGD_PIT_V_Pond"]*100).tolist(),mode="lines",name=" Pondéré",
            line=dict(color=C["purple"],width=3)))
        fig_prof.update_layout(title="Profil LGD_PIT Vasicek — Obligations triées par LGD_TTC",
            height=380,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
            xaxis_title="Obligations",yaxis_title="LGD (%)",
            legend=dict(orientation="h",y=-0.28))
        st.plotly_chart(fig_prof,use_container_width=True, key="chart_057")

    # ── SUB 3 : Frye-Jacobs ──────────────────────────────────────────────────
    with sub_lgd[3]:
        st.subheader("LGD_PIT Frye-Jacobs — Formule Exacte")
        st.info("""**Formule Frye-Jacobs :**
        p = PD_TTC × LGD_TTC  |  δ = [Φ⁻¹(PD_TTC) − Φ⁻¹(p)] / √(1−ρ)
        **LGD_PIT(Z) = (1/PD_PIT(Z)) × Φ[Φ⁻¹(PD_PIT(Z)) − δ]**
        Propriété : E[PD_PIT × LGD_PIT] = EL_TTC = PD_TTC × LGD_TTC""")

        k1,k2,k3,k4 = st.columns(4)
        k1.metric("LGD_TTC moy",    f"{df_lgd[lgd_col].mean()*100:.3f}%")
        k2.metric("LGD_FJ (Pond)",  f"{df_lgd['LGD_PIT_FJ_Pond'].mean()*100:.3f}%",
                   delta=f"{(df_lgd['LGD_PIT_FJ_Pond'].mean()-df_lgd[lgd_col].mean())*100:+.3f}pp")
        k3.metric("LGD_FJ min",     f"{df_lgd['LGD_PIT_FJ_Pond'].min()*100:.2f}%")
        k4.metric("LGD_FJ max",     f"{df_lgd['LGD_PIT_FJ_Pond'].max()*100:.2f}%")

        col_f1,col_f2 = st.columns(2)
        with col_f1:
            # Courbe analytique LGD_PIT(Z) pour différentes PD
            Z_range=np.linspace(-3,3,200)
            rho_m=RHO_LGD; lgd_ref=lgd_res["mu_lgd"]
            def lgd_fj_v(pd_ttc,lgd_ttc,rho,Z):
                p=np.clip(pd_ttc*lgd_ttc,1e-8,1-1e-8)
                pd_p=np.clip(_norm.cdf((_norm.ppf(pd_ttc)-np.sqrt(rho)*Z)/np.sqrt(1-rho)),1e-8,1-1e-8)
                delta=(_norm.ppf(pd_ttc)-_norm.ppf(p))/np.sqrt(1-rho)
                return np.clip(_norm.cdf(_norm.ppf(pd_p)-delta)/pd_p,0.005,0.999)

            fig_fj=go.Figure()
            pd_vals_fj=[0.01,0.03,0.05,0.10,0.20]
            clrs_fj=["#185FA5","#1D9E75","#EF9F27","#E24B4A","#534AB7"]
            for pd_v,col_v in zip(pd_vals_fj,clrs_fj):
                fig_fj.add_trace(go.Scatter(x=Z_range.tolist(),
                    y=(lgd_fj_v(pd_v,lgd_ref,rho_m,Z_range)*100).tolist(),
                    mode="lines",name=f"PD_TTC={pd_v:.0%}",
                    line=dict(color=col_v,width=2)))
            fig_fj.add_hline(y=lgd_ref*100,line_color="gray",line_dash="dot",
                annotation_text=f"LGD_TTC={lgd_ref*100:.1f}%")
            fig_fj.add_vline(x=0,line_color="gray",line_dash="dash")
            fig_fj.update_layout(title="LGD_PIT Frye-Jacobs(Z) par PD_TTC<br>Z<0=récession ↑LGD | Z>0=expansion ↓LGD",
                height=400,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="Z (facteur systémique)",yaxis_title="LGD_PIT FJ (%)",
                legend=dict(orientation="h",y=-0.28))
            st.plotly_chart(fig_fj,use_container_width=True, key="chart_058")

        with col_f2:
            # Scatter LGD_TTC vs LGD_PIT_FJ
            fig_sfj=go.Figure()
            if sect_col_lgd in df_lgd.columns:
                for i_s,sec in enumerate(df_lgd[sect_col_lgd].unique()):
                    g=df_lgd[df_lgd[sect_col_lgd]==sec]
                    fig_sfj.add_trace(go.Scatter(x=(g[lgd_col]*100).tolist(),
                        y=(g["LGD_PIT_FJ_Pond"]*100).tolist(),mode="markers",name=sec[:14],
                        marker=dict(size=7,opacity=0.75,color=clrs_lgd[i_s%len(clrs_lgd)],
                        line=dict(color="white",width=0.5))))
            lm_fj=max(df_lgd[lgd_col].max(),df_lgd["LGD_PIT_FJ_Pond"].max())*105
            fig_sfj.add_shape(type="line",x0=0,x1=lm_fj,y0=0,y1=lm_fj,
                line=dict(dash="dash",color="gray"))
            fig_sfj.update_layout(title="LGD_TTC vs LGD_PIT Frye-Jacobs",
                height=400,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
                xaxis_title="LGD_TTC (%)",yaxis_title="LGD_PIT FJ (%)",
                legend=dict(orientation="h",y=-0.28))
            st.plotly_chart(fig_sfj,use_container_width=True, key="chart_059")


        # ── Projections Frye-Jacobs par scénario ────────────────────────────────────
        st.subheader("Projections LGD_PIT Frye-Jacobs — Scénarios Z sur T=1..5 ans")
        T_fj=[1,2,3,4,5]
        Z_scen_fj={" Optimiste":+1.5," Central":0.0," Pessimiste":-1.5}
        fig_fjp=go.Figure()
        rho_m_fj=RHO_LGD
        # Courbes par scénario pour le portefeuille moyen
        pd_moy_fj=float(df_lgd[pd_col].mean()); lgd_moy_fj=float(df_lgd[lgd_col].mean())
        for sc_fj,(sc_n,Z_v) in enumerate(Z_scen_fj.items()):
            col_fj=[C["green"],C["blue"],C["red"]][sc_fj]
            # LGD_FJ varie avec Z dans le temps
            lgd_t=[]; pd_pit_t=[]
            for t in T_fj:
                Z_t=Z_v*min(1.0,t/3.0)  # convergence progressive
                p_t=np.clip(_norm.cdf((_norm.ppf(pd_moy_fj)-np.sqrt(rho_m_fj)*Z_t)/np.sqrt(1-rho_m_fj)),1e-8,1-1e-8)
                p_ttc=np.clip(pd_moy_fj*lgd_moy_fj,1e-8,1-1e-8)
                delta_fj=(_norm.ppf(pd_moy_fj)-_norm.ppf(p_ttc))/np.sqrt(1-rho_m_fj)
                lgd_t_val=np.clip(_norm.cdf(_norm.ppf(p_t)-delta_fj)/p_t,0.005,0.999)
                lgd_t.append(float(lgd_t_val)); pd_pit_t.append(float(p_t))
            fig_fjp.add_trace(go.Scatter(x=T_fj,y=[v*100 for v in lgd_t],
                mode="lines+markers",name=f"LGD {sc_n}",
                line=dict(color=col_fj,width=2.5),marker=dict(size=7)))
        fig_fjp.add_hline(y=lgd_moy_fj*100,line_color="gray",line_dash="dash",
            annotation_text=f"LGD_TTC moy={lgd_moy_fj*100:.1f}%")
        fig_fjp.add_hline(y=float(df_lgd["LGD_PIT_FJ_Pond"].mean()*100),
            line_color=C["navy"],line_dash="dot",
            annotation_text=f"LGD_FJ_Pond={df_lgd['LGD_PIT_FJ_Pond'].mean()*100:.2f}%")
        fig_fjp.update_layout(title="Trajectoire LGD_PIT Frye-Jacobs(Z_t) — Projection 5 ans<br>"
            f"Z<0=récession ↑LGD | Z>0=expansion ↓LGD | ρ={rho_m_fj:.3f}",
            height=420,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
            xaxis_title="Horizon T (ans)",yaxis_title="LGD_PIT FJ (%)",
            xaxis=dict(tickvals=T_fj),legend=dict(orientation="h",y=-0.22))
        st.plotly_chart(fig_fjp,use_container_width=True, key="chart_060")
        # Heatmap LGD_FJ par secteur × scénario
        if sect_col_lgd in df_lgd.columns:
            hm_fj=[]; Z_vals_hm={"Opt":+1.5,"Cent":0.0,"Pess":-1.5}
            def lgd_fj_hm(pd_v,lgd_v,rho,Z):
                p=np.clip(pd_v*lgd_v,1e-8,1-1e-8)
                pd_p=np.clip(_norm.cdf((_norm.ppf(pd_v)-np.sqrt(rho)*Z)/np.sqrt(1-rho)),1e-8,1-1e-8)
                delta=(_norm.ppf(pd_v)-_norm.ppf(p))/np.sqrt(1-rho)
                return float(np.clip(_norm.cdf(_norm.ppf(pd_p)-delta)/pd_p,0.005,0.999))
            for sec in df_lgd[sect_col_lgd].unique():
                g=df_lgd[df_lgd[sect_col_lgd]==sec]
                if len(g)==0: continue
                row_fj={"Secteur":sec[:16],"LGD_TTC":round(g[lgd_col].mean()*100,3)}
                for sc_lbl,Z_v in Z_vals_hm.items():
                    lgd_v=float(g[lgd_col].mean()); pd_v=float(g[pd_col].mean())
                    row_fj[f"LGD_{sc_lbl}"]=round(lgd_fj_hm(pd_v,lgd_v,rho_m_fj,Z_v)*100,3)
                hm_fj.append(row_fj)
            df_hm_fj=pd.DataFrame(hm_fj).set_index("Secteur")
            fig_hm_fj=go.Figure(go.Heatmap(z=df_hm_fj.values,x=df_hm_fj.columns.tolist(),
                y=df_hm_fj.index.tolist(),colorscale="YlOrRd",
                text=df_hm_fj.round(2).values,texttemplate="%{text}%",
                textfont=dict(size=9),colorbar=dict(title="LGD%")))
            fig_hm_fj.update_layout(title="Heatmap LGD_PIT Frye-Jacobs par Secteur × Scénario",
                height=max(300,len(hm_fj)*40+80),paper_bgcolor="white")
            st.plotly_chart(fig_hm_fj,use_container_width=True, key="chart_061")
        # Comparaison Vasicek vs Frye-Jacobs
        df_lgd_s=df_lgd.sort_values(lgd_col).reset_index(drop=True)
        fig_comp=go.Figure()
        fig_comp.add_trace(go.Scatter(x=list(range(len(df_lgd_s))),
            y=(df_lgd_s[lgd_col]*100).tolist(),mode="lines",name="LGD_TTC",
            line=dict(color="black",width=1.5,dash="dash")))
        fig_comp.add_trace(go.Scatter(x=list(range(len(df_lgd_s))),
            y=(df_lgd_s["LGD_PIT_V_Pond"]*100).tolist(),mode="lines",name="LGD Vasicek ",
            line=dict(color=C["blue"],width=2.5)))
        fig_comp.add_trace(go.Scatter(x=list(range(len(df_lgd_s))),
            y=(df_lgd_s["LGD_PIT_FJ_Pond"]*100).tolist(),mode="lines",name="LGD Frye-Jacobs ",
            line=dict(color=C["red"],width=2.5,dash="dot")))
        fig_comp.update_layout(title="Comparaison LGD_PIT : Vasicek vs Frye-Jacobs",
            height=360,paper_bgcolor="white",plot_bgcolor="#F8F9FC",
            xaxis_title="Obligations (triées)",yaxis_title="LGD (%)",
            legend=dict(orientation="h",y=-0.28))
        st.plotly_chart(fig_comp,use_container_width=True, key="chart_062")

    # ── SUB 4 : Tableau par obligation ───────────────────────────────────────
    with sub_lgd[4]:
        st.subheader("LGD_PIT par Obligation — Vasicek & Frye-Jacobs")

        # Colonnes à afficher
        desc_col="NOM_EMETTEUR" if "NOM_EMETTEUR" in df_lgd.columns else "DESCRIPTION"
        sect_col="SECTEUR_GICS" if "SECTEUR_GICS" in df_lgd.columns else "SECTEUR"
        cols_tbl=[c for c in [desc_col,sect_col,lgd_col,"LGD_PIT_V_Pond",
                               "delta_LGD_V","LGD_PIT_FJ_Pond","delta_LGD_FJ",pd_col] if c in df_lgd.columns]
        df_tbl=df_lgd[cols_tbl].copy()
        for c in [lgd_col,"LGD_PIT_V_Pond","delta_LGD_V","LGD_PIT_FJ_Pond","delta_LGD_FJ",pd_col]:
            if c in df_tbl.columns:
                df_tbl[c]=(df_tbl[c]*100).round(4).astype(str)+"%"
        df_tbl.columns=[c.replace(lgd_col,"LGD_TTC%").replace("LGD_PIT_V_Pond","LGD_PIT_Vasicek%")
                          .replace("delta_LGD_V","ΔLGD_V(pp)").replace("LGD_PIT_FJ_Pond","LGD_PIT_FJ%")
                          .replace("delta_LGD_FJ","ΔLGD_FJ(pp)").replace(pd_col,"PD_TTC%")
                          .replace(desc_col,"Emetteur").replace(sect_col,"Secteur")
                          for c in df_tbl.columns]
        st.dataframe(df_tbl,use_container_width=True,hide_index=True,height=500)

        # Export
        xlsx_lgd=io.BytesIO()
        with pd.ExcelWriter(xlsx_lgd,engine="openpyxl") as wr:
            df_lgd.to_excel(wr,sheet_name="LGD_PIT",index=False)
        st.download_button("Exporter LGD_PIT (Excel)",data=xlsx_lgd.getvalue(),
            file_name="IFRS9_LGD_PIT.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key="dl_005")



with TABS[tab_ead]:
    import math as _math_ead
    import io as _io_ead

    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>EAD — Exposition au Defaut avec PD PIT (CPV)</h2>
      <p>EAD_t = EAD_(t-1) x (1 - PD_PIT_t) x (1 - AM_t) | PD_PIT = sigmoid(logit(PD_TTC) + delta_cycle)</p>
    </div></div>""", unsafe_allow_html=True)

    if df is None:
        st.warning("Chargez vos donnees dans la sidebar."); st.stop()

    # =========================================================================
    # PARAMETRES SIDEBAR
    # =========================================================================
    with st.sidebar:
        st.markdown("--- \nParametres EAD")
        ead_horizon = st.slider("Horizon de projection (ans)", 1, 10, 5, key="ead_hz_pit")
        delta_pess  = st.number_input("delta Pessimiste", -2.0, 2.0, 0.420, 0.01, key="ead_dp")
        delta_cent  = st.number_input("delta Central",    -2.0, 2.0,-0.252, 0.01, key="ead_dc")
        delta_opti  = st.number_input("delta Optimiste",  -2.0, 2.0,-0.850, 0.01, key="ead_do")

    HZON = int(ead_horizon)
    T_AX = list(range(HZON + 1))
    DELTA = {"Pessimiste": delta_pess, "Central": delta_cent, "Optimiste": delta_opti}
    SC_COL = {"Pessimiste": "#E24B4A", "Central": "#EF9F27", "Optimiste": "#1D9E75"}
    MC_COL = {"ZC": "#185FA5", "AA": "#E24B4A", "TA": "#1D9E75"}
    SCENARIO_REF = "Central"

    # =========================================================================
    # FORMULES EXACTES DU NOTEBOOK
    # =========================================================================
    def _sigmoid(x):
        return 1.0 / (1.0 + np.exp(-np.clip(x, -30, 30)))

    def _logit(p):
        p = np.clip(p, 1e-7, 1 - 1e-7)
        return np.log(p / (1 - p))

    def _pd_pit(pd_ttc, delta):
        """PD_PIT = sigmoid( logit(PD_TTC) + delta_cycle )"""
        return float(_sigmoid(_logit(pd_ttc) + delta))

    def _ead0(meth, mat_res, mat_tot, nominal):
        """ZC/TA : NOMINAL | AA : NOMINAL x (mat_res / mat_tot)"""
        m = str(meth).strip().upper()
        if mat_res <= 0: return 0.0
        if m in ("ZC", "TA"): return float(nominal)
        if pd.isna(mat_tot) or mat_tot <= 0: return float(nominal)
        return float(nominal) * (float(mat_res) / float(mat_tot))

    def _am_t(meth, mat_res, t):
        """ZC/TA : 0 (IN FINE) | AA : 1/N_restant"""
        if float(t) > float(mat_res): return 1.0
        m = str(meth).strip().upper()
        if m in ("ZC", "TA"): return 0.0
        return 1.0 / max(float(mat_res) - float(t) + 1, 1.0)

    # =========================================================================
    # PREPARATION DONNEES
    # =========================================================================
    @st.cache_data(show_spinner=False)
    @st.cache_data(show_spinner=False)
    def prep_ead_pit(js):
        d = pd.read_json(io.StringIO(js))
        d.columns = [c.strip() for c in d.columns]
        # PD TTC
        pd_col = next((c for c in ["PD_TTC","PD_decimal"] if c in d.columns), None)
        if pd_col:
            d["PD_TTC"] = pd.to_numeric(d[pd_col], errors="coerce").clip(1e-6, 1-1e-6)
        else:
            d["PD_TTC"] = pd.to_numeric(d.get("PD",0), errors="coerce") / 100
        d["PD_TTC"] = d["PD_TTC"].clip(1e-6, 1-1e-6)
        d["NOMINAL"] = pd.to_numeric(d.get("NOMINAL",1e5), errors="coerce").fillna(1e5)
        # Maturites
        mc = next((c for c in ["MATURITE_RESIDUELLE","Maturite (ans)","Maturite (ans)","mat_res"] if c in d.columns), None)
        if mc is None:
            mc = next((c for c in d.columns if "aturit" in c and "ans" in c.lower()), None)
        d["mat_res"] = pd.to_numeric(d[mc], errors="coerce").fillna(5.0).clip(lower=0) if mc else 5.0
        # Maturite totale DATE_JOUISSANCE -> DATE_ECHEANCE
        for col in ["DATE_JOUISSANCE","DATE_ECHEANCE","DATE_EMISSION"]:
            if col in d.columns:
                try:    d[col] = pd.to_datetime(d[col], unit="ms", errors="coerce")
                except: d[col] = pd.to_datetime(d[col], errors="coerce")
        if "DATE_JOUISSANCE" in d.columns and "DATE_ECHEANCE" in d.columns:
            d["mat_tot"] = (d["DATE_ECHEANCE"] - d["DATE_JOUISSANCE"]).dt.days / 365.25
        elif "mat_tot_precomp" in d.columns:
            d["mat_tot"] = pd.to_numeric(d["mat_tot_precomp"], errors="coerce")
        else:
            d["mat_tot"] = None
        d["ann_ec"] = (d["mat_tot"] - d["mat_res"]).clip(lower=0) if d["mat_tot"] is not None else 0
        if "METHODE_VALO" not in d.columns: d["METHODE_VALO"] = "ZC"
        d["METHODE_VALO"] = d["METHODE_VALO"].astype(str).str.strip().str.upper()
        d["horizon"] = np.ceil(d["mat_res"]).astype(int).clip(lower=0)
        # EAD_0
        d["EAD_0"] = d.apply(lambda r: _ead0(
            r["METHODE_VALO"], float(r["mat_res"]),
            float(r["mat_tot"]) if pd.notna(r.get("mat_tot")) else None,
            float(r["NOMINAL"])), axis=1)
        d["cap_amor"] = (d["NOMINAL"] - d["EAD_0"]).clip(lower=0)
        return d.reset_index(drop=True)

    with st.spinner("Calcul EAD_0 et PD PIT..."):
        _df_prep = df.copy()
        for _dc in ["DATE_JOUISSANCE","DATE_ECHEANCE","DATE_EMISSION"]:
            if _dc in _df_prep.columns:
                _df_prep[_dc] = pd.to_datetime(_df_prep[_dc], errors="coerce")
        if "DATE_JOUISSANCE" in _df_prep.columns and "DATE_ECHEANCE" in _df_prep.columns:
            _dd = (_df_prep["DATE_ECHEANCE"] - _df_prep["DATE_JOUISSANCE"]).dt.days
            _df_prep["mat_tot_precomp"] = (_dd / 365.25).clip(lower=0.1)
        df_ead = prep_ead_pit(_df_prep.to_json())
    # Validation: ensure all 328 obligations are present
    if len(df_ead) < len(df):
        st.error(f"ATTENTION: EAD calcule sur {len(df_ead)} obligations au lieu de {len(df)}. Rechargez les donnees.")
        st.stop()
    # Show row count
    st.caption(f"EAD calcule sur {len(df_ead)} obligations | ZC:{(df_ead['METHODE_VALO']=='ZC').sum()} | AA:{(df_ead['METHODE_VALO']=='AA').sum()} | TA:{(df_ead['METHODE_VALO']=='TA').sum()}")

    # Calcul PD PIT par scenario
    for sc_name, delta in DELTA.items():
        df_ead[f"PD_PIT_{sc_name}"] = df_ead["PD_TTC"].apply(lambda p: _pd_pit(p, delta))
    df_ead["PD_PIT"] = df_ead[f"PD_PIT_{SCENARIO_REF}"]

    # =========================================================================
    # PROJECTIONS EAD PAR SCENARIO
    # =========================================================================
    @st.cache_data(show_spinner=False)
    @st.cache_data(show_spinner=False)
    def proj_ead_pit(ead_json, hz, d_pess, d_cent, d_opti):
        """Projection EAD vectorisee — numpy arrays, pas de boucle Python par obligation."""
        d = pd.read_json(io.StringIO(ead_json))
        deltas = {"Pessimiste": d_pess, "Central": d_cent, "Optimiste": d_opti}
        n = len(d)
        results = {}
        # Pre-calculer les taux d amortissement pour chaque (methode, mat_res, t)
        meth_arr  = d["METHODE_VALO"].astype(str).str.strip().str.upper().values
        mr_arr    = d["mat_res"].values.astype(float)
        ead0_arr  = d["EAD_0"].values.astype(float)
        hz_arr    = np.ceil(mr_arr).astype(int).clip(max=hz)
        # AM matrix: shape (n, hz)
        am_matrix = np.zeros((n, hz), dtype=float)
        for t in range(1, hz + 1):
            for j in range(n):
                if t > mr_arr[j]:
                    am_matrix[j, t-1] = 1.0
                elif meth_arr[j] in ("ZC", "TA"):
                    am_matrix[j, t-1] = 0.0
                else:  # AA
                    am_matrix[j, t-1] = 1.0 / max(mr_arr[j] - t + 1.0, 1.0)

        for sc_name, delta in deltas.items():
            # PD PIT vectorise
            pd_ttc = d["PD_TTC"].values.astype(float)
            pd_pit = np.array([_pd_pit(float(p), delta) for p in pd_ttc])
            # Projection EAD annee par annee — array operations
            ead_matrix = np.zeros((n, hz + 1), dtype=float)
            ead_matrix[:, 0] = ead0_arr
            for t in range(1, hz + 1):
                am_t = am_matrix[:, t-1]
                ead_prev = ead_matrix[:, t-1]
                ead_matrix[:, t] = np.maximum(0.0, ead_prev * (1 - pd_pit) * (1 - am_t))
            # Build result DataFrame
            df_res = pd.DataFrame({"i": np.arange(n), "EAD_0": np.round(ead0_arr, 2)})
            for t in range(1, hz + 1):
                df_res[f"EAD_{t}"] = np.round(ead_matrix[:, t], 2)
            results[sc_name] = df_res
        return results

    with st.spinner(f"Projections EAD {HZON} ans x 3 scenarios..."):
        proj = proj_ead_pit(df_ead.to_json(), HZON, DELTA["Pessimiste"], DELTA["Central"], DELTA["Optimiste"])

    # Totaux EAD par annee
    max_hz = int(df_ead["horizon"].max())
    ead_tot = {}
    for sc_name in DELTA:
        df_sc = proj[sc_name]
        ead_tot[sc_name] = [df_ead["EAD_0"].sum()]
        for t in range(1, HZON + 1):
            col = f"EAD_{t}"
            val = df_sc[col].sum() if col in df_sc.columns else 0
            ead_tot[sc_name].append(val)

    NE_COL = next((c for c in ["Description complète","Description complete","NOM_EMETTEUR","DESCRIPTION"] if c in df_ead.columns), "")
    SC_COL_EAD = next((c for c in ["SECTEUR_GICS","Secteur","SECTEUR"] if c in df_ead.columns), "")
    ISIN_COL = next((c for c in ["CODE_ISIN","CODE"] if c in df_ead.columns), "")

    # =========================================================================
    # KPI GLOBAUX
    # =========================================================================
    tot_ead = df_ead["EAD_0"].sum(); tot_nom = df_ead["NOMINAL"].sum()
    pd_ttc_m = df_ead["PD_TTC"].mean(); pd_pit_m = df_ead["PD_PIT"].mean()

    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric("EAD_0 Total",       f"{tot_ead/1e6:.3f} M MAD")
    k2.metric("NOMINAL Total",     f"{tot_nom/1e6:.3f} M MAD")
    k3.metric("Ratio EAD/NOM",     f"{tot_ead/tot_nom:.1%}")
    k4.metric("PD TTC moy.",       f"{pd_ttc_m:.4%}")
    k5.metric(f"PD PIT Central", f"{pd_pit_m:.4%}",
               delta=f"{(pd_pit_m-pd_ttc_m)*100:+.3f}pp")

    # =========================================================================
    # SOUS-ONGLETS
    # =========================================================================
    sub_ead = st.tabs([
        "EAD par Methode",
        "PD TTC vs PD PIT",
        "Projection Portefeuille",
        "Profil par Obligation",
        "Export Excel",
    ])

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 0 : EAD PAR METHODE
    # ─────────────────────────────────────────────────────────────────────────
    with sub_ead[0]:
        st.subheader("EAD_0 — Capital Restant Du par Methode")
        st.info(
            "ZC (Zero Coupon) : IN FINE — EAD_0 = NOMINAL — AM_t = 0\n"
            "TA (Taux Amortissable) : IN FINE — EAD_0 = NOMINAL — AM_t = 0\n"
            "AA (Amortissement Annuel) : EAD_0 = NOMINAL x mat_res/mat_tot — AM_t = 1/N_restant"
        )
        # Cartes methode
        cols_m = st.columns(3)
        for i_m,(meth,colm) in enumerate(MC_COL.items()):
            g = df_ead[df_ead["METHODE_VALO"]==meth]
            with cols_m[i_m]:
                if len(g)==0: st.info(f"Aucune obligation {meth}"); continue
                r = g["EAD_0"].sum()/g["NOMINAL"].sum() if g["NOMINAL"].sum()>0 else 0
                pd_m = g["PD_PIT"].mean(); pd_ttc = g["PD_TTC"].mean()
                st.markdown(
                    f'<div style="background:{colm};color:white;padding:16px;border-radius:10px">'
                    f'<div style="font-size:0.9rem;font-weight:800">{meth} — {len(g)} obligations</div>'
                    f'<div style="height:1px;background:rgba(255,255,255,0.3);margin:8px 0"></div>'
                    f'<div style="display:grid;grid-template-columns:auto 1fr;gap:4px 12px;font-size:0.80rem">'
                    f'<span style="opacity:0.8">EAD_0</span><b>{g["EAD_0"].sum()/1e6:.4f} M MAD</b>'
                    f'<span style="opacity:0.8">NOMINAL</span><b>{g["NOMINAL"].sum()/1e6:.4f} M MAD</b>'
                    f'<span style="opacity:0.8">Ratio CRD/NOM</span><b>{r:.2%}</b>'
                    f'<span style="opacity:0.8">PD TTC moy.</span><b>{pd_ttc:.4%}</b>'
                    f'<span style="opacity:0.8">PD PIT Central</span><b>{pd_m:.4%}</b>'
                    f'</div></div>', unsafe_allow_html=True)

        col_e1,col_e2 = st.columns(2)
        with col_e1:
            fig_bar = go.Figure()
            for meth,colm in MC_COL.items():
                g = df_ead[df_ead["METHODE_VALO"]==meth]
                if len(g)==0: continue
                fig_bar.add_trace(go.Bar(name=f"{meth} NOM", x=[meth], y=[g["NOMINAL"].sum()/1e6],
                    marker_color=colm, opacity=0.22,
                    text=f"{g['NOMINAL'].sum()/1e6:.3f}M", textposition="outside"))
                fig_bar.add_trace(go.Bar(name=f"{meth} EAD_0", x=[meth], y=[g["EAD_0"].sum()/1e6],
                    marker_color=colm, opacity=0.92,
                    text=f"{g['EAD_0'].sum()/1e6:.3f}M", textposition="outside"))
            fig_bar.update_layout(title="EAD_0 (CRD) vs NOMINAL par Methode (M MAD)",
                barmode="group", height=380, paper_bgcolor="white",
                plot_bgcolor="#F5F6FA", showlegend=False)
            st.plotly_chart(fig_bar, use_container_width=True, key="ead_pit_bar")
        with col_e2:
            if SC_COL_EAD:
                ead_sec = df_ead.groupby(SC_COL_EAD)["EAD_0"].sum().sort_values(ascending=False)/1e6
                fig_hsec = go.Figure(go.Bar(x=ead_sec.values.tolist(), y=ead_sec.index.tolist(),
                    orientation="h", marker_color=px.colors.qualitative.Plotly[:len(ead_sec)],
                    opacity=0.88, text=[f"{v:.3f}M" for v in ead_sec.values], textposition="outside"))
                fig_hsec.update_layout(title="EAD_0 par Secteur (M MAD)", height=380,
                    paper_bgcolor="white", plot_bgcolor="#F5F6FA", showlegend=False,
                    xaxis_title="M MAD")
                st.plotly_chart(fig_hsec, use_container_width=True, key="ead_pit_sect")

        # Scatter maturite vs EAD_0
        fig_sc = go.Figure()
        for meth,colm in MC_COL.items():
            g = df_ead[df_ead["METHODE_VALO"]==meth]
            if len(g)==0: continue
            hover_txt = [str(g.iloc[j].get(NE_COL,""))[:30] for j in range(len(g))] if NE_COL else []
            fig_sc.add_trace(go.Scatter(x=g["mat_res"].tolist(), y=g["EAD_0"].tolist(),
                mode="markers", name=f"{meth} ({len(g)})",
                text=hover_txt,
                hovertemplate="%{text}<br>Mat=%{x:.2f}a<br>EAD=%{y:,.0f} MAD",
                marker=dict(size=8, color=colm, opacity=0.75, line=dict(color="white",width=0.5))))
        fig_sc.update_layout(
            title="Maturite Residuelle vs EAD_0 (MAD) — ZC/TA constant | AA decroissant",
            height=360, paper_bgcolor="white", plot_bgcolor="#F5F6FA",
            xaxis_title="Maturite residuelle (ans)", yaxis_title="EAD_0 (MAD)",
            legend=dict(orientation="h", y=-0.25))
        st.plotly_chart(fig_sc, use_container_width=True, key="ead_pit_scatter_mat")

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 1 : PD TTC vs PD PIT
    # ─────────────────────────────────────────────────────────────────────────
    with sub_ead[1]:
        st.subheader("PD TTC vs PD PIT — sigmoid( logit(PD_TTC) + delta_cycle )")
        st.info(
            "delta > 0 : cycle defavorable — PD_PIT > PD_TTC (recession)\n"
            "delta < 0 : cycle favorable  — PD_PIT < PD_TTC (expansion)\n"
            "delta = 0 : PD_PIT = PD_TTC (cycle neutre)"
        )
        # PD moyennes par scenario
        pd_m_rows=[]
        for sc,delta in DELTA.items():
            pm=df_ead[f"PD_PIT_{sc}"].mean()
            pd_m_rows.append({"Scenario":sc,"delta":delta,
                              "PD_TTC moy":f"{pd_ttc_m:.4%}",
                              "PD_PIT moy":f"{pm:.4%}",
                              "Delta PD":f"{(pm-pd_ttc_m)*100:+.4f}pp"})
        st.dataframe(pd.DataFrame(pd_m_rows), use_container_width=True, hide_index=True)

        col_p1,col_p2 = st.columns(2)
        with col_p1:
            fig_sc2 = go.Figure()
            for sc,col_sc in SC_COL.items():
                fig_sc2.add_trace(go.Scatter(
                    x=(df_ead["PD_TTC"]*100).tolist(),
                    y=(df_ead[f"PD_PIT_{sc}"]*100).tolist(),
                    mode="markers", name=sc,
                    marker=dict(size=6, color=col_sc, opacity=0.65)))
            mn=0; mx=df_ead["PD_TTC"].max()*100*1.05
            fig_sc2.add_trace(go.Scatter(x=[mn,mx],y=[mn,mx],mode="lines",
                line=dict(color="gray",dash="dash",width=1),name="PD_PIT = PD_TTC"))
            fig_sc2.update_layout(title="PD TTC vs PD PIT par Scenario",
                height=380,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                xaxis_title="PD TTC (%)",yaxis_title="PD PIT (%)",
                legend=dict(orientation="h",y=-0.28))
            st.plotly_chart(fig_sc2,use_container_width=True,key="ead_pit_sc2")
        with col_p2:
            # Distribution delta PD
            fig_hd=go.Figure()
            for sc,col_sc in SC_COL.items():
                delta_pd=(df_ead[f"PD_PIT_{sc}"]-df_ead["PD_TTC"])*100
                fig_hd.add_trace(go.Histogram(x=delta_pd.tolist(),name=sc,
                    marker_color=col_sc,opacity=0.60,nbinsx=30))
            fig_hd.add_vline(x=0,line_color="#333",line_dash="dot")
            fig_hd.update_layout(title="Distribution Ajustement Cyclique (PD PIT - PD TTC)",
                barmode="overlay",height=380,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                xaxis_title="Delta PD (%)",yaxis_title="Nb obligations",
                legend=dict(orientation="h",y=-0.28))
            st.plotly_chart(fig_hd,use_container_width=True,key="ead_pit_hd")

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 2 : PROJECTION PORTEFEUILLE
    # ─────────────────────────────────────────────────────────────────────────
    with sub_ead[2]:
        st.subheader(f"Projection EAD Portefeuille — {HZON} ans")
        col_p1,col_p2=st.columns(2)
        with col_p1:
            fig_ead_sc=go.Figure()
            fig_ead_sc.add_hline(y=tot_nom/1e6,line_color="#263238",line_dash="dot",
                annotation_text=f"NOMINAL = {tot_nom/1e6:.1f}M")
            for sc in ["Pessimiste","Central","Optimiste"]:
                vals=[ead_tot[sc][t]/1e6 for t in range(HZON+1)]
                fig_ead_sc.add_trace(go.Scatter(x=T_AX,y=vals,mode="lines+markers",
                    name=f"{sc} (PD_PIT={df_ead[f'PD_PIT_{sc}'].mean():.3%})",
                    line=dict(color=SC_COL[sc],width=2.5),marker=dict(size=8),
                    fill="tonexty" if sc=="Pessimiste" else None))
                for t,v in zip(T_AX,vals):
                    fig_ead_sc.add_annotation(x=t,y=v,text=f"{v:.1f}M",
                        showarrow=False,font=dict(size=8,color=SC_COL[sc]),
                        yshift=10 if sc=="Optimiste" else (-14 if sc=="Pessimiste" else 0))
            fig_ead_sc.update_layout(title="EAD Total Portefeuille (M MAD)",
                height=400,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                xaxis_title="Annees",yaxis_title="EAD (M MAD)",
                xaxis=dict(tickvals=T_AX),legend=dict(orientation="h",y=-0.3))
            st.plotly_chart(fig_ead_sc,use_container_width=True,key="ead_pit_port_sc")
        with col_p2:
            # Erosion relative
            fig_pct=go.Figure()
            for sc in ["Pessimiste","Central","Optimiste"]:
                ref=ead_tot[sc][0]
                vals_p=[ead_tot[sc][t]/ref*100 for t in range(HZON+1)]
                fig_pct.add_trace(go.Scatter(x=T_AX,y=vals_p,mode="lines+markers",
                    name=sc,line=dict(color=SC_COL[sc],width=2.5),marker=dict(size=8)))
                for t,v in zip(T_AX,vals_p):
                    fig_pct.add_annotation(x=t,y=v,text=f"{v:.1f}%",
                        showarrow=False,font=dict(size=8,color=SC_COL[sc]),
                        yshift=8 if sc=="Optimiste" else (-12 if sc=="Pessimiste" else 0))
            fig_pct.add_hline(y=100,line_color="#263238",line_dash="dot")
            fig_pct.update_layout(title="Erosion EAD (%  de EAD_0)",
                height=400,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                xaxis_title="Annees",yaxis_title="EAD_t / EAD_0 (%)",
                xaxis=dict(tickvals=T_AX),legend=dict(orientation="h",y=-0.3))
            st.plotly_chart(fig_pct,use_container_width=True,key="ead_pit_pct")

        # EAD par methode — scenario central
        fig_meth=go.Figure()
        for meth,colm in MC_COL.items():
            mask=(df_ead["METHODE_VALO"]==meth)
            g=df_ead[mask]
            if len(g)==0: continue
            df_sc_c=proj[SCENARIO_REF]
            vals=[g["EAD_0"].sum()/1e6]
            for t in T_AX[1:]:
                col=f"EAD_{t}"
                vals.append(df_sc_c.loc[mask.values,col].sum()/1e6 if col in df_sc_c.columns else 0)
            fig_meth.add_trace(go.Scatter(x=T_AX,y=vals,mode="lines+markers",
                name=f"{meth} ({len(g)} oblig.)",
                line=dict(color=colm,width=2.5),marker=dict(size=8)))
        fig_meth.update_layout(title=f"EAD par Methode — Scenario {SCENARIO_REF} (M MAD)",
            height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
            xaxis_title="Annees",yaxis_title="EAD (M MAD)",
            xaxis=dict(tickvals=T_AX),legend=dict(orientation="h",y=-0.25))
        st.plotly_chart(fig_meth,use_container_width=True,key="ead_pit_meth_cent")

        # Tableau recapitulatif
        st.subheader("Tableau EAD Total — 3 Scenarios")
        recap_rows=[]
        for t in T_AX:
            vp=ead_tot["Pessimiste"][t]; vc=ead_tot["Central"][t]; vo=ead_tot["Optimiste"][t]
            recap_rows.append({"Annee":t,
                "Pessimiste (MAD)":f"{vp:,.0f}","Central (MAD)":f"{vc:,.0f}",
                "Optimiste (MAD)":f"{vo:,.0f}",
                "Ecart Pess-Opt":f"{vp-vo:+,.0f}"})
        st.dataframe(pd.DataFrame(recap_rows),use_container_width=True,hide_index=True)

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 3 : PROFIL PAR OBLIGATION
    # ─────────────────────────────────────────────────────────────────────────
    with sub_ead[3]:
        st.subheader("Profil EAD — Obligation Individuelle")

        # Filtres par methode
        col_flt1, col_flt2 = st.columns([1, 3])
        with col_flt1:
            meth_filter = st.selectbox("Filtrer par methode :",
                options=["Toutes"] + sorted(df_ead["METHODE_VALO"].unique().tolist()),
                key="ead_pit_meth_flt")
        df_ead_flt = df_ead if meth_filter=="Toutes" else df_ead[df_ead["METHODE_VALO"]==meth_filter]
        with col_flt2:
            if NE_COL:
                all_labels_flt=[f"{str(df_ead_flt.iloc[i].get(NE_COL,''))[:65]} [{df_ead_flt.iloc[i]['METHODE_VALO']} | {df_ead_flt.iloc[i]['mat_res']:.1f}a]"
                                for i in range(len(df_ead_flt))]
            else:
                all_labels_flt=[f"Oblig {df_ead_flt.index[i]} [{df_ead_flt.iloc[i]['METHODE_VALO']}]"
                                for i in range(len(df_ead_flt))]
            if len(df_ead_flt)==0:
                st.warning("Aucune obligation pour cette methode."); st.stop()
            flt_idx_local = st.selectbox(f"Choisir parmi {len(df_ead_flt)} obligations :",
                options=list(range(len(df_ead_flt))),
                format_func=lambda i: all_labels_flt[i],
                key="ead_pit_oblig_sel")
        idx_ob = df_ead_flt.index[flt_idx_local]  # indice dans df_ead original
        all_labels = [f"{str(df_ead.iloc[i].get(NE_COL,''))[:65]} [{df_ead.iloc[i]['METHODE_VALO']}]"
                     for i in range(len(df_ead))] if NE_COL else [f"Oblig {i}" for i in range(len(df_ead))]

        row_ob=df_ead.loc[idx_ob]
        hz_ob=int(row_ob["horizon"]); meth_ob=str(row_ob["METHODE_VALO"])
        pd_ttc_ob=float(row_ob["PD_TTC"]); ead0_ob=float(row_ob["EAD_0"])
        nom_ob=float(row_ob["NOMINAL"]); mat_r_ob=float(row_ob["mat_res"])

        # Info cards
        c1,c2,c3=st.columns(3)
        c1.markdown(
            f'<div style="background:#1B1F6B;color:white;padding:14px;border-radius:10px">'
            f'<div style="font-size:0.7rem;opacity:0.8;letter-spacing:0.05em">OBLIGATION</div>'
            f'<div style="font-size:0.82rem;font-weight:700;margin-top:4px">{str(row_ob.get(NE_COL,""))[:45]}</div>'
            f'<div style="display:grid;grid-template-columns:auto 1fr;gap:3px 8px;font-size:0.78rem;margin-top:8px">'
            f'<span style="opacity:0.7">Methode</span><b>{meth_ob}</b>'
            f'<span style="opacity:0.7">Maturite</span><b>{mat_r_ob:.2f} ans</b>'
            f'<span style="opacity:0.7">NOMINAL</span><b>{nom_ob:,.0f} MAD</b>'
            f'</div></div>', unsafe_allow_html=True)
        c2.markdown(
            f'<div style="background:#009FE3;color:white;padding:14px;border-radius:10px">'
            f'<div style="font-size:0.7rem;opacity:0.8;letter-spacing:0.05em">PD</div>'
            f'<div style="display:grid;grid-template-columns:auto 1fr;gap:3px 8px;font-size:0.80rem;margin-top:8px">'
            f'<span style="opacity:0.8">PD TTC</span><b>{pd_ttc_ob*100:.4f}%</b>'
            f'<span style="opacity:0.8">PD PIT Pess.</span><b>{float(row_ob.get("PD_PIT_Pessimiste",0))*100:.4f}%</b>'
            f'<span style="opacity:0.8">PD PIT Cent.</span><b>{float(row_ob.get("PD_PIT",0))*100:.4f}%</b>'
            f'<span style="opacity:0.8">PD PIT Opti.</span><b>{float(row_ob.get("PD_PIT_Optimiste",0))*100:.4f}%</b>'
            f'</div></div>', unsafe_allow_html=True)
        c3.markdown(
            f'<div style="background:#0F6E56;color:white;padding:14px;border-radius:10px">'
            f'<div style="font-size:0.7rem;opacity:0.8;letter-spacing:0.05em">EAD</div>'
            f'<div style="display:grid;grid-template-columns:auto 1fr;gap:3px 8px;font-size:0.80rem;margin-top:8px">'
            f'<span style="opacity:0.8">EAD_0 (CRD)</span><b>{ead0_ob:,.0f} MAD</b>'
            f'<span style="opacity:0.8">Cap. amorti</span><b>{float(row_ob.get("cap_amor",0)):,.0f} MAD</b>'
            f'<span style="opacity:0.8">Ratio CRD/NOM</span><b>{ead0_ob/nom_ob:.2%}</b>'
            f'</div></div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        # Graphique EAD par scenario
        fig_ob=go.Figure()
        # Map original index to position in proj dataframe
        _pos_ob = df_ead.index.get_loc(idx_ob)
        for sc,col_sc in SC_COL.items():
            df_sc=proj[sc]
            vals=[ead0_ob]
            for t in range(1,hz_ob+1):
                col_t=f"EAD_{t}"
                v=df_sc.iloc[_pos_ob].get(col_t,0) if col_t in df_sc.columns else 0
                vals.append(float(v))
            t_ax_ob=list(range(len(vals)))
            fig_ob.add_trace(go.Scatter(x=t_ax_ob,y=vals,mode="lines+markers",
                name=f"{sc} (PD_PIT={float(row_ob.get(f'PD_PIT_{sc}',0))*100:.4f}%)",
                line=dict(color=col_sc,width=2.5),marker=dict(size=9)))
            for t,v in zip(t_ax_ob,vals):
                fig_ob.add_annotation(x=t,y=v,text=f"{v:,.0f}",
                    showarrow=False,font=dict(size=8,color=col_sc),yshift=10)
        fig_ob.update_layout(
            title=f"EAD annuel — {str(df_ead.loc[idx_ob].get(NE_COL, str(idx_ob)))[:50]}<br>"
                  f"<sup>EAD_0={ead0_ob:,.0f} MAD | PD_TTC={pd_ttc_ob*100:.4f}% | Methode={meth_ob}</sup>",
            height=400,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
            xaxis_title="Annee",yaxis_title="EAD (MAD)",
            legend=dict(orientation="h",y=-0.3))
        st.plotly_chart(fig_ob,use_container_width=True,key="ead_pit_ob_sc")

        # Tableau annuel detail
        st.subheader(f"Detail annuel — {all_labels[idx_ob][:40]}")
        _pos_ob2 = df_ead.index.get_loc(idx_ob)
        tab_ob=[]
        for sc,col_sc in SC_COL.items():
            pd_pit_v=float(row_ob.get(f"PD_PIT_{sc}",pd_ttc_ob))
            ead_t=ead0_ob
            for t in range(hz_ob+1):
                am=_am_t(meth_ob,mat_r_ob,t) if t>0 else 0
                if t==0:
                    tab_ob.append({"Annee":t,"Scenario":sc,"PD_PIT (%)":round(pd_pit_v*100,4),
                        "AM_t":0.0,"EAD (MAD)":round(ead_t,0),"Erosion (%)":100.0})
                else:
                    ead_t=max(0,ead_t*(1-pd_pit_v)*(1-am))
                tab_ob.append({"Annee":t,"Scenario":sc,"PD_PIT (%)":round(pd_pit_v*100,4),
                        "AM_t":round(am,4),"EAD (MAD)":round(ead_t,0),
                        "Erosion (%)":round(ead_t/ead0_ob*100,2)})
        df_tab_ob=pd.DataFrame(tab_ob)
        st.dataframe(df_tab_ob,use_container_width=True,hide_index=True,height=380)

    # ─────────────────────────────────────────────────────────────────────────
    # SUB 4 : ANALYSE SECTORIELLE
    # ─────────────────────────────────────────────────────────────────────────
    with sub_ead[4]:
        st.subheader("Export Excel — Format Notebook EAD__1_.ipynb")
        st.info(
            f"Fichier Excel — {len(df_ead)} obligations x 4 onglets:\n"
            "- Detail_Pessimiste / Detail_Central / Detail_Optimiste : 1 ligne/obligation + "
            "EAD_0..EAD_Tn + PD_TTC + PD_PIT\n"
            "- Resume : synthese toutes methodes + 3 scenarios"
        )
        r1,r2,r3,r4=st.columns(4)
        r1.metric("Obligations",str(len(df_ead)))
        r2.metric("EAD_0 Total",f"{df_ead['EAD_0'].sum()/1e6:.3f} M MAD")
        r3.metric("PD PIT Central",f"{df_ead['PD_PIT'].mean():.4%}")
        r4.metric("Horizon",f"{HZON} ans")

        # ── Export Excel: genere une fois, stocke en session_state ──────────────
        _ead_key = "ead_xl_{}_{}_{}".format(len(df_ead), HZON, round(DELTA.get('Central',0),3))
        col_gen1, col_gen2 = st.columns([1, 2])
        with col_gen1:
            gen_btn = st.button("Generer le fichier Excel", key="btn_ead_pit_xl",
                                use_container_width=True, type="primary")
        with col_gen2:
            if _ead_key in st.session_state:
                st.download_button(
                    "Telecharger IFRS9_EAD_portefeuille.xlsx",
                    data=st.session_state[_ead_key],
                    file_name="IFRS9_EAD_portefeuille.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_ead_pit_xl",
                )
        if gen_btn:
            prog = st.progress(0, text="Construction Excel...")
            import io as _xlio
            buf = _xlio.BytesIO()
            df_e = df_ead.copy()
            ne_e = next((c for c in ["Description complète","NOM_EMETTEUR","DESCRIPTION"] if c in df_e.columns), "")
            sc_e = next((c for c in ["SECTEUR_GICS","Secteur","SECTEUR"] if c in df_e.columns), "")
            is_e = next((c for c in ["CODE_ISIN","CODE"] if c in df_e.columns), "")
            SCEN_NAMES = ["Pessimiste","Central","Optimiste"]
            with pd.ExcelWriter(buf, engine="openpyxl") as wr:
                for s_idx, sc_name in enumerate(SCEN_NAMES):
                    prog.progress((s_idx+1)*25, text=f"Onglet {sc_name}...")
                    df_sc = proj[sc_name]
                    base = pd.DataFrame({
                        "CODE_ISIN":    df_e[is_e].astype(str) if is_e in df_e.columns else "",
                        "Secteur":      df_e[sc_e].astype(str) if sc_e in df_e.columns else "",
                        "Description":  df_e[ne_e].astype(str).str[:60] if ne_e in df_e.columns else "",
                        "METHODE_VALO": df_e["METHODE_VALO"].astype(str),
                        "mat_res (ans)":df_e["mat_res"].round(2),
                        "EAD_0 (MAD)":  df_e["EAD_0"].round(0),
                        "PD_TTC (%)":   (df_e["PD_TTC"]*100).round(4),
                        f"PD_PIT (%)":  (df_e.get(f"PD_PIT_{sc_name}", df_e["PD_TTC"])*100).round(4),
                        "delta_cycle":  round(DELTA[sc_name],4),
                    })
                    for t in range(1, HZON+1):
                        col_t = f"EAD_{t}"
                        ead_col = pd.to_numeric(df_sc.get(col_t, 0), errors="coerce").fillna(0)
                        base[f"EAD_T{t}(MAD)"] = ead_col.where(df_e["mat_res"]>=t, 0.0).round(0)
                    base.to_excel(wr, sheet_name=f"Detail_{sc_name[:5]}", index=False)
                # Resume
                prog.progress(90, text="Onglet Resume...")
                resume = pd.DataFrame({
                    "CODE_ISIN":   df_e[is_e].astype(str) if is_e in df_e.columns else "",
                    "Secteur":     df_e[sc_e].astype(str) if sc_e in df_e.columns else "",
                    "Description": df_e[ne_e].astype(str).str[:55] if ne_e in df_e.columns else "",
                    "METHODE":     df_e["METHODE_VALO"].astype(str),
                    "mat_res (a)": df_e["mat_res"].round(2),
                    "EAD_0 (MAD)": df_e["EAD_0"].round(0),
                    "PD_TTC (%)":  (df_e["PD_TTC"]*100).round(4),
                })
                for sc_name in SCEN_NAMES:
                    df_sc = proj[sc_name]
                    resume[f"PD_PIT_{sc_name[:4]}(%)"] = (df_e.get(f"PD_PIT_{sc_name}", df_e["PD_TTC"])*100).round(4)
                    resume[f"{sc_name[:4]}_T1"]  = pd.to_numeric(df_sc.get("EAD_1",0),errors="coerce").fillna(0).round(0)
                    resume[f"{sc_name[:4]}_T{HZON}"] = pd.to_numeric(df_sc.get(f"EAD_{HZON}",0),errors="coerce").fillna(0).round(0)
                resume.to_excel(wr, sheet_name="Resume", index=False)
            prog.progress(100, text="Termine!")
            st.session_state[_ead_key] = buf.getvalue()
            st.success(f"Fichier pret — {len(df_ead)} obligations · {HZON} ans · 4 onglets")
            st.rerun()

        # Apercu
        st.subheader(f"Apercu — Scenario {SCENARIO_REF} (50 premieres)")
        _isin_col2=next((c for c in ["CODE_ISIN","CODE"] if c in df_ead.columns),"")
        prev_c=[c for c in [_isin_col2,NE_COL,SC_COL_EAD,"METHODE_VALO","mat_res",
                             "cap_amor","EAD_0","PD_TTC","PD_PIT"]
                if c and c in df_ead.columns]
        df_prev2=df_ead[prev_c].head(50).copy()
        for c in ["PD_TTC","PD_PIT"]:
            if c in df_prev2.columns: df_prev2[c]=(df_prev2[c]*100).round(4)
        for c in ["EAD_0","cap_amor"]:
            if c in df_prev2.columns: df_prev2[c]=df_prev2[c].round(0)
        df_sc_prev=proj[SCENARIO_REF]
        for t in range(1,min(HZON+1,4)):
            col_t=f"EAD_{t}"
            if col_t in df_sc_prev.columns:
                df_prev2[f"EAD_T{t} (MAD)"]=df_sc_prev[col_t].head(50).round(0).values
        st.dataframe(df_prev2,use_container_width=True,hide_index=True,height=430)

with TABS[tab_ecl]:

    # ── Header ────────────────────────────────────────────────────────────
    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>ECL IFRS 9 — Expected Credit Loss</h2>
      <p>S1 = ECL 12 mois &nbsp;|&nbsp; S2/S3 = ECL Lifetime &nbsp;|&nbsp;
         PD Reversion TTC &nbsp;|&nbsp; LGD Frye-Jacobs</p>
    </div></div>""", unsafe_allow_html=True)


    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 1 — COURBE DES TAUX BAM (Nelson-Siegel-Svensson)
    # ═══════════════════════════════════════════════════════════════════════
    with st.expander("Construction Courbe des Taux BAM (Nelson-Siegel-Svensson)", expanded=True):

        _bam_col1, _bam_col2 = st.columns([1, 2])
        with _bam_col1:
            _bam_file = st.file_uploader(
                "Fichier BAM (CSV Bons du Tresor)",
                type=["csv"], key="bam_upload",
                help="Taux-de-reference-des-bons-du-tresor_*.csv"
            )

        # Build NSS curve from BAM file or hardcoded defaults
        _bam_key = "bam_nss_curve"
        if _bam_file is not None:
            # Parse uploaded BAM CSV
            try:
                from scipy.optimize import least_squares as _lsq_bam
                from datetime import datetime as _dtb, date as _dateb
                import io as _iobam

                _DATE_VAL = _dateb(2026, 5, 7)
                _lines_bam = _bam_file.read().decode("latin-1").splitlines()
                _rows_bam = []
                for _line in _lines_bam[3:]:
                    _parts = _line.strip().split(";")
                    if len(_parts) < 3: continue
                    _ech_s = _parts[0].strip().strip('"')
                    _tx_s  = _parts[2].strip().strip('"').replace("%","").replace(",",".").strip()
                    if _ech_s.upper() in ["TOTAL",""] or not _tx_s: continue
                    try:
                        _ech = _dtb.strptime(_ech_s, "%d/%m/%Y").date()
                        _tx  = float(_tx_s) / 100.0
                        _rows_bam.append({"echeance":_ech, "taux":_tx,
                                          "T":(_ech - _DATE_VAL).days/365.25})
                    except: continue

                if len(_rows_bam) >= 3:
                    _dfb = pd.DataFrame(_rows_bam).sort_values("T").reset_index(drop=True)
                    _T_obs = _dfb["T"].values
                    _R_obs = _dfb["taux"].values

                    # Bootstrap zero-coupon
                    _zs = np.zeros(len(_T_obs))
                    for _i, (_T, _r) in enumerate(zip(_T_obs, _R_obs)):
                        if _T <= 1.0:
                            _zs[_i] = _r
                        else:
                            _n_f = int(np.floor(_T)); _tc = list(range(1,_n_f+1)); _tc[-1] = _T
                            _pv = 0.0
                            for _t_c in _tc[:-1]:
                                _zt = float(np.interp(_t_c, _T_obs[:_i], _zs[:_i])) if _i>0 else _zs[0]
                                _pv += _r/(1+_zt)**_t_c
                            _dm = 1 - _pv
                            _zs[_i] = ((1+_r)/_dm)**(1/_T)-1 if _dm>0 else _r

                    # NSS calibration
                    def _nss_b(T,b0,b1,b2,b3,t1,t2):
                        T=np.atleast_1d(T).astype(float); t1=max(t1,0.01); t2=max(t2,0.01)
                        x1=T/t1; x2=T/t2
                        f1=np.where(x1>1e-6,(1-np.exp(-x1))/x1,1-x1/2)
                        f2=np.where(x2>1e-6,(1-np.exp(-x2))/x2,1-x2/2)
                        return b0+b1*f1+b2*(f1-np.exp(-x1))+b3*(f2-np.exp(-x2))

                    _res_b = _lsq_bam(
                        lambda p,T,z: _nss_b(T,*p)-z,
                        x0=[_zs[-1], _zs[0]-_zs[-1], 0.01, 0.0, 2.0, 4.0],
                        args=(_T_obs, _zs),
                        bounds=([0.001,-0.20,-0.20,-0.20,0.1,0.1],
                                [0.20, 0.20, 0.20, 0.20,30.,30.]),
                        method="trf", max_nfev=20000, ftol=1e-14)
                    _P_b = _res_b.x
                    _rmse_b = float(np.sqrt(np.mean(_res_b.fun**2)))*10000

                    def _r_zc_b(T):
                        return float(np.clip(_nss_b(np.atleast_1d([T]),*_P_b)[0],0.001,0.25))

                    st.session_state[_bam_key] = {
                        "T_obs": _T_obs.tolist(), "R_obs": _R_obs.tolist(),
                        "z_spot": _zs.tolist(), "P_NSS": _P_b.tolist(),
                        "rmse": _rmse_b, "n_pts": len(_dfb)
                    }
                    with _bam_col2:
                        st.success(f"Courbe BAM calibree — {len(_dfb)} points | RMSE NSS = {_rmse_b:.4f} bps")
                else:
                    with _bam_col2:
                        st.error("Fichier BAM invalide — moins de 3 points.")
            except Exception as _ebam:
                with _bam_col2:
                    st.error(f"Erreur lecture BAM: {_ebam}")
        else:
            with _bam_col2:
                if _bam_key in st.session_state:
                    _bd = st.session_state[_bam_key]
                    st.info(f"Courbe BAM en memoire — {_bd['n_pts']} points | RMSE = {_bd['rmse']:.4f} bps")
                else:
                    st.info("Courbe BAM par defaut (BAM 08/05/2026) utilisee pour actualisation.")

        # Visualisation courbe
        if _bam_key in st.session_state or _bam_file is None:
            # Get curve data
            if _bam_key in st.session_state:
                _bdat = st.session_state[_bam_key]
                _T_obs_p = np.array(_bdat["T_obs"])
                _R_obs_p = np.array(_bdat["R_obs"])
                _zs_p    = np.array(_bdat["z_spot"])
                _P_p     = _bdat["P_NSS"]
                _rmse_p  = _bdat["rmse"]
                def _r_zc_p(T):
                    T=np.atleast_1d(T).astype(float); t1=max(_P_p[4],0.01); t2=max(_P_p[5],0.01)
                    x1=T/t1; x2=T/t2
                    f1=np.where(x1>1e-6,(1-np.exp(-x1))/x1,1-x1/2)
                    f2=np.where(x2>1e-6,(1-np.exp(-x2))/x2,1-x2/2)
                    return np.clip(_P_p[0]+_P_p[1]*f1+_P_p[2]*(f1-np.exp(-x1))+_P_p[3]*(f2-np.exp(-x2)),0.001,0.25)
            else:
                # Hardcoded BAM default
                _bam_pts = [(0.20,0.0221),(0.28,0.0226),(0.78,0.0229),(1.37,0.0234),
                            (3.94,0.0270),(5.45,0.0282),(9.11,0.0308),(13.20,0.0343),
                            (19.27,0.0368),(28.95,0.0401)]
                _T_obs_p = np.array([x[0] for x in _bam_pts])
                _R_obs_p = np.array([x[1] for x in _bam_pts])
                _zs_p    = _R_obs_p.copy()  # approx for display
                _r_zc_p  = lambda T: np.interp(T, _T_obs_p, _R_obs_p,
                                                left=_R_obs_p[0], right=_R_obs_p[-1])
                _rmse_p  = 0.0

            # Courbe fine
            _T_fine = np.concatenate([
                np.arange(0.25, 1.0, 0.25),
                np.arange(1.0, 5.0, 0.5),
                np.arange(5.0, 31.0, 0.5)
            ])
            _Z_fine = np.array([float(np.atleast_1d(_r_zc_p(t))[0]) for t in _T_fine])
            _D_fine = 1/(1+_Z_fine)**_T_fine

            # Tableau standard
            _T_std = [0.25, 0.5, 1, 2, 3, 4, 5, 6, 7, 8, 10, 12, 15, 20, 25, 30]
            _Z_std = np.array([float(np.atleast_1d(_r_zc_p(t))[0]) for t in _T_std])
            _D_std = 1/(1+_Z_std)**np.array(_T_std)

            # TIE distribution si df disponible
            _has_tie = df is not None and "Spread émis (bps)" in df.columns

            # Graphes
            _gc1, _gc2, _gc3 = st.columns(3)

            with _gc1:
                _fig_bam = go.Figure()
                # Zone court/moyen/long terme
                for _xa,_xb,_cl,_lb in [(0,2,"#1A6B2E","CT"),(2,5,"#2E4D7B","MT"),
                                          (5,15,"#EF9F27","LT"),(15,31,"#A71E1E","TLT")]:
                    _fig_bam.add_vrect(x0=_xa,x1=_xb,fillcolor=_cl,opacity=0.05,layer="below")
                    _fig_bam.add_annotation(x=(_xa+_xb)/2,y=0.021,text=_lb,
                        font=dict(size=8,color=_cl),showarrow=False)
                # Courbe NSS
                _fig_bam.add_trace(go.Scatter(x=_T_fine,y=_Z_fine*100,
                    mode="lines",name=f"NSS (RMSE={_rmse_p:.2f}bp)",
                    line=dict(color="#1B1F6B",width=3)))
                # Points bootstrappes
                _fig_bam.add_trace(go.Scatter(x=_T_obs_p,y=_zs_p*100,
                    mode="markers",name="Taux ZC (bootstrap)",
                    marker=dict(size=10,color="#A71E1E",symbol="circle",
                                line=dict(color="white",width=1.5))))
                # Points marche
                _fig_bam.add_trace(go.Scatter(x=_T_obs_p,y=_R_obs_p*100,
                    mode="markers",name="Taux marche BAM",
                    marker=dict(size=8,color="#1D9E75",symbol="triangle-up",
                                line=dict(color="white",width=1))))
                # Annotations
                for _t,_z in zip(_T_obs_p,_zs_p):
                    _fig_bam.add_annotation(x=_t,y=_z*100,text=f"{_z*100:.3f}%",
                        showarrow=False,font=dict(size=7,color="#1B1F6B"),
                        yshift=12,xshift=5)
                _fig_bam.update_layout(
                    title="Courbe Zero-Coupon BAM — Nelson-Siegel-Svensson",
                    height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                    xaxis_title="Maturite (ans)",yaxis_title="Taux zero-coupon (%)",
                    xaxis=dict(range=[0,31]),
                    legend=dict(orientation="h",y=-0.3,font=dict(size=9)))
                st.plotly_chart(_fig_bam,use_container_width=True,key="bam_curve_main")

            with _gc2:
                # Facteurs actualisation D(t)
                _fig_disc = go.Figure()
                _fig_disc.add_trace(go.Scatter(x=_T_fine,y=_D_fine,
                    mode="lines",name="D(T) = 1/(1+r_ZC)^T",
                    line=dict(color="#185FA5",width=2.5),fill="tozeroy",
                    fillcolor="rgba(24,95,165,0.08)"))
                # Points standards
                _fig_disc.add_trace(go.Scatter(x=_T_std,y=_D_std,
                    mode="markers",name="Points standards",
                    marker=dict(size=7,color="#E24B4A"),
                    text=[f"T={t} | D={d:.4f}" for t,d in zip(_T_std,_D_std)],
                    hovertemplate="%{text}"))
                for _t,_d in zip(_T_std,_D_std):
                    if _t in [1,5,10,20,30]:
                        _fig_disc.add_annotation(x=_t,y=_d,
                            text=f"{_d:.4f}",showarrow=False,
                            font=dict(size=8,color="#185FA5"),yshift=12)
                _fig_disc.update_layout(
                    title="Facteurs d'Actualisation D(T) = 1/(1+r_ZC)^T",
                    height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                    xaxis_title="Maturite (ans)",yaxis_title="D(T)",
                    legend=dict(orientation="h",y=-0.3))
                st.plotly_chart(_fig_disc,use_container_width=True,key="bam_disc_main")

            with _gc3:
                # TIE individuel si donnees dispo
                if _has_tie and df is not None:
                    _mc_b=next((c for c in df.columns if "aturit" in c and "ans" in c.lower()),None)
                    _sp_b=next((c for c in df.columns if "pread" in c.lower()),None)
                    if _mc_b and _sp_b:
                        _mat_b=pd.to_numeric(df[_mc_b],errors="coerce").fillna(5)
                        _sp_dec=pd.to_numeric(df[_sp_b],errors="coerce").fillna(0)/10000
                        _is_bdt=(df.get("Secteur","x")=="souverain")
                        _rzc_b=np.array([float(np.atleast_1d(_r_zc_p(float(t)))[0]) for t in _mat_b])
                        _tie_b=np.where(_is_bdt,_rzc_b,np.clip(_rzc_b+_sp_dec.values,0.001,0.30))
                        _sect_b=df.get("Secteur",pd.Series(["?"]*len(df)))
                        _fig_tie=go.Figure()
                        _T_fine2=np.linspace(0.1,32,300)
                        _Z_fine2=np.array([float(np.atleast_1d(_r_zc_p(t))[0]) for t in _T_fine2])
                        _fig_tie.add_trace(go.Scatter(x=_T_fine2,y=_Z_fine2*100,
                            mode="lines",name="Courbe BDT (spread=0)",
                            line=dict(color="#1B1F6B",width=2.5),zorder=3))
                        _sects_b=_sect_b.unique()
                        _cpals=px.colors.qualitative.Plotly
                        for _si,_sec in enumerate(_sects_b):
                            _mask=(_sect_b==_sec)
                            _fig_tie.add_trace(go.Scatter(
                                x=_mat_b[_mask].tolist(),
                                y=(_tie_b[_mask]*100).tolist(),
                                mode="markers",name=f"{str(_sec)[:14]} ({_mask.sum()})",
                                marker=dict(size=5,color=_cpals[_si%len(_cpals)],opacity=0.75)))
                        _fig_tie.update_layout(
                            title=f"TIE Individuel (r_ZC + spread) — {len(df)} obligations",
                            height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                            xaxis_title="Maturite (ans)",yaxis_title="TIE (%)",
                            legend=dict(orientation="h",y=-0.3,font=dict(size=7)))
                        st.plotly_chart(_fig_tie,use_container_width=True,key="bam_tie_scatter")
                else:
                    st.info("Chargez le portefeuille pour voir la distribution des TIE.")

            # Tableau synthese courbe
            _df_courbe=pd.DataFrame({
                "Maturite (ans)": _T_std,
                "r_ZC (%)": [f"{z*100:.4f}" for z in _Z_std],
                "D(T)":     [f"{d:.6f}" for d in _D_std],
                "r_ZC+100bp TIE (%)": [f"{min((z+0.01)*100,25):.4f}" for z in _Z_std],
            })
            st.dataframe(_df_courbe,use_container_width=True,hide_index=True,height=200)

            # Telecharger courbe
            _buf_bam=io.BytesIO()
            pd.DataFrame({"T":_T_std,"r_ZC(%)":(_Z_std*100).round(4),"D(T)":_D_std.round(6)}).to_excel(
                _buf_bam,index=False)
            st.download_button("Telecharger Courbe BAM (Excel)",data=_buf_bam.getvalue(),
                file_name="Courbe_BAM_NSS.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_bam_curve")


    # ── Parametres (dans la page, pas sidebar) ────────────────────────────
    with st.expander("Parametres ECL", expanded=False):
        _pa, _pb, _pc = st.columns(3)
        with _pa:
            _Tp5 = st.slider("Horizon projection macro (ans)", 1, 10, 5, key="e5_tp")
            _Tr5 = st.slider("Periode reversion TTC (ans)",    1, 10, 3, key="e5_tr")
        with _pb:
            _wp5 = st.slider("Poids Pessimiste", 0.0, 1.0, 0.20, 0.05, key="e5_wp")
            _wc5 = st.slider("Poids Central",    0.0, 1.0, 0.70, 0.05, key="e5_wc")
            _wo5 = round(max(0.0, 1.0-_wp5-_wc5), 2)
            st.caption(f"Poids Optimiste = {_wo5:.0%}")
        with _pc:
            _ra5 = st.number_input("Taux actualisation", 0.0, 0.15, 0.035,
                                    0.001, format="%.3f", key="e5_ra")
            _dp5 = st.number_input("Delta Pessimiste CPV", -3.0, 3.0, 0.420,
                                    0.01, key="e5_dp")
            _dc5 = st.number_input("Delta Central CPV",    -3.0, 3.0,-0.252,
                                    0.01, key="e5_dc")
            _do5 = st.number_input("Delta Optimiste CPV",  -3.0, 3.0,-0.850,
                                    0.01, key="e5_do")

    _key5 = f"ecl5_{len(df)}_{_Tp5}_{_Tr5}_{round(_ra5,4)}"

    # ── Bouton calcul ─────────────────────────────────────────────────────
    _ba5, _bb5 = st.columns([1, 3])
    with _ba5:
        _btn5 = st.button("Calculer ECL IFRS 9", key="e5_btn",
                           type="primary", use_container_width=True)
    with _bb5:
        if _key5 in st.session_state:
            _tot5 = st.session_state[_key5]["ECL_FINAL"].sum()
            _n5   = len(st.session_state[_key5])
            st.success(f"ECL calcule — {_n5} obligations — ECL Total = {_tot5/1e3:.2f} K MAD")
        else:
            st.info(f"{len(df)} obligations chargees. Cliquez Calculer ECL IFRS 9.")

    # ── Execution ─────────────────────────────────────────────────────────
    if _btn5:
        with st.spinner(f"Calcul ECL — {len(df)} obligations x 3 scenarios..."):
            try:
                _res5 = _run_ecl_global(df, _Tp5, _Tr5, _dp5, _dc5, _do5,
                                         _wp5, _wc5, _wo5, _ra5)
                st.session_state[_key5] = _res5
                st.rerun()
            except Exception as _ex5:
                import traceback as _tb5
                st.error(f"Erreur calcul ECL: {_ex5}")
                st.code(_tb5.format_exc())

    # ── Affichage (si calcule) ────────────────────────────────────────────
    if _key5 in st.session_state:
        _df5 = st.session_state[_key5]
        _s15 = (_df5["STADE"]=="S1")
        _s25 = (_df5["STADE"]=="S2")
        _s35 = (_df5["STADE"]=="S3")

        # KPI
        _k1,_k2,_k3,_k4,_k5,_k6 = st.columns(6)
        _k1.metric("ECL Total",   f"{_df5['ECL_FINAL'].sum()/1e3:.2f} K MAD")
        _k2.metric("EAD Total",   f"{_df5['EAD_0'].sum()/1e6:.3f} M MAD")
        _k3.metric("ECL/EAD",     f"{_df5['ECL_FINAL'].sum()/_df5['EAD_0'].sum():.4%}")
        _k4.metric(f"S1({_s15.sum()})", f"{_df5.loc[_s15,'ECL_FINAL'].sum()/1e3:.2f}K")
        _k5.metric(f"S2({_s25.sum()})", f"{_df5.loc[_s25,'ECL_FINAL'].sum()/1e3:.2f}K")
        _k6.metric(f"S3({_s35.sum()})", f"{_df5.loc[_s35,'ECL_FINAL'].sum()/1e3:.2f}K")

        # Sous-onglets
        _t1,_t2,_t3,_t4 = st.tabs(["Tableau","Profil Obligation","Par Stade","Par Secteur"])

        with _t1:
            st.subheader("ECL IFRS 9 — 328 obligations")
            _nc5=next((c for c in["Description","NOM_EMETTEUR"] if c in _df5.columns),"")
            _sc5=next((c for c in["Secteur","SECTEUR_GICS"] if c in _df5.columns),"")
            _c15,_c25=st.columns(2)
            with _c15:
                _fg1=go.Figure(go.Bar(x=["S1","S2","S3"],
                    y=[_df5[_df5["STADE"]==s]["ECL_FINAL"].sum()/1e3 for s in["S1","S2","S3"]],
                    marker_color=["#1A6B2E","#8B6914","#A31919"],opacity=0.88,
                    text=[f"{_df5[_df5['STADE']==s]['ECL_FINAL'].sum()/1e3:.2f}K"
                          for s in["S1","S2","S3"]],textposition="outside"))
                _fg1.update_layout(title="ECL par Stade (K MAD)",height=300,
                    paper_bgcolor="white",plot_bgcolor="#F5F6FA",showlegend=False)
                st.plotly_chart(_fg1,use_container_width=True,key="e5_g1")
            with _c25:
                _fg2=go.Figure(go.Bar(x=["Pessimiste","Central","Optimiste"],
                    y=[_df5["ECL_Life_Pess"].sum()/1e3,
                       _df5["ECL_Life_Cent"].sum()/1e3,
                       _df5["ECL_Life_Opti"].sum()/1e3],
                    marker_color=["#A71E1E","#2E4D7B","#1A6B2E"],opacity=0.88,
                    text=[f"{v:.2f}K" for v in [_df5["ECL_Life_Pess"].sum()/1e3,
                          _df5["ECL_Life_Cent"].sum()/1e3,
                          _df5["ECL_Life_Opti"].sum()/1e3]],textposition="outside"))
                _fg2.update_layout(title="ECL Lifetime par Scenario",height=300,
                    paper_bgcolor="white",plot_bgcolor="#F5F6FA",showlegend=False)
                st.plotly_chart(_fg2,use_container_width=True,key="e5_g2")

            _cols5=["CODE_ISIN",_nc5,_sc5,"METHODE","STADE","mat_res","EAD_0",
                    "PD_TTC","PD_PIT_Central","LGD_TTC","LGD_PIT","ECL_12M_Pond","ECL_Life_Pond","ECL_FINAL"]
            _ds5=_df5[[c for c in _cols5 if c and c in _df5.columns]].copy() if _df5 is not None else pd.DataFrame()
            for _col_pct in["PD_TTC","PD_PIT_Central","LGD_TTC","LGD_PIT"]:
                if _col_pct in _ds5.columns: _ds5[_col_pct]=(_ds5[_col_pct]*100).round(4)
            def _hl5(row):
                bg={"S1":"#D6EAD7","S2":"#FFF3CD","S3":"#F8D7DA"}.get(str(row.get("STADE","")),"")  # noqa
                return [f"background-color:{bg}" if c=="STADE" else "" for c in row.index]
            st.dataframe(_ds5.reset_index(drop=True).style.apply(_hl5,axis=1),
                         use_container_width=True,hide_index=True,height=430)
            _buf5=io.BytesIO()
            with pd.ExcelWriter(_buf5, engine="openpyxl") as _wr5:
                # Onglet 1 : Resultats ECL complets
                _df5_exp = _df5.drop(columns=["_det_c"],errors="ignore").copy()
                _df5_exp.to_excel(_wr5, sheet_name="ECL_Resultats", index=False)
                # Onglet 2 : Courbe BAM + TIE par obligation
                _tie_df = _df5[["CODE_ISIN","Description","Secteur","METHODE","mat_res","r_ZC","TIE"]].copy()
                _tie_df.to_excel(_wr5, sheet_name="TIE_par_Obligation", index=False)
                # Onglet 3 : Resume par stade
                _stade_rows=[]
                for _s5 in ["S1","S2","S3"]:
                    _gx=_df5[_df5["STADE"]==_s5]
                    _stade_rows.append({"Stade":_s5,"N":len(_gx),
                        "EAD_0":_gx["EAD_0"].sum(),"EL_TTC":_gx["EL_TTC"].sum(),
                        "ECL_12M":_gx["ECL_12M_Pond"].sum(),"ECL_Vie":_gx["ECL_Life_Pond"].sum(),
                        "ECL_FINAL":_gx["ECL_FINAL"].sum()})
                pd.DataFrame(_stade_rows).to_excel(_wr5, sheet_name="Resume_Stades", index=False)
            st.download_button(
                "Telecharger ECL + TIE BAM (Excel)",
                data=_buf5.getvalue(),
                file_name="ECL_IFRS9_complet.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="e5_dl")

        with _t2:
            st.subheader("Profil ECL — Obligation individuelle")
            _fa5,_fb5=st.columns([1,3])
            with _fa5:
                _sf5=st.selectbox("Stade",["Tous","S1","S2","S3"],key="e5_sf")
                _mf5=st.selectbox("Methode",["Toutes","ZC","AA","TA"],key="e5_mf")
            _dff5=_df5.copy()
            if _sf5!="Tous": _dff5=_dff5[_dff5["STADE"]==_sf5]
            if _mf5!="Toutes": _dff5=_dff5[_dff5["METHODE"]==_mf5]
            _dff5=_dff5.reset_index(drop=True)
            with _fb5:
                if len(_dff5)==0:
                    st.warning("Aucune obligation.")
                else:
                    _nc5b=next((c for c in["Description","NOM_EMETTEUR"] if c in _dff5.columns),"")
                    _lb5=[f"{str(_dff5.iloc[i].get(_nc5b,''))[:50]} "
                          f"[{_dff5.iloc[i]['METHODE']}|{_dff5.iloc[i]['STADE']}|"
                          f"{_dff5.iloc[i]['mat_res']:.1f}a]"
                          for i in range(len(_dff5))]
                    _ix5=st.selectbox(f"Parmi {len(_dff5)} obligations",
                        list(range(len(_dff5))),format_func=lambda i:_lb5[i],key="e5_ix")

            if len(_dff5)>0:
                _r5=_dff5.iloc[_ix5]
                _st5r=str(_r5["STADE"])
                _sb5={"S1":"#1A6B2E","S2":"#8B6914","S3":"#A31919"}.get(_st5r,"#333")
                _xc1,_xc2,_xc3=st.columns(3)
                _xc1.markdown(
                    f'''<div style="background:#1B1F6B;color:white;padding:14px;border-radius:10px">
                    <div style="font-size:0.7rem;opacity:0.8">OBLIGATION</div>
                    <div style="font-size:0.8rem;font-weight:700;margin:4px 0">
                    {str(_r5.get(_nc5b if _nc5b else "",""))[:50]}</div>
                    Methode: <b>{_r5["METHODE"]}</b> | Mat: <b>{_r5["mat_res"]:.2f}a</b><br>
                    Stade: <span style="background:{_sb5};padding:1px 8px;border-radius:3px;
                    font-weight:700">{_st5r}</span>
                    </div>''',unsafe_allow_html=True)
                _xc2.markdown(
                    f'''<div style="background:#009FE3;color:white;padding:14px;border-radius:10px">
                    <div style="font-size:0.7rem;opacity:0.8">PD &amp; LGD &amp; ACTUALISATION</div>
                    PD_TTC: <b>{float(_r5["PD_TTC"])*100:.4f}%</b><br>
                    PD_PIT (Cent.): <b>{float(_r5.get("PD_PIT_Central",_r5["PD_TTC"]))*100:.4f}%</b><br>
                    LGD_TTC: <b>{float(_r5["LGD_TTC"])*100:.4f}%</b><br>
                    LGD_PIT (FJ): <b>{float(_r5["LGD_PIT"])*100:.4f}%</b><br>
                    EAD_0: <b>{float(_r5["EAD_0"]):,.0f} MAD</b><br>
                    r_ZC (BAM): <b>{float(_r5.get("r_ZC",0)):.4f}%</b><br>
                    TIE individuel: <b>{float(_r5.get("TIE",0)):.4f}%</b>
                    </div>''',unsafe_allow_html=True)
                _xc3.markdown(
                    f'''<div style="background:{_sb5};color:white;padding:14px;border-radius:10px">
                    <div style="font-size:0.7rem;opacity:0.8">ECL IFRS 9</div>
                    ECL 12M: <b>{float(_r5["ECL_12M_Pond"]):,.2f} MAD</b><br>
                    ECL Vie: <b>{float(_r5["ECL_Life_Pond"]):,.2f} MAD</b><br>
                    <b style="font-size:1.05rem">ECL FINAL: {float(_r5["ECL_FINAL"]):,.2f} MAD</b>
                    </div>''',unsafe_allow_html=True)
                _det5=_r5.get("_det_c",[])
                if _det5:
                    _dd5=pd.DataFrame(_det5)
                    _fg3=go.Figure()
                    _fg3.add_trace(go.Bar(x=[r["T"] for r in _det5],
                        y=[r["EL_t"] for r in _det5],name="EL_t (Central)",
                        marker_color="#2E4D7B",opacity=0.78))
                    _fg3.add_trace(go.Scatter(
                        x=[r["T"] for r in _det5],
                        y=[r["ECL_cumul"] for r in _det5],
                        name="ECL cumule",mode="lines+markers",
                        line=dict(color="#1B1F6B",width=2,dash="dot"),yaxis="y2"))
                    _fg3.update_layout(title=f"EL annuel — {_lb5[_ix5][:55]}",
                        height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                        xaxis_title="Annee",yaxis_title="EL_t (MAD)",
                        yaxis2=dict(overlaying="y",side="right",showgrid=False),
                        legend=dict(orientation="h",y=-0.3))
                    st.plotly_chart(_fg3,use_container_width=True,key="e5_det")
                    _dd5b=_dd5.copy()
                    # Colonnes: T, PD_ann(%), PD_marg(%), EAD_t, LGD_PIT(%), TIE(%), D_t, EL_t, ECL_cumul
                    _dd5b.columns=["Annee","PD_ann(%)","PD_marg(%)","EAD_t(MAD)","LGD_PIT(%)","TIE(%)","D_t","EL_t(MAD)","ECL_cumul(MAD)"]
                    # PD_ann et PD_marg sont stockes en fraction → convertir en %
                    _dd5b["PD_ann(%)"] = (_dd5b["PD_ann(%)"]).round(4)
                    _dd5b["PD_marg(%)"] = (_dd5b["PD_marg(%)"]).round(4)
                    _dd5b["LGD_PIT(%)"] = (_dd5b["LGD_PIT(%)"]).round(4)
                    _dd5b["TIE(%)"] = (_dd5b["TIE(%)"]).round(4)
                    _dd5b["D_t"] = _dd5b["D_t"].round(6)
                    _dd5b["EAD_t(MAD)"] = _dd5b["EAD_t(MAD)"].round(0)
                    _dd5b["EL_t(MAD)"] = _dd5b["EL_t(MAD)"].round(2)
                    _dd5b["ECL_cumul(MAD)"] = _dd5b["ECL_cumul(MAD)"].round(2)
                    st.dataframe(_dd5b,use_container_width=True,hide_index=True,height=400)

        with _t3:
            st.subheader("ECL par Stade")
            _rec5=[{"Stade":s,"N":int((_df5["STADE"]==s).sum()),
                    "EAD_0":f"{_df5[_df5['STADE']==s]['EAD_0'].sum():,.0f}",
                    "EL_TTC":f"{_df5[_df5['STADE']==s]['EL_TTC'].sum():,.2f}",
                    "ECL_12M":f"{_df5[_df5['STADE']==s]['ECL_12M_Pond'].sum():,.2f}",
                    "ECL_Vie":f"{_df5[_df5['STADE']==s]['ECL_Life_Pond'].sum():,.2f}",
                    "ECL_FINAL":f"{_df5[_df5['STADE']==s]['ECL_FINAL'].sum():,.2f}",
                    "ECL/EAD":f"{_df5[_df5['STADE']==s]['ECL_FINAL'].sum()/max(_df5[_df5['STADE']==s]['EAD_0'].sum(),1):.4%}"}
                   for s in["S1","S2","S3"]]
            st.dataframe(pd.DataFrame(_rec5),use_container_width=True,hide_index=True)

        with _t4:
            _scc5=next((c for c in["Secteur","SECTEUR_GICS"] if c in _df5.columns),None)
            if _scc5:
                _gs5=_df5.groupby(_scc5).agg(
                    N=("ECL_FINAL","count"),EAD=("EAD_0","sum"),
                    ECL=("ECL_FINAL","sum")).reset_index().sort_values("ECL",ascending=False)
                _gc1,_gc2=st.columns(2)
                with _gc1:
                    _fgs=go.Figure(go.Bar(x=_gs5["ECL"]/1e3,y=_gs5[_scc5],
                        orientation="h",
                        marker_color=px.colors.qualitative.Plotly[:len(_gs5)],opacity=0.88,
                        text=(_gs5["ECL"]/1e3).round(2).astype(str)+"K",
                        textposition="outside"))
                    _fgs.update_layout(title="ECL par Secteur (K MAD)",height=400,
                        paper_bgcolor="white",plot_bgcolor="#F5F6FA",showlegend=False)
                    st.plotly_chart(_fgs,use_container_width=True,key="e5_sec")
                with _gc2:
                    _fps=go.Figure(go.Pie(labels=_gs5[_scc5],values=_gs5["ECL"]/1e3,
                        hole=0.5,textinfo="label+percent",textfont_size=9,
                        marker=dict(colors=px.colors.qualitative.Plotly[:len(_gs5)])))
                    _fps.update_layout(title="Repartition ECL",height=400,
                        paper_bgcolor="white",
                        legend=dict(orientation="h",y=-0.3,font=dict(size=7)))
                    st.plotly_chart(_fps,use_container_width=True,key="e5_pie")
                st.dataframe(_gs5.reset_index(drop=True),
                             use_container_width=True,hide_index=True)


# ════════════════════════════════════════════════════════════════
# TAB 7 — ANALYSE & INTERPRETATION (IA Explicable)
# ════════════════════════════════════════════════════════════════
with TABS[7]:
    import json as _json_ai

    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>Analyse &amp; Interpretation — IA Explicable</h2>
      <p>Decomposition ECL &nbsp;|&nbsp; Drivers PD &nbsp;|&nbsp;
         Early Warning &nbsp;|&nbsp; Regime Macro &nbsp;|&nbsp; Stress Test</p>
    </div></div>""", unsafe_allow_html=True)

    if df is None:
        st.warning("Chargez vos donnees dans la sidebar.")

    # ─── Variables globales recuperees ───────────────────────────────────
    _df_ecl_ai = None
    _ecl_keys_ai = [k for k in st.session_state if str(k).startswith(f"ecl5_{len(df)}_")]
    if _ecl_keys_ai:
        _key5_ai = sorted(_ecl_keys_ai)[-1]
        _df_ecl_ai = st.session_state.get(_key5_ai, None)
    if _df_ecl_ai is None:
        st.info("Calculez d'abord l'ECL dans l'onglet 'ECL IFRS 9'.")

    # ─── Sub-onglets ─────────────────────────────────────────────────────
    sub_ai = st.tabs([
        "Decomposition ECL",
        "Drivers de la PD",
        "Early Warning",
        "Regime Macro",
        "Stress Test Interactif",
    ])

    # ══════════════════════════════════════════════════════════════════════
    # SUB 0 : DECOMPOSITION ECL PAR COMPOSANTE (Waterfall)
    # ══════════════════════════════════════════════════════════════════════
    with sub_ai[0]:
        st.subheader("Decomposition ECL — Methode des Effets Isoles")
        st.info(
            "Principe : on fixe successivement PD, LGD, EAD et D(t) a leur valeur "
            "de reference (TTC) et on mesure l'impact de chaque composante sur l'ECL. "
            "ECL = PD_marg × LGD_PIT × EAD × D(t)"
        )

        if _df_ecl_ai is None:
            st.warning("Calculez d'abord l'ECL dans l'onglet 'ECL IFRS 9' pour voir la decomposition.")
        else:
            # Selecteur obligation
            _ne_ai = next((c for c in ["Description","NOM_EMETTEUR"] if c in _df_ecl_ai.columns), "")
            _lbls_ai = [
                f"{str(_df_ecl_ai.iloc[i].get(_ne_ai,''))[:50]} "
                f"[{_df_ecl_ai.iloc[i]['METHODE']}|{_df_ecl_ai.iloc[i]['STADE']}|{_df_ecl_ai.iloc[i]['mat_res']:.1f}a]"
                for i in range(len(_df_ecl_ai))
            ]
            _ai_idx = st.selectbox("Choisir une obligation",
                options=list(range(len(_df_ecl_ai))),
                format_func=lambda i: _lbls_ai[i],
                key="ai_oblig_sel")

            _r_ai = _df_ecl_ai.iloc[_ai_idx]
            _pt   = float(_r_ai["PD_TTC"])
            _lt   = float(_r_ai["LGD_TTC"])
            _lp   = float(_r_ai["LGD_PIT"])
            _e0   = float(_r_ai["EAD_0"])
            _mr   = float(_r_ai["mat_res"])
            _meth = str(_r_ai["METHODE"])
            _st   = str(_r_ai["STADE"])
            _ecl_final = float(_r_ai["ECL_FINAL"])
            _tie  = float(_r_ai.get("TIE", 0.035))

            # Recalcul decomposition
            from scipy.stats import norm as _Nd
            def _phi_ai(x): return _Nd.cdf(np.clip(x,-30,30))
            def _phi_inv_ai(p): return _Nd.ppf(np.clip(p,1e-8,1-1e-8))
            def _rho_ai(p): w=(1-np.exp(-50*p))/(1-np.exp(-50)); return 0.12*w+0.24*(1-w)
            import math as _m_ai

            _d_cent = -0.252
            _pp_cent = float(1/(1+np.exp(-(_phi_inv_ai(_pt)+_d_cent))))
            _rho_v   = _rho_ai(_pt)
            _Z_ai    = (_phi_inv_ai(_pt)-np.sqrt(1-_rho_v)*_phi_inv_ai(_pp_cent))/np.sqrt(_rho_v)
            _Z_ai    = float(np.clip(_Z_ai,-5,5))

            def _cpd_ai(pa,t): return 1-(1-pa)**t
            def _cpd_rev_ai(pt,pp,t,Tp,Tr):
                if t<=Tp: return _cpd_ai(pp,t)
                di=_cpd_ai(pp,Tp)-_cpd_ai(pt,Tp)
                if t>Tp+Tr: return _cpd_ai(pt,t)
                fr=(Tr-(t-Tp))/Tr
                return _cpd_ai(pt,t)+(max(di*fr,0) if di>=0 else min(di*fr,0))
            def _pdm_ai(pt,pp,t,Tp,Tr,cpd_prev):
                pa=_pp_cent if t<=5 else pt
                if t>5: pa=pt+(_pp_cent-pt)*(5+3-t)/3 if t<=8 else pt
                return max(pa*(1-cpd_prev),0)
            def _ead_ai(e0,meth,mr,pp,t):
                e=float(e0)
                for ti in range(1,int(t)+1):
                    am=0.0 if str(meth).upper() in("ZC","TA") else 1.0/max(float(mr)-ti+1,1.0)
                    e=max(0.0,e*(1-float(pp))*(1-am))
                return e

            _tmax = 1 if _st=="S1" else int(_m_ai.ceil(_mr))

            # ECL reference (tout TTC)
            _ecl_ref=0; cpd_prev=0
            for t in range(1,_tmax+1):
                pm=_pdm_ai(_pt,_pt,t,5,3,cpd_prev)
                cpd_prev=1-(1-cpd_prev)*(1-_pt)
                et=_ead_ai(_e0,_meth,_mr,_pt,t-1)
                dt=1/(1+_tie)**t
                _ecl_ref+=pm*_lt*et*dt

            # ECL PD_PIT (changer PD seulement)
            _ecl_pd=0; cpd_prev=0
            for t in range(1,_tmax+1):
                pm=_pdm_ai(_pt,_pp_cent,t,5,3,cpd_prev)
                cpd_prev=1-(1-cpd_prev)*(1-_pp_cent)
                et=_ead_ai(_e0,_meth,_mr,_pp_cent,t-1)
                dt=1/(1+_tie)**t
                _ecl_pd+=pm*_lt*et*dt

            # ECL PD_PIT + LGD_PIT
            _ecl_pd_lgd=0; cpd_prev=0
            for t in range(1,_tmax+1):
                pm=_pdm_ai(_pt,_pp_cent,t,5,3,cpd_prev)
                cpd_prev=1-(1-cpd_prev)*(1-_pp_cent)
                et=_ead_ai(_e0,_meth,_mr,_pp_cent,t-1)
                dt=1/(1+_tie)**t
                _ecl_pd_lgd+=pm*_lp*et*dt

            # ECL final (PD_PIT + LGD_PIT + EAD dynamique + D(t) TIE)
            _ecl_full = _ecl_final

            _contrib_pd  = _ecl_pd    - _ecl_ref
            _contrib_lgd = _ecl_pd_lgd- _ecl_pd
            _contrib_ead = _ecl_full  - _ecl_pd_lgd

            # Waterfall
            _wf_vals  = [_ecl_ref, _contrib_pd, _contrib_lgd, _contrib_ead, _ecl_full]
            _wf_lbls  = ["EL TTC (base)","Effet PD_PIT macro","Effet LGD_PIT (FJ)","Effet EAD dynamique","ECL FINAL"]
            _wf_meas  = ["absolute","relative","relative","relative","total"]
            _wf_colors= ["#185FA5","#E24B4A" if _contrib_pd>0 else "#1A6B2E",
                         "#E24B4A" if _contrib_lgd>0 else "#1A6B2E",
                         "#E24B4A" if _contrib_ead>0 else "#1A6B2E","#1B1F6B"]

            fig_wf = go.Figure(go.Waterfall(
                name="ECL Decomposition",
                orientation="v",
                measure=_wf_meas,
                x=_wf_lbls,
                y=_wf_vals,
                connector=dict(line=dict(color="#999",width=1,dash="dot")),
                increasing=dict(marker_color="#E24B4A"),
                decreasing=dict(marker_color="#1A6B2E"),
                totals=dict(marker_color="#1B1F6B"),
                text=[f"{v:+.2f}" for v in _wf_vals],
                textposition="outside",
                textfont=dict(size=11,color="#333"),
            ))
            fig_wf.update_layout(
                title=f"Decomposition ECL — {_lbls_ai[_ai_idx][:50]}",
                yaxis_title="ECL (MAD)", height=420,
                paper_bgcolor="white", plot_bgcolor="#F5F6FA",
                showlegend=False)
            st.plotly_chart(fig_wf, use_container_width=True, key="ai_wf")

            # Tableau pct
            _tot_chg = abs(_contrib_pd)+abs(_contrib_lgd)+abs(_contrib_ead)
            _pct = lambda x: f"{x/_tot_chg*100:+.1f}%" if _tot_chg>0 else "0%"
            _df_dec = pd.DataFrame({
                "Composante":  ["EL TTC (reference)","Effet PD_PIT macro","Effet LGD_PIT Frye-Jacobs","Effet EAD dynamique","ECL FINAL"],
                "Contribution (MAD)": [f"{_ecl_ref:.4f}",f"{_contrib_pd:+.4f}",f"{_contrib_lgd:+.4f}",f"{_contrib_ead:+.4f}",f"{_ecl_full:.4f}"],
                "Part de l'ajustement": ["Base",_pct(_contrib_pd),_pct(_contrib_lgd),_pct(_contrib_ead),"100%"],
                "Interpretation": [
                    "PD TTC × LGD TTC × EAD fixe — vision moyenne cycle",
                    "Delta du au cycle macro actuel (CPV delta="+f"{_d_cent:+.3f})",
                    "Ajustement LGD par regime systémique (Z="+f"{_Z_ai:+.3f})",
                    "Erosion du capital selon la methode d'amortissement",
                    "ECL IFRS 9 conforme"
                ]
            })
            st.dataframe(_df_dec, use_container_width=True, hide_index=True)

            # Interpretation automatique
            _dom = max([("PD macro",abs(_contrib_pd)),("LGD PIT",abs(_contrib_lgd)),("EAD",abs(_contrib_ead))],key=lambda x:x[1])
            _dir = "hausse" if _ecl_full > _ecl_ref else "baisse"
            st.markdown(f"""
            <div style="background:#EEF2FF;border-left:4px solid #1B1F6B;padding:14px 18px;border-radius:6px;margin-top:8px">
            <b>Interpretation automatique :</b><br>
            L'ECL IFRS 9 de <b>{_ecl_full:.2f} MAD</b> est en <b>{_dir}</b> de
            <b>{abs(_ecl_full-_ecl_ref):.2f} MAD</b> par rapport a la reference TTC ({_ecl_ref:.2f} MAD).<br>
            L'effet dominant est <b>{_dom[0]}</b> ({_dom[1]:.2f} MAD, {_dom[1]/_tot_chg*100:.0f}% de l'ajustement).
            {"La PD PIT est superieure a la PD TTC, signalant des conditions macro defavorables pour cet emetteur." if _contrib_pd>0 else "La PD PIT est inferieure a la PD TTC, les conditions macro actuelles sont favorables."}
            </div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════
    # SUB 1 : DRIVERS DE LA PD (contribution macro)
    # ══════════════════════════════════════════════════════════════════════
    with sub_ai[1]:
        st.subheader("Drivers de la PD PIT — Attribution par Facteur Macro")

        from scipy.stats import norm as _Ndrv
        _phi_drv   = lambda x: _Ndrv.cdf(np.clip(x,-30,30))
        _phi_i_drv = lambda p: _Ndrv.ppf(np.clip(p,1e-8,1-1e-8))

        if dm_cpv is None:
            st.warning("Chargez Data facteur macro.xlsx pour voir les drivers.")
        elif "res_cpv4" not in dir():
            st.warning("Lancez d'abord le modele CPV (onglet PD CPV) pour voir les drivers.")
        else:
            try:
                _coefs_all = res_cpv4.params.drop("const",errors="ignore")
                _pvals_all = res_cpv4.pvalues.drop("const",errors="ignore")

                # Agreger coefs lagged → variable de base (somme des beta_lag)
                _base_coefs={}; _base_pvals={}
                for _cn,_cv in _coefs_all.items():
                    _bn=_cn.split("_lag")[0] if "_lag" in _cn else _cn
                    _base_coefs[_bn]=_base_coefs.get(_bn,0.0)+float(_cv)
                    _base_pvals[_bn]=min(_base_pvals.get(_bn,1.0),float(_pvals_all.get(_cn,1.0)))

                # Variables disponibles dans dm_cpv
                _mac_vars_d=[v for v in _base_coefs if v in dm_cpv.columns]
                _last_d=dm_cpv.iloc[-1]; _mean_d=dm_cpv[_mac_vars_d].mean(); _std_d=dm_cpv[_mac_vars_d].std()

                # Calcul contributions actuelles + sensibilite
                _rows_drv=[]
                for _v in _mac_vars_d:
                    _coef=_base_coefs[_v]; _pval=_base_pvals[_v]
                    _dx=float(_last_d.get(_v,0))-float(_mean_d.get(_v,0))
                    _contrib=_coef*_dx
                    _sensib=abs(_coef)*float(_std_d.get(_v,1.0))
                    _rows_drv.append({
                        "variable":_v,"coef":_coef,"pval":_pval,"dx":_dx,
                        "contrib":_contrib,"sensib":_sensib,
                    })
                _df_drv_full=pd.DataFrame(_rows_drv).sort_values("sensib",ascending=False).reset_index(drop=True)

                st.markdown("""
                <div style="background:#EEF2FF;border-left:4px solid #1B1F6B;padding:12px 16px;border-radius:6px;margin:8px 0;font-size:0.85rem">
                <b>Deux lectures complémentaires :</b><br>
                 <b>Contribution actuelle</b> = β × ΔX (valeur 2024 − moyenne historique) → ce qui EXPLIQUE la PD_PIT actuelle<br>
                 <b>Sensibilité potentielle</b> = |β| × σ(X) → quel facteur AURAIT LE PLUS D'IMPACT si il variait de 1 écart-type
                </div>""", unsafe_allow_html=True)

                # Delta total pour PD_PIT
                _delta_cpv_est=sum(_base_coefs.get(v,0)*float(_mean_d.get(v,0)) for v in _mac_vars_d)
                _sum_abs_d=sum(abs(r["contrib"]) for _,r in _df_drv_full.iterrows()) or 1

                _gc1,_gc2=st.columns(2)
                with _gc1:
                    # Graphe 1: CONTRIBUTION ACTUELLE (barres)
                    _labs_c=[f"{r['variable']}" for _,r in _df_drv_full.iterrows()]
                    _vals_c=[r["contrib"] for _,r in _df_drv_full.iterrows()]
                    _cols_c=["#E24B4A" if v>0 else ("#1A6B2E" if v<0 else "#999") for v in _vals_c]
                    _txt_c=[f"β={r['coef']:+.3f} | ΔX={r['dx']:+.4f}" for _,r in _df_drv_full.iterrows()]
                    fig_dc=go.Figure()
                    fig_dc.add_trace(go.Bar(
                        x=_vals_c,y=_labs_c,orientation="h",
                        marker_color=_cols_c,opacity=0.88,
                        text=[f"{v:+.5f}" if abs(v)>1e-6 else "≈ 0" for v in _vals_c],
                        textposition="outside",
                        customdata=_txt_c,
                        hovertemplate="%{y}<br>%{customdata}<br>Contribution: %{x:+.5f}"))
                    fig_dc.add_vline(x=0,line_color="#333",line_width=1)
                    fig_dc.update_layout(
                        title="Contribution ACTUELLE au Δlogit(DR)<br><sup>β × (X_2024 − X_moyen) — Rouge=hausse risque, Vert=baisse risque</sup>",
                        height=max(300,40*len(_mac_vars_d)+80),
                        paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                        xaxis_title="Contribution au logit(DR)",showlegend=False,
                        margin=dict(l=130,r=80,t=70,b=40))
                    st.plotly_chart(fig_dc,use_container_width=True,key="ai_drv_c")

                with _gc2:
                    # Graphe 2: SENSIBILITE POTENTIELLE (barres horizontales)
                    _df_sens=_df_drv_full.sort_values("sensib",ascending=True)
                    _labs_s=[r["variable"] for _,r in _df_sens.iterrows()]
                    _vals_s=[r["sensib"] for _,r in _df_sens.iterrows()]
                    _cols_s=["#A71E1E" if _base_coefs.get(r["variable"],0)>0 else "#1A6B2E"
                             for _,r in _df_sens.iterrows()]
                    fig_ds=go.Figure()
                    fig_ds.add_trace(go.Bar(
                        x=_vals_s,y=_labs_s,orientation="h",
                        marker_color=_cols_s,opacity=0.80,
                        text=[f"|β|={abs(_base_coefs.get(r['variable'],0)):.3f} × σ={float(_std_d.get(r['variable'],1)):.4f}"
                              for _,r in _df_sens.iterrows()],
                        textposition="outside"))
                    fig_ds.update_layout(
                        title="Sensibilité POTENTIELLE par Facteur (1 écart-type)<br><sup>|β| × σ(X) — Montre l'impact maximum potentiel de chaque facteur</sup>",
                        height=max(300,40*len(_mac_vars_d)+80),
                        paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                        xaxis_title="|β| × σ",showlegend=False,
                        margin=dict(l=130,r=80,t=70,b=40))
                    st.plotly_chart(fig_ds,use_container_width=True,key="ai_drv_s")

                # Tableau detaille
                _df_tab_drv=pd.DataFrame({
                    "Facteur": _df_drv_full["variable"],
                    "β agrégé (Σ lags)": _df_drv_full["coef"].round(4),
                    "p-value": _df_drv_full["pval"].apply(lambda p: f"{p:.3f}"),
                    "Sign. (5%)": [" Oui" if p<0.05 else "Non" for p in _df_drv_full["pval"]],
                    "ΔX actuel": _df_drv_full["dx"].apply(lambda x: f"{x:+.4f}" if abs(x)>1e-6 else "≈ 0"),
                    "Contribution": _df_drv_full["contrib"].apply(lambda v: f"{v:+.5f}" if abs(v)>1e-6 else "≈ 0 (X stable)"),
                    "Sensib. 1σ": _df_drv_full["sensib"].round(4),
                    "Signe β": ["↑ Risque quand ↑" if c>0 else "↓ Risque quand ↑" for c in _df_drv_full["coef"]],
                })
                st.dataframe(_df_tab_drv,use_container_width=True,hide_index=True)

                # Interpretation
                _top_contrib=[r for _,r in _df_drv_full.iterrows() if abs(r["contrib"])>1e-6]
                _top_sensib=_df_drv_full.iloc[0]  # highest sensitivity
                _zero_vars=[r["variable"] for _,r in _df_drv_full.iterrows() if abs(r["contrib"])<1e-6]

                _pd_ttc_moy=float(df["PD_TTC"].mean()) if "PD_TTC" in df.columns else 0.041
                _pp_cent=float(1/(1+np.exp(-(_phi_i_drv(_pd_ttc_moy)-0.252))))
                _delta_obs=_phi_i_drv(_pp_cent)-_phi_i_drv(_pd_ttc_moy)

                st.markdown(f"""
                <div style="background:#F0FDF4;border-left:4px solid #1A6B2E;padding:14px 18px;border-radius:6px;margin-top:12px">
                <b> Interpretation des drivers CPV :</b><br>
                La PD_PIT Centrale est <b>{"supérieure" if _delta_obs>0 else "inférieure"} à la PD_TTC</b>
                (Δlogit CPV ≈ {_delta_obs:+.3f}).<br><br>

                {"<b> Contribution actuelle :</b><br>" +
                ("<br>".join([f"• <b>{r['variable']}</b> : β={r['coef']:+.3f} × ΔX={r['dx']:+.4f} = <b>{r['contrib']:+.5f}</b>"
                 for r in _top_contrib[:5]])) if _top_contrib else
                "• Toutes les variables sont à leur <b>valeur moyenne historique</b> → contributions ≈ 0."}<br><br>

                {"<b> Variables stables en 2024</b> (ΔX≈0, mais potentiellement impactantes) : " +
                ", ".join(_zero_vars[:4]) + "<br>" if _zero_vars else ""}

                <b> Facteur le plus sensible</b> : <b>{_top_sensib["variable"]}</b>
                (sensibilité 1σ = {_top_sensib["sensib"]:.4f}) —
                {"Si ce facteur augmente d'1σ, le logit(DR) change de " + f"{abs(_top_sensib['sensib']):.4f}"}.
                </div>""", unsafe_allow_html=True)

            except Exception as _e_drv2:
                st.warning(f"Erreur drivers: {_e_drv2}")
                import traceback; st.code(traceback.format_exc())

    # ══════════════════════════════════════════════════════════════════════
    # SUB 2 : EARLY WARNING
    # ══════════════════════════════════════════════════════════════════════
    with sub_ai[2]:
        st.subheader("Early Warning System — Score d'Alerte par Obligation")
        st.info(
            "Score 0–100 base sur 3 signaux : "
            "(1) Proximity au seuil de stade, "
            "(2) Ecart PD_PIT / PD_TTC, "
            "(3) LGD PIT vs LGD TTC. "
            "Rouge > 70 | Orange 40–70 | Vert < 40"
        )

        # Calcul du score EWS
        _ews_rows = []
        for _i_ews in range(len(df)):
            _r_ews = df.iloc[_i_ews]
            _spread = float(_r_ews.get("SPREAD_ADJ", _r_ews.get("SPREAD_EMISSION", 0)))
            _pd_ttc_ews = float(_r_ews.get("PD_TTC", _r_ews.get("PD",0.05)/100 if _r_ews.get("PD",100)>1 else _r_ews.get("PD",0.05)))
            _pd_pit_ews = float(_r_ews.get("PD_PIT_CPV", _pd_ttc_ews))
            _lgd_ttc_ews = float(_r_ews.get("LGD_TTC", _r_ews.get("LGD", 0.30)))
            _lgd_pit_ews = float(_r_ews.get("LGD_PIT_V_Pond", _lgd_ttc_ews))
            _snum = int(_r_ews.get("SNUM", 1))

            # Signal 1 : proximite seuil (0-40)
            if _snum == 1:
                _prox = min((_spread / S12 * 40) if S12>0 else 0, 40)
            elif _snum == 2:
                _prox = 40 + min((_spread / S23 * 30) if S23>0 else 0, 30)
            else:
                _prox = 70

            # Signal 2 : ecart PD_PIT/PD_TTC (0-30)
            _ratio_pd = (_pd_pit_ews / _pd_ttc_ews) if _pd_ttc_ews > 0 else 1
            _sig_pd = min(max((_ratio_pd - 1) * 30, 0), 30)

            # Signal 3 : ecart LGD_PIT/LGD_TTC (0-30)
            _ratio_lgd = (_lgd_pit_ews / _lgd_ttc_ews) if _lgd_ttc_ews > 0 else 1
            _sig_lgd = min(max((_ratio_lgd - 1) * 30, 0), 30)

            _score = min(_prox + _sig_pd + _sig_lgd, 100)
            _niveau = "Critique" if _score >= 70 else ("Alerte" if _score >= 40 else "Normal")
            _marg_seuil = (S23 - _spread) if _snum <= 2 else 0

            _ews_rows.append({
                "CODE_ISIN":       str(_r_ews.get("CODE_ISIN",""))[:15],
                "Emetteur":        str(_r_ews.get("NOM_EMETTEUR",""))[:30],
                "Stade":           f"S{_snum}",
                "Spread (bps)":    round(_spread,1),
                "PD_TTC (%)":      round(_pd_ttc_ews*100,3),
                "PD_PIT (%)":      round(_pd_pit_ews*100,3),
                "Ratio PD":        round(_ratio_pd,3),
                "LGD_PIT/TTC":     round(_ratio_lgd,3),
                "Marge seuil (bps)": round(_marg_seuil,1),
                "Score EWS":       round(_score,1),
                "Niveau":          _niveau,
            })

        _df_ews = pd.DataFrame(_ews_rows).sort_values("Score EWS", ascending=False).reset_index(drop=True)

        # KPIs EWS
        _n_crit = (_df_ews["Niveau"]=="Critique").sum()
        _n_alrt = (_df_ews["Niveau"]=="Alerte").sum()
        _n_norm = (_df_ews["Niveau"]=="Normal").sum()
        ea,eb,ec,ed = st.columns(4)
        ea.metric("Obligations Critiques",  str(_n_crit), delta=None)
        eb.metric("Obligations en Alerte",  str(_n_alrt))
        ec.metric("Obligations Normales",   str(_n_norm))
        ed.metric("Score moyen portefeuille", f"{_df_ews['Score EWS'].mean():.1f}/100")

        # Graphe scatter score vs spread
        col_ew1, col_ew2 = st.columns(2)
        with col_ew1:
            _ews_colors = _df_ews["Niveau"].map({"Critique":"#A71E1E","Alerte":"#EF9F27","Normal":"#1A6B2E"})
            fig_ews = go.Figure(go.Scatter(
                x=_df_ews["Spread (bps)"].tolist(),
                y=_df_ews["Score EWS"].tolist(),
                mode="markers",
                marker=dict(size=8, color=_ews_colors.tolist(), opacity=0.80,
                            line=dict(color="white",width=0.5)),
                text=(_df_ews["Emetteur"]+" | S="+_df_ews["Score EWS"].astype(str)).tolist(),
                hovertemplate="%{text}<br>Spread=%{x:.0f}bps<br>Score=%{y:.0f}"))
            fig_ews.add_hline(y=70,line_color="#A71E1E",line_dash="dash",
                annotation_text="Seuil Critique (70)")
            fig_ews.add_hline(y=40,line_color="#EF9F27",line_dash="dot",
                annotation_text="Seuil Alerte (40)")
            fig_ews.update_layout(title="Early Warning : Score vs Spread",
                height=380, paper_bgcolor="white", plot_bgcolor="#F5F6FA",
                xaxis_title="Spread (bps)", yaxis_title="Score EWS (0-100)")
            st.plotly_chart(fig_ews, use_container_width=True, key="ai_ews_sc")

        with col_ew2:
            # Gauge portefeuille
            _score_moy = _df_ews["Score EWS"].mean()
            fig_gauge = go.Figure(go.Indicator(
                mode="gauge+number+delta",
                value=_score_moy,
                delta={"reference": 40, "increasing": {"color":"#A71E1E"}},
                title={"text": "Score EWS Portefeuille"},
                gauge={
                    "axis": {"range": [0,100]},
                    "bar":  {"color": "#A71E1E" if _score_moy>=70 else ("#EF9F27" if _score_moy>=40 else "#1A6B2E")},
                    "steps": [
                        {"range":[0,40],   "color":"#D6EAD7"},
                        {"range":[40,70],  "color":"#FFF3CD"},
                        {"range":[70,100], "color":"#F8D7DA"},
                    ],
                    "threshold": {"line":{"color":"#A71E1E","width":3},"thickness":0.75,"value":70}
                }))
            fig_gauge.update_layout(height=380, paper_bgcolor="white")
            st.plotly_chart(fig_gauge, use_container_width=True, key="ai_gauge")

        # Top 15 obligations critiques
        st.subheader("Top 15 Obligations — Score le plus eleve")
        def _hl_ews(row):
            c={"Critique":"#F8D7DA","Alerte":"#FFF3CD","Normal":"#D6EAD7"}.get(row.get("Niveau",""),"")
            return [f"background-color:{c}" if col=="Niveau" else "" for col in row.index]
        st.dataframe(_df_ews.head(15).style.apply(_hl_ews,axis=1),
                     use_container_width=True, hide_index=True, height=420)

        # Interpretation EWS
        _top1_ews = _df_ews.iloc[0]
        _pct_crit = _n_crit/len(df)*100
        _ews_msg = "elevee" if _pct_crit>15 else ("moderee" if _pct_crit>5 else "faible")
        st.markdown(f"""
        <div style="background:#FFF8EE;border-left:4px solid #EF9F27;padding:14px 18px;border-radius:6px;margin-top:8px">
        <b>Synthese Early Warning :</b><br>
        {_n_crit} obligations ({_pct_crit:.1f}%) sont en zone critique — concentration de risque {_ews_msg}.<br>
        L'obligation la plus a risque est <b>{_top1_ews['Emetteur']}</b>
        (Score {_top1_ews['Score EWS']:.0f}/100, Spread={_top1_ews['Spread (bps)']:.0f} bps,
        PD_PIT={_top1_ews['PD_PIT (%)']:.3f}%).
        {f"La marge avant passage en stade superieur est de {_top1_ews['Marge seuil (bps)']:.0f} bps." if _top1_ews['Marge seuil (bps)']>0 else ""}
        </div>""", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════
    # SUB 3 : REGIME MACRO (K-Means sur donnees macro historiques)
    # ══════════════════════════════════════════════════════════════════════
    with sub_ai[3]:
        st.subheader("Detection de Regime Macro-Economique (K-Means)")
        st.info(
            "Les annees 1994-2024 sont clusterisees selon 6 facteurs macro. "
            "On identifie le regime actuel et on compare l'ECL observe dans des periodes similaires."
        )

        if dm_cpv is None:
            st.warning("Chargez les donnees macro (Data facteur macro.xlsx) pour la detection de regime.")
        else:
            from sklearn.preprocessing import StandardScaler as _SS
            from sklearn.cluster import KMeans as _KM

            _mac_vars_ai = [c for c in ["inflation","croissance_pib","chomage","taux_change","masse_monetaire","taux_directeur"]
                           if c in dm_cpv.columns]
            _df_mac_ai   = dm_cpv[_mac_vars_ai + ["annee"]].dropna().copy()
            _X_mac       = _df_mac_ai[_mac_vars_ai].values
            _sc_mac      = _SS(); _X_sc = _sc_mac.fit_transform(_X_mac)

            _n_reg_ai    = st.slider("Nombre de regimes",2,5,3,key="ai_n_reg")
            _km_ai       = _KM(n_clusters=_n_reg_ai, random_state=42, n_init=20)
            _km_ai.fit(_X_sc)
            _labels_ai   = _km_ai.labels_

            _df_mac_ai   = _df_mac_ai.copy()
            _df_mac_ai["Regime"] = [f"Regime {l+1}" for l in _labels_ai]

            # Caracteriser chaque regime
            _regime_colors = ["#1A6B2E","#185FA5","#EF9F27","#A71E1E","#5B2D8E"]
            _regime_info = {}
            for _rl in range(_n_reg_ai):
                _mask_r = (_labels_ai==_rl)
                _mean_r = _df_mac_ai[_mask_r][_mac_vars_ai].mean()
                _annees_r = _df_mac_ai[_mask_r]["annee"].values
                _regime_info[f"Regime {_rl+1}"] = {
                    "annees": list(_annees_r),
                    "mean":   _mean_r,
                    "n":      _mask_r.sum()
                }

            # Identifier le regime actuel (derniere annee)
            _last_scaled = _sc_mac.transform(_df_mac_ai.iloc[[-1]][_mac_vars_ai])
            _current_regime = f"Regime {_km_ai.predict(_last_scaled)[0]+1}"
            _current_year   = int(_df_mac_ai["annee"].iloc[-1])

            col_rm1, col_rm2 = st.columns(2)
            with col_rm1:
                # Timeline des regimes
                fig_reg = go.Figure()
                for _rl in range(_n_reg_ai):
                    _mask_r = (_labels_ai==_rl)
                    _ann_r  = _df_mac_ai[_mask_r]["annee"].values
                    fig_reg.add_trace(go.Scatter(
                        x=_ann_r, y=[_rl+1]*len(_ann_r), mode="markers",
                        name=f"Regime {_rl+1}",
                        marker=dict(size=14, color=_regime_colors[_rl],
                                    symbol="square", line=dict(color="white",width=1)),
                        text=[f"Regime {_rl+1} — Annee {int(a)}" for a in _ann_r],
                        hovertemplate="%{text}"))
                fig_reg.update_layout(
                    title=f"Regimes macro-economiques 1994-{_current_year}",
                    height=320, paper_bgcolor="white", plot_bgcolor="#F5F6FA",
                    xaxis_title="Annee", yaxis=dict(tickvals=list(range(1,_n_reg_ai+1)),
                    ticktext=[f"Regime {i+1}" for i in range(_n_reg_ai)]),
                    legend=dict(orientation="h",y=-0.25))
                st.plotly_chart(fig_reg, use_container_width=True, key="ai_reg_tl")

            with col_rm2:
                # Profil du regime actuel vs autres
                _cur_info = _regime_info[_current_regime]
                fig_prof_r = go.Figure()
                _theta_r = _mac_vars_ai + [_mac_vars_ai[0]]
                for _rl in range(_n_reg_ai):
                    _rname = f"Regime {_rl+1}"
                    _means_r = [float(_regime_info[_rname]["mean"].get(v,0)) for v in _mac_vars_ai]
                    _norm_r  = [_means_r[i]/(abs(_df_mac_ai[_mac_vars_ai[i]].max())+1e-8)*100 for i in range(len(_mac_vars_ai))]
                    _norm_r  = _norm_r + [_norm_r[0]]
                    fig_prof_r.add_trace(go.Scatterpolar(
                        r=_norm_r, theta=_theta_r,
                        fill="toself" if _rname==_current_regime else None,
                        name=_rname + (" (ACTUEL)" if _rname==_current_regime else ""),
                        line=dict(color=_regime_colors[_rl], width=2.5 if _rname==_current_regime else 1.5,
                                  dash="solid" if _rname==_current_regime else "dot"),
                        opacity=0.85 if _rname==_current_regime else 0.55))
                fig_prof_r.update_layout(
                    polar=dict(radialaxis=dict(visible=True)),
                    title=f"Profil des regimes — {_current_regime} = Regime actuel",
                    height=320, paper_bgcolor="white",
                    legend=dict(orientation="h",y=-0.25))
                st.plotly_chart(fig_prof_r, use_container_width=True, key="ai_reg_rad")

            # Tableau caracteristiques
            _rows_reg = []
            for _rl in range(_n_reg_ai):
                _rname = f"Regime {_rl+1}"
                _ri    = _regime_info[_rname]
                _annees_str = str(sorted([int(a) for a in _ri["annees"]])[:4])[1:-1].replace(",","") + "..."
                _row_r = {"Regime": _rname + (" [ACTUEL]" if _rname==_current_regime else ""),
                           "Nb annees": _ri["n"], "Annees": _annees_str}
                for v in _mac_vars_ai[:4]:
                    _row_r[v] = round(float(_ri["mean"].get(v,0)),3)
                _rows_reg.append(_row_r)
            st.dataframe(pd.DataFrame(_rows_reg), use_container_width=True, hide_index=True)

            # Interpretation
            _cur_annees = sorted([int(a) for a in _cur_info["annees"]])
            st.markdown(f"""
            <div style="background:#EEF2FF;border-left:4px solid #1B1F6B;padding:14px 18px;border-radius:6px;margin-top:8px">
            <b>Regime actuel ({_current_year}) : {_current_regime}</b><br>
            Ce regime est similaire aux annees : {_cur_annees}.<br>
            Ce regime est similaire aux annees : {_cur_annees}.<br>
            </div>""", unsafe_allow_html=True)
    # ══════════════════════════════════════════════════════════════════════

    with sub_ai[4]:
        st.subheader("Stress Test — Impact par Obligation")
        st.info(
            "Ce stress test recalcule l'ECL pour chaque obligation en appliquant "
            "les chocs macro directement sur la PD_PIT (via Δlogit)."
        )

        # ── Scenario + valeurs ──────────────────────────────────────────────
        _SCENARIOS_ST={
            "Personnalise":        dict(pib=0.0,infl=0.0,chom=0.0,taux=0.0,chg=0.0,mm=0.0),
            "Recession moderee":   dict(pib=-2.0,infl=1.5,chom=2.0,taux=0.5,chg=0.0,mm=0.0),
            "Crise severe":        dict(pib=-5.0,infl=3.0,chom=4.0,taux=2.0,chg=10.0,mm=0.0),
            "Choc inflationniste": dict(pib=-1.0,infl=4.5,chom=0.0,taux=3.0,chg=0.0,mm=0.0),
            "Reprise favorable":   dict(pib=2.5,infl=-1.0,chom=-2.0,taux=-0.5,chg=0.0,mm=2.0),
        }
        _SENS_ECONO={  # Sensibilites heuristiques (impact sur logit DR)
            "pib":  -0.35,   # PIB ↑ → risque ↓
            "infl": +0.20,   # Inflation ↑ → risque ↑
            "chom": +0.25,   # Chomage ↑ → risque ↑
            "taux": +0.15,   # Taux ↑ → risque ↑
            "chg":  +0.05,   # Change ↑ → risque ↑
            "mm":   -0.05,   # MM ↑ → risque ↓
        }

        # Selectbox
        _preset_st=st.selectbox("Scenario predifini",list(_SCENARIOS_ST.keys()),key="st_preset")
        _pv_st=_SCENARIOS_ST[_preset_st]

        st.info(
            f"**{_preset_st}** : "
            f"PIB={_pv_st['pib']:+.1f}% | Inflation={_pv_st['infl']:+.1f}pp | "
            f"Chomage={_pv_st['chom']:+.1f}pp | Taux BAM={_pv_st['taux']:+.2f}pp | "
            f"Taux change={_pv_st['chg']:+.1f}% | Masse mon.={_pv_st['mm']:+.1f}%"
        )

        # Si personnalise: permettre modification
        if _preset_st=="Personnalise":
            _st_c1,_st_c2=st.columns(2)
            with _st_c1:
                _pv_st["pib"]  =st.number_input("PIB (variation %)",-10.0,5.0,0.0,0.5,key="st_pib",format="%.1f")
                _pv_st["infl"] =st.number_input("Inflation (Δ pp)",-5.0,10.0,0.0,0.5,key="st_infl",format="%.1f")
                _pv_st["chom"] =st.number_input("Chomage (Δ pp)",-5.0,10.0,0.0,0.5,key="st_chom",format="%.1f")
            with _st_c2:
                _pv_st["taux"] =st.number_input("Taux BAM (Δ pp)",-3.0,6.0,0.0,0.25,key="st_taux",format="%.2f")
                _pv_st["chg"]  =st.number_input("Taux change (Δ %)",-10.0,20.0,0.0,1.0,key="st_chg",format="%.1f")
                _pv_st["mm"]   =st.number_input("Masse mon. (Δ%%)",-10.0,15.0,0.0,1.0,key="st_mm",format="%.1f")

        # Calcul Δlogit via sensibilites ECONOMIQUES intuitives (toujours)
        _delta_heur=(
            _SENS_ECONO["pib"]*_pv_st["pib"]+_SENS_ECONO["infl"]*_pv_st["infl"]+
            _SENS_ECONO["chom"]*_pv_st["chom"]+_SENS_ECONO["taux"]*_pv_st["taux"]+
            _SENS_ECONO["chg"]*_pv_st["chg"]+_SENS_ECONO["mm"]*_pv_st["mm"])

        # Optionnel: afficher aussi Δlogit via coefs CPV pour reference
        _delta_cpv_st=None
        if "res_cpv4" in dir():
            try:
                _cc2=res_cpv4.params.drop("const",errors="ignore")
                _shk2={"inflation":_pv_st["infl"],"croissance_pib":_pv_st["pib"],
                        "chomage":_pv_st["chom"],"taux_directeur":_pv_st["taux"],
                        "taux_change":_pv_st["chg"],"masse_monetaire":_pv_st["mm"]}
                _dl2=sum(float(cv)*_shk2.get(cn.split("_lag")[0] if "_lag" in cn else cn,0.0)
                         for cn,cv in _cc2.items())
                _delta_cpv_st=float(np.clip(_dl2,-3.0,3.0))
            except: pass

        # Afficher les deux Δlogit
        _col_st_h="#1A6B2E" if _delta_heur<0 else ("#A71E1E" if _delta_heur>0 else "#555")
        _sens_str_h="→ Baisse risque (PD ↓, ECL ↓) " if _delta_heur<0 else ("→ Hausse risque (PD ↑, ECL ↑) " if _delta_heur>0 else "→ Neutre")
        st.markdown(f"""
        <div style="background:{_col_st_h};color:white;padding:10px 16px;border-radius:8px;margin:8px 0">
        <b>Δ logit(DR) = {_delta_heur:+.4f}</b> &nbsp; {_sens_str_h}
        <br><span style="font-size:0.8rem;opacity:0.85">Source: Sensibilités économiques calibrées
        {f" | CPV (modèle): {_delta_cpv_st:+.4f}" if _delta_cpv_st is not None else ""}
        </span></div>""",unsafe_allow_html=True)

        # Bouton
        if _df_ecl_ai is not None:
            _skey3=f"stress3_{len(df)}_{round(_delta_heur,4)}"
            if st.button("Lancer le Stress Test",type="primary",key="st_run"):
                from scipy.stats import norm as _Ns3; import math as _ms3
                _ph3=lambda x:_Ns3.cdf(np.clip(x,-30,30))
                _phi3=lambda p:_Ns3.ppf(np.clip(p,1e-8,1-1e-8))
                def _rh3(p): w=(1-np.exp(-50*p))/(1-np.exp(-50)); return 0.12*w+0.24*(1-w)
                def _lg3(pt,lt,rho,Z):
                    p=np.clip(pt*lt,1e-8,1-1e-8); pp=np.clip(_ph3((_phi3(pt)-np.sqrt(rho)*Z)/np.sqrt(1-rho)),1e-8,1-1e-8)
                    return float(np.clip(_ph3(_phi3(pp)-(_phi3(pt)-_phi3(p))/np.sqrt(1-rho))/pp,0.001,0.999))
                def _pd3(pt,pp,t,Tp,Tr): 
                    if t<=Tp: return pp
                    if t>Tp+Tr: return pt
                    return pt+(pp-pt)*(Tp+Tr-t)/Tr
                def _pdm3(pt,pp,t,Tp,Tr,cpd_p): return max(_pd3(pt,pp,t,Tp,Tr)*(1-cpd_p),0)
                def _ea3(e0,meth,mr,pp,t):
                    e=float(e0)
                    for ti in range(1,int(t)+1):
                        am=0.0 if str(meth).upper() in("ZC","TA") else 1.0/max(float(mr)-ti+1,1.0)
                        e=max(0.0,e*(1-float(pp))*(1-am))
                    return e

                with st.spinner(f"Calcul ECL stresse — {len(_df_ecl_ai)} obligations..."):
                    _pg3=st.progress(0); _rs3=[]
                    for _ix3 in range(len(_df_ecl_ai)):
                        _r3=_df_ecl_ai.iloc[_ix3]
                        _pt3=float(_r3["PD_TTC"]); _lt3=float(_r3["LGD_TTC"])
                        _e03=float(_r3["EAD_0"]); _mr3=float(_r3["mat_res"])
                        _me3=str(_r3["METHODE"]); _sd3=str(_r3["STADE"])
                        _tie3=float(_r3.get("TIE",3.5)); 
                        if _tie3>1: _tie3=_tie3/100
                        _ecl3b=float(_r3["ECL_FINAL"])
                        # PD_PIT stresse via Δlogit heuristique
                        _pp3b=float(np.clip(1/(1+np.exp(-(_phi3(np.clip(_pt3,1e-6,1-1e-6))-0.252))),1e-6,1-1e-6))
                        _pp3s=float(np.clip(1/(1+np.exp(-(_phi3(np.clip(_pt3,1e-6,1-1e-6))-0.252+_delta_heur))),1e-6,1-1e-6))
                        _rv3=_rh3(_pt3)
                        _Zv3=(_phi3(_pt3)-np.sqrt(1-_rv3)*_phi3(_pp3s))/np.sqrt(_rv3)
                        _lp3s=_lg3(_pt3,_lt3,_rv3,float(np.clip(_Zv3,-5,5)))
                        _tmx3=1 if _sd3=="S1" else int(_ms3.ceil(_mr3))
                        _ecl3s=0.0; _cpd3=0.0
                        for _t3 in range(1,_tmx3+1):
                            _pm3=_pdm3(_pt3,_pp3s,_t3,5,3,_cpd3)
                            _pa3=_pd3(_pt3,_pp3s,_t3,5,3)
                            _cpd3=1-(1-_cpd3)*(1-_pa3)
                            _et3=_ea3(_e03,_me3,_mr3,_pp3s,_t3-1)
                            _dt3=1/(1+_tie3)**_t3
                            _ecl3s+=_pm3*_lp3s*_et3*_dt3
                        _ne3=next((c for c in["Description","NOM_EMETTEUR"] if c in _df_ecl_ai.columns),"")
                        _rs3.append({"CODE_ISIN":str(_r3.get("CODE_ISIN",""))[:15],
                            "Description":str(_r3.get(_ne3,""))[:45] if _ne3 else "",
                            "METHODE":_me3,"STADE":_sd3,"mat_res":round(_mr3,2),
                            "PD_TTC(%)":round(_pt3*100,4),
                            "PD_PIT_base(%)":round(_pp3b*100,4),
                            "PD_PIT_stress(%)":round(_pp3s*100,4),
                            "DeltaPD_pp":round((_pp3s-_pp3b)*100,4),
                            "LGD_base(%)":round(float(_r3["LGD_PIT"])*100,4),
                            "LGD_stress(%)":round(_lp3s*100,4),
                            "ECL_base":round(_ecl3b,2),"ECL_stress":round(_ecl3s,2),
                            "DeltaECL":round(_ecl3s-_ecl3b,2),
                            "Hausse(%)":round((_ecl3s/_ecl3b-1)*100,1) if _ecl3b>0.001 else 0})
                        if _ix3%50==0: _pg3.progress(int(_ix3/len(_df_ecl_ai)*95))
                    _pg3.progress(100)
                    st.session_state[_skey3]=pd.DataFrame(_rs3)

            if _skey3 in st.session_state:
                _ds3=st.session_state[_skey3]
                _eb3=_ds3["ECL_base"].sum(); _es3=_ds3["ECL_stress"].sum()
                _pd3b=_ds3["PD_PIT_base(%)"].mean(); _pd3s=_ds3["PD_PIT_stress(%)"].mean()
                _hausse_pct=(_es3/_eb3-1)*100 if _eb3>0 else 0

                _color_kpi="#A71E1E" if _hausse_pct>0 else "#1A6B2E"
                _k1,_k2,_k3,_k4,_k5=st.columns(5)
                _k1.metric("ECL base",f"{_eb3/1e3:.2f} K MAD")
                _k2.metric("ECL stresse",f"{_es3/1e3:.2f} K MAD",delta=f"{(_es3-_eb3)/1e3:+.2f}K")
                _k3.metric("Hausse ECL",f"{_hausse_pct:+.1f}%")
                _k4.metric("PD_PIT base",f"{_pd3b:.3f}%")
                _k5.metric("PD_PIT stress",f"{_pd3s:.3f}%",delta=f"{_pd3s-_pd3b:+.3f}pp")

                _dc1,_dc2=st.columns(2)
                with _dc1:
                    _n1b=(_ds3["STADE"]=="S1").sum();_n2b=(_ds3["STADE"]=="S2").sum();_n3b=(_ds3["STADE"]=="S3").sum()
                    _rp3=max(0,_pd3s/_pd3b-1) if _pd3b>0 else 0
                    _m12=max(0,int(_n1b*min(_rp3*0.5,0.30)));_m23=max(0,int(_n2b*min(_rp3*0.3,0.20)))
                    _fm3=go.Figure()
                    _fm3.add_trace(go.Bar(name="Base",x=["S1","S2","S3"],y=[_n1b,_n2b,_n3b],
                        marker_color=["#1A6B2E","#EF9F27","#A71E1E"],opacity=0.35,
                        text=[_n1b,_n2b,_n3b],textposition="outside"))
                    _fm3.add_trace(go.Bar(name="Stress",x=["S1","S2","S3"],
                        y=[max(0,_n1b-_m12),_n2b+_m12-_m23,_n3b+_m23],
                        marker_color=["#1A6B2E","#EF9F27","#A71E1E"],opacity=0.90,
                        text=[max(0,_n1b-_m12),_n2b+_m12-_m23,_n3b+_m23],textposition="outside"))
                    _fm3.update_layout(title="Migration des stades",barmode="group",height=300,
                        paper_bgcolor="white",plot_bgcolor="#F5F6FA",legend=dict(orientation="h",y=-0.3))
                    st.plotly_chart(_fm3,use_container_width=True,key="st3_migr")
                with _dc2:
                    _fh3=go.Figure(go.Histogram(x=_ds3["Hausse(%)"].tolist(),nbinsx=25,
                        marker_color="#E24B4A" if _delta_heur>0 else "#1A6B2E",opacity=0.78))
                    _fh3.add_vline(x=0,line_color="#1B1F6B",line_dash="dash",
                        annotation_text="0%",annotation_position="top right")
                    _fh3.update_layout(title="Distribution hausse ECL (%)",height=300,
                        paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                        xaxis_title="Hausse ECL (%)",yaxis_title="N obligations")
                    st.plotly_chart(_fh3,use_container_width=True,key="st3_hist")

                _ne3b=next((c for c in["Description","NOM_EMETTEUR"] if c in _ds3.columns),"")
                _fsc3=go.Figure()
                for _std3,_col3 in [("S1","#1A6B2E"),("S2","#EF9F27"),("S3","#A71E1E")]:
                    _gs3=_ds3[_ds3["STADE"]==_std3]
                    if not len(_gs3): continue
                    _fsc3.add_trace(go.Scatter(x=_gs3["DeltaPD_pp"].tolist(),
                        y=_gs3["DeltaECL"].tolist(),mode="markers",name=f"{_std3}({len(_gs3)})",
                        text=_gs3[_ne3b].tolist() if _ne3b else [],
                        hovertemplate="%{text}<br>ΔPD=%{x:+.3f}pp<br>ΔECL=%{y:+.2f} MAD",
                        marker=dict(size=6,color=_col3,opacity=0.75)))
                _fsc3.update_layout(title="ΔPD_PIT (pp) vs ΔECL (MAD) — par obligation",
                    height=340,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                    xaxis_title="ΔPD_PIT (pp)",yaxis_title="ΔECL (MAD)",
                    legend=dict(orientation="h",y=-0.3))
                st.plotly_chart(_fsc3,use_container_width=True,key="st3_scat")

                st.subheader("Top 15 — Obligations les plus impactees")
                st.dataframe(_ds3.sort_values("DeltaECL",ascending=(_delta_heur<0)).head(15).reset_index(drop=True),
                             use_container_width=True,hide_index=True,height=400)
                _buf3=io.BytesIO(); _ds3.to_excel(_buf3,index=False)
                st.download_button("Telecharger resultats (Excel)",data=_buf3.getvalue(),
                    file_name=f"Stress_{_preset_st.replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_st3")

                _niv_st="severe" if abs(_hausse_pct)>30 else ("modere" if abs(_hausse_pct)>10 else "limite")
                st.markdown(f"""
                <div style="background:{"#FEF2F2" if _hausse_pct>0 else "#F0FDF4"};
                border-left:4px solid {"#A71E1E" if _hausse_pct>0 else "#1A6B2E"};
                padding:14px 18px;border-radius:6px;margin-top:8px">
                <b>Scenario "{_preset_st}" — Impact {_niv_st} :</b><br>
                Δlogit = <b>{_delta_heur:+.4f}</b> → PD_PIT {_pd3b:.3f}% → <b>{_pd3s:.3f}%</b><br>
                ECL : <b>{_eb3/1e3:.2f}K MAD</b> → <b>{_es3/1e3:.2f}K MAD</b>
                (<b>{_hausse_pct:+.1f}%</b>)<br>
                Migration estimee : {_m12} S1→S2 | {_m23} S2→S3
                </div>""",unsafe_allow_html=True)
        else:
            st.warning("Calculez d'abord l'ECL (onglet ECL IFRS 9) pour activer le stress test.")


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 8 — COMPARAISON SCENARIOS (Pessimiste | Central | Optimiste)
# ════════════════════════════════════════════════════════════════════════════
with TABS[tab_comp]:
    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>Comparaison des Scenarios — Pessimiste | Central | Optimiste</h2>
      <p>Vision cote a cote des 3 scenarios sur PD · LGD · EAD · ECL · Stades</p>
    </div></div>""", unsafe_allow_html=True)

    # Recuperer ECL calcule
    _ecl_keys_cp=[k for k in st.session_state if str(k).startswith("ecl5_")]
    _df_cp=None
    if _ecl_keys_cp:
        _df_cp=st.session_state[sorted(_ecl_keys_cp)[-1]]

    if df is None:
        st.warning("Chargez les donnees dans la sidebar.")
    elif _df_cp is None:
        st.info("Calculez l'ECL dans l'onglet 'ECL IFRS 9' pour activer la comparaison.")
    else:
        _DELTA_CP={"Pessimiste":0.420,"Central":-0.252,"Optimiste":-0.850}
        _W_CP={"Pessimiste":0.20,"Central":0.70,"Optimiste":0.10}
        _COLS_SC={"Pessimiste":{"bg":"#FEF2F2","border":"#A71E1E","txt":"#A71E1E"},
                  "Central":  {"bg":"#EEF2FF","border":"#1B1F6B","txt":"#1B1F6B"},
                  "Optimiste":{"bg":"#F0FDF4","border":"#1A6B2E","txt":"#1A6B2E"}}

        # ── KPI par scenario ──────────────────────────────────────────────
        st.subheader("Indicateurs cles par scenario")
        _kcols=st.columns(3)
        for _sci,_sc in enumerate(["Pessimiste","Central","Optimiste"]):
            _c=_COLS_SC[_sc]
            _ecl_col=f"ECL_Life_{_sc[:4]}" if f"ECL_Life_{_sc[:4]}" in _df_cp.columns else "ECL_FINAL"
            _pd_col=f"PD_PIT_{_sc}" if f"PD_PIT_{_sc}" in _df_cp.columns else "PD_TTC"
            _ecl_sc=_df_cp[_ecl_col].sum() if _ecl_col in _df_cp.columns else 0
            _pd_sc=(_df_cp[_pd_col]*100).mean() if _pd_col in _df_cp.columns else _df_cp["PD_TTC"].mean()*100
            _lgd_sc=_df_cp["LGD_PIT"].mean()*100 if "LGD_PIT" in _df_cp.columns else 0
            with _kcols[_sci]:
                st.markdown(f"""
                <div style="background:{_c['bg']};border:2px solid {_c['border']};
                     border-radius:10px;padding:18px;margin-bottom:8px">
                <div style="font-size:0.95rem;font-weight:800;color:{_c['txt']};
                     text-transform:uppercase;letter-spacing:0.08em;margin-bottom:12px;
                     border-bottom:1px solid {_c['border']};padding-bottom:6px">
                  {_sc}
                  <span style="font-size:0.70rem;font-weight:500;opacity:0.75;margin-left:8px">
                  (poids {int(_W_CP[_sc]*100)}%)
                  </span>
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
                  <div style="text-align:center;padding:8px;background:white;border-radius:6px">
                    <div style="font-size:0.60rem;color:#666;text-transform:uppercase;font-weight:700">ECL Total</div>
                    <div style="font-size:1.25rem;font-weight:800;color:{_c['txt']}">{_ecl_sc/1e3:.1f}K</div>
                    <div style="font-size:0.60rem;color:#888">MAD</div>
                  </div>
                  <div style="text-align:center;padding:8px;background:white;border-radius:6px">
                    <div style="font-size:0.60rem;color:#666;text-transform:uppercase;font-weight:700">PD_PIT moy.</div>
                    <div style="font-size:1.25rem;font-weight:800;color:{_c['txt']}">{_pd_sc:.3f}%</div>
                  </div>
                  <div style="text-align:center;padding:8px;background:white;border-radius:6px">
                    <div style="font-size:0.60rem;color:#666;text-transform:uppercase;font-weight:700">LGD_PIT moy.</div>
                    <div style="font-size:1.25rem;font-weight:800;color:{_c['txt']}">{_lgd_sc:.2f}%</div>
                  </div>
                  <div style="text-align:center;padding:8px;background:white;border-radius:6px">
                    <div style="font-size:0.60rem;color:#666;text-transform:uppercase;font-weight:700">ECL/EAD</div>
                    <div style="font-size:1.25rem;font-weight:800;color:{_c['txt']}">{_ecl_sc/_df_cp['EAD_0'].sum()*100:.3f}%</div>
                  </div>
                </div>
                </div>""", unsafe_allow_html=True)

        st.markdown("---")

        # ── Graphiques comparatifs ────────────────────────────────────────
        _cg1,_cg2=st.columns(2)
        with _cg1:
            # ECL par stade et scenario
            _stades_cp=["S1","S2","S3"]
            _sc_names=["Pessimiste","Central","Optimiste"]
            _fig_comp_ecl=go.Figure()
            for _sc in _sc_names:
                _col_k=f"ECL_Life_{_sc[:4]}" if f"ECL_Life_{_sc[:4]}" in _df_cp.columns else "ECL_FINAL"
                _clr={"Pessimiste":"#A71E1E","Central":"#1B1F6B","Optimiste":"#1A6B2E"}[_sc]
                _vals=[_df_cp[_df_cp["STADE"]==s][_col_k].sum()/1e3 if _col_k in _df_cp.columns else 0
                       for s in _stades_cp]
                _fig_comp_ecl.add_trace(go.Bar(name=_sc,x=_stades_cp,y=_vals,
                    marker_color=_clr,opacity=0.85,
                    text=[f"{v:.1f}K" for v in _vals],textposition="outside"))
            _fig_comp_ecl.update_layout(title="ECL par Stade et Scenario (K MAD)",
                barmode="group",height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                yaxis_title="ECL (K MAD)",legend=dict(orientation="h",y=-0.25))
            st.plotly_chart(_fig_comp_ecl,use_container_width=True,key="cp_ecl_stade")

        with _cg2:
            # PD_PIT distribution par scenario
            _fig_pd_dist=go.Figure()
            for _sc in _sc_names:
                _pd_c=f"PD_PIT_{_sc}" if f"PD_PIT_{_sc}" in _df_cp.columns else "PD_TTC"
                _clr={"Pessimiste":"#A71E1E","Central":"#1B1F6B","Optimiste":"#1A6B2E"}[_sc]
                _fig_pd_dist.add_trace(go.Box(
                    y=(_df_cp[_pd_c]*100).tolist() if _pd_c in _df_cp.columns else (_df_cp["PD_TTC"]*100).tolist(),
                    name=_sc,marker_color=_clr,
                    boxmean=True,fillcolor="rgba(150,150,150,0.15)",line=dict(color=_clr,width=2)))
            _fig_pd_dist.update_layout(title="Distribution PD_PIT par Scenario (%)",
                height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                yaxis_title="PD_PIT (%)",showlegend=True,
                legend=dict(orientation="h",y=-0.25))
            st.plotly_chart(_fig_pd_dist,use_container_width=True,key="cp_pd_box")

        _cg3,_cg4=st.columns(2)
        with _cg3:
            # Waterfall : ECL pondere
            _ecl_values=[]
            for _sc in _sc_names:
                _col_k=f"ECL_Life_{_sc[:4]}" if f"ECL_Life_{_sc[:4]}" in _df_cp.columns else "ECL_FINAL"
                _ecl_values.append(_df_cp[_col_k].sum()/1e3 if _col_k in _df_cp.columns else 0)
            _ecl_pond=sum(_W_CP[s]*_ecl_values[i] for i,s in enumerate(_sc_names))
            _fig_wf_cp=go.Figure(go.Bar(
                x=_sc_names+["Pondere"],
                y=_ecl_values+[_ecl_pond],
                marker_color=["#A71E1E","#1B1F6B","#1A6B2E","#EF9F27"],
                opacity=0.88,text=[f"{v:.2f}K" for v in _ecl_values+[_ecl_pond]],
                textposition="outside"))
            _fig_wf_cp.add_hline(y=_ecl_pond,line_dash="dot",line_color="#EF9F27",
                annotation_text=f"ECL Pondere = {_ecl_pond:.2f}K MAD",
                annotation_position="top right")
            _fig_wf_cp.update_layout(title="ECL Total par Scenario + ECL Pondere (K MAD)",
                height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                yaxis_title="ECL (K MAD)",showlegend=False)
            st.plotly_chart(_fig_wf_cp,use_container_width=True,key="cp_ecl_wf")

        with _cg4:
            # Radar : profil risque par scenario
            _cats_r=["PD_PIT","LGD_PIT","ECL/EAD","S2+S3 %"]
            _fig_rad_cp=go.Figure()
            for _sci2,_sc in enumerate(_sc_names):
                _pd_c2=f"PD_PIT_{_sc}" if f"PD_PIT_{_sc}" in _df_cp.columns else "PD_TTC"
                _ecl_c2=f"ECL_Life_{_sc[:4]}" if f"ECL_Life_{_sc[:4]}" in _df_cp.columns else "ECL_FINAL"
                _clr2={"Pessimiste":"#A71E1E","Central":"#1B1F6B","Optimiste":"#1A6B2E"}[_sc]
                _pd_m=(_df_cp[_pd_c2]*100).mean() if _pd_c2 in _df_cp.columns else _df_cp["PD_TTC"].mean()*100
                _lgd_m=_df_cp["LGD_PIT"].mean()*100 if "LGD_PIT" in _df_cp.columns else 30
                _ecl_m=_df_cp[_ecl_c2].sum()/_df_cp["EAD_0"].sum()*100 if _ecl_c2 in _df_cp.columns else 0
                _s23_m=((_df_cp["STADE"]!="S1").sum()/len(_df_cp))*100
                # Normaliser 0-100 pour radar
                _maxv=[20,80,5,100]
                _vals_r=[min(_pd_m/_maxv[0]*100,100),min(_lgd_m/_maxv[1]*100,100),
                         min(_ecl_m/_maxv[2]*100,100),min(_s23_m/_maxv[3]*100,100)]
                _vals_r.append(_vals_r[0])
                _cats_r_c=_cats_r+[_cats_r[0]]
                _fig_rad_cp.add_trace(go.Scatterpolar(r=_vals_r,theta=_cats_r_c,
                    fill="toself",name=_sc,line=dict(color=_clr2,width=2),
                    fillcolor="rgba(150,150,150,0.13)"))
            _fig_rad_cp.update_layout(
                title="Profil de Risque par Scenario (normalisé 0-100)",
                polar=dict(radialaxis=dict(range=[0,100],tickfont=dict(size=8))),
                height=360,paper_bgcolor="white",
                legend=dict(orientation="h",y=-0.15))
            st.plotly_chart(_fig_rad_cp,use_container_width=True,key="cp_radar")

        # ── Tableau comparatif complet ────────────────────────────────────
        st.subheader("Synthese comparative — tous indicateurs")
        _rows_synth=[]
        for _sc in _sc_names:
            _ecl_c3=f"ECL_Life_{_sc[:4]}" if f"ECL_Life_{_sc[:4]}" in _df_cp.columns else "ECL_FINAL"
            _pd_c3=f"PD_PIT_{_sc}" if f"PD_PIT_{_sc}" in _df_cp.columns else "PD_TTC"
            _ecl_tot=_df_cp[_ecl_c3].sum() if _ecl_c3 in _df_cp.columns else 0
            _ead_tot=_df_cp["EAD_0"].sum()
            _rows_synth.append({"Scenario":_sc,
                "Poids":f"{int(_W_CP[_sc]*100)}%",
                "PD_PIT moy (%)":round((_df_cp[_pd_c3]*100).mean(),4) if _pd_c3 in _df_cp.columns else "—",
                "LGD_PIT moy (%)":round(_df_cp["LGD_PIT"].mean()*100,4) if "LGD_PIT" in _df_cp.columns else "—",
                "EAD Total (K MAD)":round(_ead_tot/1e3,1),
                "ECL Total (K MAD)":round(_ecl_tot/1e3,2),
                "ECL/EAD (%)":f"{_ecl_tot/_ead_tot*100:.4f}%" if _ead_tot>0 else "—",
                "S1 (N)":int((_df_cp["STADE"]=="S1").sum()),
                "S2 (N)":int((_df_cp["STADE"]=="S2").sum()),
                "S3 (N)":int((_df_cp["STADE"]=="S3").sum()),
            })
        _df_synth=pd.DataFrame(_rows_synth)
        def _hl_sc(row):
            sc=str(row.get("Scenario",""))
            bg={"Pessimiste":"#FFF5F5","Central":"#F5F5FF","Optimiste":"#F5FFF5"}.get(sc,"white")
            return [f"background-color:{bg}" for _ in row.index]
        st.dataframe(_df_synth.style.apply(_hl_sc,axis=1),
                     use_container_width=True,hide_index=True)


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 9 — RISQUE UL & VaR CREDIT
# ════════════════════════════════════════════════════════════════════════════
with TABS[tab_risq]:
    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>Risque de Credit — Perte Inattendue (UL) et VaR Credit</h2>
      <p>Capital economique · VaR 99.9% · Distribution des pertes · Bale III</p>
    </div></div>""", unsafe_allow_html=True)

    from scipy.stats import norm as _Nvar
    import math as _mvar

    _ecl_keys_var=[k for k in st.session_state if str(k).startswith("ecl5_")]
    _df_var=None
    if _ecl_keys_var:
        _df_var=st.session_state[sorted(_ecl_keys_var)[-1]]

    if df is None:
        st.warning("Chargez les donnees dans la sidebar.")
    elif _df_var is None:
        st.info("Calculez l'ECL dans l'onglet 'ECL IFRS 9' pour activer cette analyse.")
    else:
        _phiv=lambda x: _Nvar.cdf(np.clip(x,-30,30))
        _phiiv=lambda p: _Nvar.ppf(np.clip(p,1e-8,1-1e-8))
        def _rho_v(p): w=(1-np.exp(-50*p))/(1-np.exp(-50)); return 0.12*w+0.24*(1-w)

        # ── Parametres ────────────────────────────────────────────────────
        with st.expander("Parametres du modele de risque", expanded=False):
            _c_v1,_c_v2,_c_v3=st.columns(3)
            with _c_v1:
                _conf_level=st.slider("Niveau de confiance VaR (%)",90.0,99.9,99.9,0.1,key="var_conf")/100
                _n_sim_var=st.select_slider("Simulations Monte Carlo",
                    [1000,5000,10000,50000,100000],10000,key="var_nsim")
            with _c_v2:
                _rho_sys=st.slider("Correlation systemique ρ global (%)",1.0,30.0,12.0,0.5,key="var_rho")/100
                _horizon_var=st.selectbox("Horizon",["1 an","1 an (Lifetime S2/S3)"],key="var_hz")
            with _c_v3:
                st.markdown("**Formule Vasicek (Bale II) :**")
                st.latex(r"EL_i = PD_i \times LGD_i \times EAD_i")
                st.latex(r"K_i = \Phi\!\left(\frac{\Phi^{-1}(PD)+\sqrt{\rho}\,\Phi^{-1}(0.999)}{\sqrt{1-\rho}}\right) \times LGD \times EAD")

        # ── Calcul UL et VaR ─────────────────────────────────────────────
        _btn_var=st.button("Calculer UL et VaR Credit",type="primary",key="btn_var")
        _vkey=f"var_res_{len(df)}_{_conf_level}"

        if _btn_var:
            with st.spinner(f"Simulation Monte Carlo ({_n_sim_var:,} scenarios)..."):
                _pg_v=st.progress(0)
                _n=len(_df_var)

                # Parametres par obligation
                _pd_arr=_df_var["PD_TTC"].values.astype(float)
                _lgd_arr=_df_var["LGD_PIT"].values.astype(float) if "LGD_PIT" in _df_var.columns else _df_var["LGD_TTC"].values.astype(float)
                _ead_arr=_df_var["EAD_0"].values.astype(float)
                _rho_arr=np.array([_rho_v(float(p)) for p in _pd_arr])

                # ── 1. EL analytique (Expected Loss) ─────────────────────
                _el_arr=_pd_arr*_lgd_arr*_ead_arr
                _el_total=_el_arr.sum()

                # ── 2. Perte inattendue par obligation (Bale II) ──────────
                _ul_arr=np.sqrt(_pd_arr*(1-_pd_arr))*_lgd_arr*_ead_arr
                _ul_indiv=_ul_arr.sum()

                # ── 3. Capital economique Bale II (VaR analytique) ────────
                # K_i = [Phi((Phi^-1(PD)+sqrt(rho)*Phi^-1(0.999))/sqrt(1-rho))-PD] * LGD * EAD
                _q999=_Nvar.ppf(_conf_level)
                _k_arr=np.array([
                    (float(_phiv((_phiiv(float(_pd_arr[i]))+np.sqrt(_rho_arr[i])*_q999)/np.sqrt(1-_rho_arr[i])))-float(_pd_arr[i]))
                    *float(_lgd_arr[i])*float(_ead_arr[i])
                    for i in range(_n)])
                _k_total=_k_arr.sum()  # Capital economique total

                # ── 4. Monte Carlo : distribution des pertes ──────────────
                _pg_v.progress(20)
                np.random.seed(42)
                _Z_sys=np.random.standard_normal(_n_sim_var)  # facteur systemique
                _pertes_sim=np.zeros(_n_sim_var)
                _batch=max(1,_n_sim_var//20)
                for _bi in range(0,_n_sim_var,_batch):
                    _be=min(_bi+_batch,_n_sim_var)
                    _Z_b=_Z_sys[_bi:_be].reshape(-1,1)
                    # PD conditionnelle : Phi((Phi^-1(PD)-sqrt(rho)*Z)/sqrt(1-rho))
                    _pd_cond=_phiv(
                        (_phiiv(_pd_arr).reshape(1,-1)-np.sqrt(_rho_arr).reshape(1,-1)*_Z_b)
                        /np.sqrt(1-_rho_arr).reshape(1,-1))  # shape (batch, n_oblig)
                    _u=np.random.uniform(0,1,(_be-_bi,_n))
                    _defaults=(_u<_pd_cond).astype(float)
                    _pertes_sim[_bi:_be]=(_defaults*(_lgd_arr*_ead_arr).reshape(1,-1)).sum(axis=1)
                    _pg_v.progress(20+int((_bi/_n_sim_var)*70))

                _pg_v.progress(95)
                _var_mc=float(np.percentile(_pertes_sim,_conf_level*100))
                _cvar_mc=float(_pertes_sim[_pertes_sim>=_var_mc].mean()) if (_pertes_sim>=_var_mc).sum()>0 else _var_mc
                _ul_mc=float(_pertes_sim.std())
                _pg_v.progress(100)

                st.session_state[_vkey]={
                    "el_total":_el_total,"ul_indiv":_ul_indiv,
                    "k_total":_k_total,"var_mc":_var_mc,"cvar_mc":_cvar_mc,
                    "ul_mc":_ul_mc,"pertes_sim":_pertes_sim.tolist(),
                    "k_arr":_k_arr.tolist(),"el_arr":_el_arr.tolist(),
                    "ul_arr":_ul_arr.tolist(),
                }

        if _vkey in st.session_state:
            _vr=st.session_state[_vkey]
            _el=_vr["el_total"]; _ul=_vr["ul_mc"]
            _var=_vr["var_mc"]; _cvar=_vr["cvar_mc"]; _k=_vr["k_total"]
            _ead_tot_v=_df_var["EAD_0"].sum()

            # ── KPI ──────────────────────────────────────────────────────
            st.markdown("### Indicateurs de Risque de Credit")
            _kv=st.columns(5)
            _kv[0].metric("EL (Perte Attendue)",f"{_el/1e3:.2f} K MAD",
                delta=f"{_el/_ead_tot_v*100:.3f}% EAD")
            _kv[1].metric("UL (Perte Inattendue)",f"{_ul/1e3:.2f} K MAD",
                delta=f"σ portefeuille")
            _kv[2].metric(f"VaR Credit {_conf_level*100:.1f}%",f"{_var/1e3:.2f} K MAD",
                delta=f"x{_var/_el:.1f} EL")
            _kv[3].metric(f"CVaR (ES)",f"{_cvar/1e3:.2f} K MAD",
                delta=f"Expected Shortfall")
            _kv[4].metric("Capital Economique",f"{_k/1e3:.2f} K MAD",
                delta=f"{_k/_ead_tot_v*100:.3f}% EAD (Bale II)")

            st.markdown(f"""
            <div style="background:#EEF2FF;border-left:4px solid #1B1F6B;
                 padding:14px 18px;border-radius:8px;margin:12px 0;font-size:0.85rem">
            <b>Interpretation :</b><br>
            • <b>EL = {_el/1e3:.2f}K MAD</b> — Provision IFRS 9 requise (perte moyenne attendue)<br>
            • <b>UL = {_ul/1e3:.2f}K MAD</b> — Volatilite des pertes (perte inattendue, 1 ecart-type)<br>
            • <b>VaR {_conf_level*100:.1f}% = {_var/1e3:.2f}K MAD</b> — Dans {_conf_level*100:.1f}% des scenarios, la perte ne depasse pas ce niveau<br>
            • <b>CVaR = {_cvar/1e3:.2f}K MAD</b> — Perte moyenne dans les scenarios les plus adverses ({(1-_conf_level)*100:.1f}%)<br>
            • <b>Capital economique = {_k/1e3:.2f}K MAD</b> — Capital requis pour couvrir les pertes inattendues (Bale II)
            </div>""", unsafe_allow_html=True)

            # ── Distribution Monte Carlo ──────────────────────────────────
            _va1,_va2=st.columns(2)
            with _va1:
                _pertes=np.array(_vr["pertes_sim"])
                _fig_dist_v=go.Figure()
                _fig_dist_v.add_trace(go.Histogram(x=_pertes/1e3,nbinsx=80,
                    name="Distribution MC",marker_color="#1B1F6B",opacity=0.72,
                    histnorm="probability density"))
                _fig_dist_v.add_vline(x=_el/1e3,line_color="#EF9F27",line_width=2.5,line_dash="solid",
                    annotation_text=f"EL={_el/1e3:.1f}K",annotation_position="top right",
                    annotation_font=dict(color="#EF9F27",size=10))
                _fig_dist_v.add_vline(x=_var/1e3,line_color="#E24B4A",line_width=2.5,line_dash="dash",
                    annotation_text=f"VaR {_conf_level*100:.0f}%={_var/1e3:.1f}K",
                    annotation_position="top right",annotation_font=dict(color="#E24B4A",size=10))
                _fig_dist_v.add_vline(x=_cvar/1e3,line_color="#A71E1E",line_width=2,line_dash="dot",
                    annotation_text=f"CVaR={_cvar/1e3:.1f}K",annotation_position="top left",
                    annotation_font=dict(color="#A71E1E",size=10))
                _q_shade=float(np.percentile(_pertes,_conf_level*100))
                _pertes_sorted=np.sort(_pertes)
                _tail=_pertes_sorted[_pertes_sorted>=_q_shade]/1e3
                if len(_tail)>0:
                    _counts,_bins=np.histogram(_tail,bins=20)
                    _fig_dist_v.add_trace(go.Bar(x=(_bins[:-1]+_bins[1:])/2,
                        y=_counts/len(_pertes)/(_bins[1]-_bins[0]),
                        width=(_bins[1]-_bins[0]),
                        marker_color="#A71E1E",opacity=0.50,name="Queue (VaR+)"))
                _fig_dist_v.update_layout(
                    title=f"Distribution des Pertes — Monte Carlo ({_n_sim_var:,} simulations)",
                    height=420,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                    xaxis_title="Pertes (K MAD)",yaxis_title="Densite",
                    legend=dict(orientation="h",y=-0.25),barmode="overlay")
                st.plotly_chart(_fig_dist_v,use_container_width=True,key="var_dist")

            with _va2:
                # Capital economique par obligation (top 20)
                _k_arr_np=np.array(_vr["k_arr"])
                _ne_v=next((c for c in["Description","NOM_EMETTEUR"] if c in _df_var.columns),"")
                _top20_idx=np.argsort(_k_arr_np)[-20:][::-1]
                _top20_labs=[str(_df_var.iloc[i].get(_ne_v,""))[:30] if _ne_v else f"Oblig {i}"
                             for i in _top20_idx]
                _top20_k=_k_arr_np[_top20_idx]/1e3
                _fig_k=go.Figure(go.Bar(x=_top20_k,y=_top20_labs,orientation="h",
                    marker_color="#1B1F6B",opacity=0.85,
                    text=[f"{v:.2f}K" for v in _top20_k],textposition="outside"))
                _fig_k.update_layout(
                    title="Top 20 — Capital Economique par Obligation (K MAD)",
                    height=420,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                    xaxis_title="Capital economique (K MAD)",showlegend=False,
                    margin=dict(l=180,r=80))
                st.plotly_chart(_fig_k,use_container_width=True,key="var_k_top")

            # ── Decomposition du risque ────────────────────────────────────
            _va3,_va4=st.columns(2)
            with _va3:
                # EL vs UL vs Capital par secteur
                _sc_v=next((c for c in["Secteur","SECTEUR_GICS"] if c in _df_var.columns),None)
                if _sc_v:
                    _el_np=np.array(_vr["el_arr"]); _ul_np=np.array(_vr["ul_arr"])
                    _df_var2=_df_var.copy()
                    _df_var2["_k"]=_k_arr_np; _df_var2["_el"]=_el_np; _df_var2["_ul"]=_ul_np
                    _g_sec=_df_var2.groupby(_sc_v).agg(
                        EL=("_el","sum"),UL=("_ul","sum"),K=("_k","sum"),N=("EAD_0","count")).reset_index()
                    _g_sec=_g_sec.sort_values("K",ascending=True)
                    _fig_sec_k=go.Figure()
                    _fig_sec_k.add_trace(go.Bar(y=_g_sec[_sc_v],x=_g_sec["EL"]/1e3,
                        orientation="h",name="EL",marker_color="#EF9F27",opacity=0.85))
                    _fig_sec_k.add_trace(go.Bar(y=_g_sec[_sc_v],x=_g_sec["K"]/1e3,
                        orientation="h",name="Capital Eco.",marker_color="#1B1F6B",opacity=0.85))
                    _fig_sec_k.update_layout(title="EL vs Capital Economique par Secteur",
                        barmode="group",height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                        xaxis_title="K MAD",legend=dict(orientation="h",y=-0.25),
                        margin=dict(l=140))
                    st.plotly_chart(_fig_sec_k,use_container_width=True,key="var_sec_k")

            with _va4:
                # Courbe VaR (confidence levels)
                _conf_levels=np.arange(0.90,0.9995,0.002)
                _var_curve=[float(np.percentile(_pertes,c*100)) for c in _conf_levels]
                _fig_var_c=go.Figure()
                _fig_var_c.add_trace(go.Scatter(x=_conf_levels*100,y=np.array(_var_curve)/1e3,
                    mode="lines",name="VaR(α)",line=dict(color="#1B1F6B",width=2.5)))
                _fig_var_c.add_scatter(x=[_conf_level*100],y=[_var/1e3],mode="markers",
                    name=f"VaR {_conf_level*100:.1f}%",
                    marker=dict(size=12,color="#E24B4A",symbol="diamond"))
                _fig_var_c.update_layout(title="Courbe VaR Credit selon le niveau de confiance",
                    height=360,paper_bgcolor="white",plot_bgcolor="#F5F6FA",
                    xaxis_title="Niveau de confiance (%)",yaxis_title="VaR (K MAD)",
                    legend=dict(orientation="h",y=-0.25))
                st.plotly_chart(_fig_var_c,use_container_width=True,key="var_curve")

            # ── Tableau de risque par obligation ─────────────────────────
            with st.expander("Tableau risque par obligation (K, EL, UL)"):
                _ne_v2=next((c for c in["Description","NOM_EMETTEUR"] if c in _df_var.columns),"")
                _df_risk_tab=pd.DataFrame({
                    "CODE_ISIN":_df_var.get("CODE_ISIN",pd.Series(["—"]*len(_df_var))),
                    "Description":_df_var[_ne_v2].str[:40] if _ne_v2 in _df_var.columns else "—",
                    "STADE":_df_var["STADE"],
                    "PD_TTC(%)":(_df_var["PD_TTC"]*100).round(4),
                    "LGD_PIT(%)":(_df_var["LGD_PIT"]*100).round(4) if "LGD_PIT" in _df_var.columns else "—",
                    "EAD_0":_df_var["EAD_0"].round(0),
                    "EL (MAD)":np.round(np.array(_vr["el_arr"]),2),
                    "UL (MAD)":np.round(np.array(_vr["ul_arr"]),2),
                    "Capital Eco (MAD)":np.round(_k_arr_np,2),
                    "K/EL ratio":np.round(_k_arr_np/np.array(_vr["el_arr"]),2)
                }).sort_values("Capital Eco (MAD)",ascending=False).reset_index(drop=True)
                st.dataframe(_df_risk_tab,use_container_width=True,hide_index=True,height=460)
                _buf_var=io.BytesIO()
                _df_risk_tab.to_excel(_buf_var,index=False)
                st.download_button("Telecharger analyse risque (Excel)",
                    data=_buf_var.getvalue(),file_name="Risque_UL_VaR_IFRS9.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_var")


# ════════════════════════════════════════════════════════════════════════════
# ONGLET 10 — RAPPORT PDF AUTOMATIQUE
# ════════════════════════════════════════════════════════════════════════════
with TABS[tab_rapp]:
    st.markdown("""<div class='hdr'><div class='hdr-txt'>
      <h2>Rapport PDF Automatique — IFRS 9 Forvis Mazars</h2>
      <p>Generation d'un rapport professionnel complet avec tous les resultats</p>
    </div></div>""", unsafe_allow_html=True)

    _ecl_keys_rp=[k for k in st.session_state if str(k).startswith("ecl5_")]
    _df_rp=st.session_state[sorted(_ecl_keys_rp)[-1]] if _ecl_keys_rp else None
    _var_rp=next((st.session_state[k] for k in st.session_state if str(k).startswith("var_res_")),None)

    # Options du rapport
    st.subheader("Configuration du rapport")
    _r1,_r2,_r3=st.columns(3)
    with _r1:
        _rp_titre=st.text_input("Titre du rapport","Rapport IFRS 9 — Portefeuille Obligataire Marocain",key="rp_titre")
        _rp_date=st.text_input("Date","25 Avril 2025",key="rp_date")
    with _r2:
        _rp_auteur=st.text_input("Auteur","PFE — Forvis Mazars Maroc",key="rp_auteur")
        _rp_version=st.text_input("Version","v1.0",key="rp_vers")
    with _r3:
        _rp_sections=st.multiselect("Sections a inclure",
            ["Resume Executif","Portefeuille & Staging","PD Vasicek & CPV",
             "LGD Frye-Jacobs","EAD","ECL IFRS 9","Analyse de Risque (UL/VaR)",
             "Comparaison Scenarios"],
            default=["Resume Executif","Portefeuille & Staging","ECL IFRS 9",
                     "Analyse de Risque (UL/VaR)"],key="rp_secs")

    if st.button("Generer le Rapport Excel Complet",type="primary",key="btn_rpp"):
        if df is None:
            st.warning("Chargez les donnees d'abord.")
        else:
            with st.spinner("Generation du rapport..."):
                try:
                    from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
                    from openpyxl.utils import get_column_letter
                    from openpyxl import Workbook as _WB
                    import io as _ior

                    _wb=_WB()
                    _nm=_wb.active; _nm.title="Page de Garde"

                    # ── Styles ────────────────────────────────────────────
                    def _fill(hex_c): return PatternFill("solid",fgColor=hex_c.replace("#",""))
                    def _font(bold=False,sz=11,color="000000",italic=False):
                        return Font(bold=bold,size=sz,color=color.replace("#",""),italic=italic)
                    def _align(h="left",v="center",wrap=False):
                        return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
                    def _border():
                        _s=Side(style="thin",color="D8DFF0")
                        return Border(left=_s,right=_s,top=_s,bottom=_s)

                    _NAVY="1B1F6B"; _CYAN="009FE3"; _LGRAY="F0F4FA"; _WHITE="FFFFFF"
                    _GREEN="1A6B2E"; _RED="A31919"; _GOLD="8B6914"

                    # ── PAGE DE GARDE PROFESSIONNELLE ─────────────────────
                    # Fond navy pour toute la page de garde
                    from openpyxl.drawing.image import Image as _XLImg
                    from openpyxl.utils import get_column_letter as _gcl2
                    import io as _img_io, base64 as _b64

                    # Largeurs colonnes
                    for _ci_g in range(1,11):
                        _nm.column_dimensions[_gcl2(_ci_g)].width=14

                    # Ligne 1: Bande navy haute (logo zone)
                    _nm.merge_cells("A1:J1")
                    _nm["A1"]=""
                    _nm["A1"].fill=_fill(_NAVY)
                    _nm.row_dimensions[1].height=12

                    # Ligne 2-4: Zone logo + nom Forvis Mazars (simulé en texte stylisé)
                    _nm.merge_cells("A2:D4")
                    _nm.merge_cells("A2:D3")
                    _nm["A2"]="forv/s mazars"
                    _nm["A2"].font=Font(bold=True,size=28,color=_CYAN,name="Arial")

                    _nm["A2"].fill=_fill(_NAVY)
                    _nm["A2"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
                    _nm.row_dimensions[2].height=30
                    _nm.row_dimensions[3].height=30
                    _nm.row_dimensions[4].height=30

                    # Ligne 2-4: Titre rapport
                    _nm.merge_cells("E2:J4")
                    _nm["E2"]=_rp_titre
                    _nm["E2"].font=Font(bold=True,size=16,color=_WHITE,name="Calibri")
                    _nm["E2"].fill=_fill(_NAVY)
                    _nm["E2"].alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)

                    # Ligne 5: Bande cyan separatrice
                    _nm.merge_cells("A5:J5")
                    _nm["A5"]=""
                    _nm["A5"].fill=_fill(_CYAN)
                    _nm.row_dimensions[5].height=6

                    # Ligne 6: Sous-titre
                    _nm.merge_cells("A6:J6")
                    _nm["A6"]="Implementation de la Norme IFRS 9 — Portefeuille Obligataire Marocain — 328 Obligations"
                    _nm["A6"].font=Font(bold=False,size=12,color=_NAVY,name="Calibri")
                    _nm["A6"].fill=_fill("F0F4FA")
                    _nm["A6"].alignment=Alignment(horizontal="center",vertical="center")
                    _nm.row_dimensions[6].height=28

                    # Ligne 7: Methodes
                    _nm.merge_cells("A7:J7")
                    _nm["A7"]="Staging Dynamique  |  PD Merton-Vasicek  |  PD CPV  |  LGD Vasicek & Frye-Jacobs  |  EAD  |  ECL IFRS 9  |  Analyse & Interpretation"
                    _nm["A7"].font=Font(bold=False,size=9,color="666666",name="Calibri",italic=True)
                    _nm["A7"].fill=_fill("F0F4FA")
                    _nm["A7"].alignment=Alignment(horizontal="center",vertical="center")
                    _nm.row_dimensions[7].height=22

                    # Ligne 8: Espace
                    _nm.row_dimensions[8].height=20

                    # Ligne 9: En-tete section infos
                    _nm.merge_cells("B9:I9")
                    _nm["B9"]="INFORMATIONS DU RAPPORT"
                    _nm["B9"].font=Font(bold=True,size=10,color=_WHITE,name="Calibri")
                    _nm["B9"].fill=_fill(_NAVY)
                    _nm["B9"].alignment=Alignment(horizontal="center",vertical="center")
                    _nm.row_dimensions[9].height=22

                    _infos=[
                        ("Auteur",_rp_auteur),("Date",_rp_date),
                        ("Version",_rp_version),("",""),
                        ("Obligations analysees",f"{len(df)}"),
                        ("Emetteurs",f"{df['NOM_EMETTEUR'].nunique() if 'NOM_EMETTEUR' in df.columns else '—'}"),
                    ]
                    if _df_rp is not None:
                        _infos+=[
                            ("ECL Total (MAD)",f"{_df_rp['ECL_FINAL'].sum():,.0f}"),
                            ("ECL / EAD",f"{_df_rp['ECL_FINAL'].sum()/_df_rp['EAD_0'].sum()*100:.4f}%"),
                            ("S1 (Sain)",f"{(_df_rp['STADE']=='S1').sum()} obligations"),
                            ("S2 (Surveillance)",f"{(_df_rp['STADE']=='S2').sum()} obligations"),
                            ("S3 (Defaut probable)",f"{(_df_rp['STADE']=='S3').sum()} obligations"),
                        ]
                    for _ri,(_lbl,_val) in enumerate(_infos,10):
                        _nm.merge_cells(f"B{_ri}:D{_ri}")
                        _nm.merge_cells(f"E{_ri}:J{_ri}")
                        _nm[f"B{_ri}"]=_lbl
                        _nm[f"B{_ri}"].font=_font(True,11,_NAVY)
                        _nm[f"B{_ri}"].fill=_fill(_LGRAY)
                        _nm[f"E{_ri}"]=_val
                        _nm[f"E{_ri}"].font=_font(False,11)
                        _nm.row_dimensions[_ri].height=22

                    _nm.column_dimensions["A"].width=3
                    for _col in ["B","C","D"]: _nm.column_dimensions[_col].width=20
                    for _col in ["E","F","G","H","I"]: _nm.column_dimensions[_col].width=18

                    # ── ONGLET PORTEFEUILLE ────────────────────────────────
                    if "Portefeuille & Staging" in _rp_sections and "STADE" in df.columns:
                        _ws_p=_wb.create_sheet("Portefeuille & Staging")
                        _hdr_cols=["CODE_ISIN","Description complète","Secteur","METHODE_VALO",
                                   "Maturité (ans)","NOMINAL","PD","LGD","STADE"]
                        _hdr_cols_ok=[c for c in _hdr_cols if c in df.columns]
                        for _ci,_cn in enumerate(_hdr_cols_ok,1):
                            _ws_p.cell(1,_ci,_cn).fill=_fill(_NAVY)
                            _ws_p.cell(1,_ci).font=_font(True,10,_WHITE)
                            _ws_p.cell(1,_ci).alignment=_align("center")
                            _ws_p.column_dimensions[get_column_letter(_ci)].width=18
                        _ws_p.row_dimensions[1].height=22
                        for _ri2,_row2 in df[_hdr_cols_ok].iterrows():
                            _bg2={"1":_LGRAY,"2":"FFF3CD","3":"F8D7DA"}.get(str(_row2.get("STADE",1)),_WHITE)
                            for _ci2,_v2 in enumerate(_row2,1):
                                _cl2=_ws_p.cell(_ri2+2,_ci2,str(_v2) if not isinstance(_v2,(int,float)) else _v2)
                                _cl2.fill=_fill(_bg2)
                                _cl2.border=_border()
                                _cl2.font=_font(sz=9)

                    # ── ONGLET ECL ─────────────────────────────────────────
                    if "ECL IFRS 9" in _rp_sections and _df_rp is not None:
                        _ws_ecl=_wb.create_sheet("ECL IFRS 9")
                        _ecl_cols=[
                            "CODE_ISIN","Description","Secteur","METHODE","STADE","mat_res",
                            "EAD_0","PD_TTC","LGD_TTC","LGD_PIT",
                            "ECL_12M_Pess","ECL_12M_Cent","ECL_12M_Opti","ECL_12M_Pond",
                            "ECL_Life_Pess","ECL_Life_Cent","ECL_Life_Opti","ECL_Life_Pond",
                            "ECL_FINAL"]
                        _ecl_cols_ok=[c for c in _ecl_cols if c in _df_rp.columns]
                        # Ajouter PD_PIT Vasicek (neutre) depuis df global si disponible
                        from scipy.stats import norm as _Nrp
                        if "PD_PIT_NEUTRE" in df.columns and "PD_PIT_NEUTRE" not in _df_rp.columns:
                            # Aligner par CODE_ISIN
                            _pit_map=df.set_index("CODE_ISIN")["PD_PIT_NEUTRE"].to_dict() if "CODE_ISIN" in df.columns else {}
                            _df_rp["PD_PIT_Vasicek"]=_df_rp["CODE_ISIN"].map(_pit_map).fillna(_df_rp["PD_TTC"]) if "CODE_ISIN" in _df_rp.columns else _df_rp["PD_TTC"]
                        elif "PD_PIT_NEUTRE" not in _df_rp.columns:
                            # Fallback: recalcul depuis delta CPV Central
                            _DELTA_RP={"Pessimiste":0.420,"Central":-0.252,"Optimiste":-0.850}
                            for _sc_rp,_d_rp in _DELTA_RP.items():
                                _col_pp=f"PD_PIT_{_sc_rp}"
                                if _col_pp not in _df_rp.columns:
                                    _df_rp[_col_pp]=_df_rp["PD_TTC"].apply(
                                        lambda p: float(1/(1+np.exp(-(_Nrp.ppf(np.clip(float(p),1e-8,1-1e-8))+_d_rp)))))
                            _df_rp["PD_PIT_Vasicek"]=_df_rp.get("PD_PIT_Central",_df_rp["PD_TTC"])
                        # Inserer PD_PIT apres PD_TTC dans la liste
                        _ecl_cols_full=[
                            "CODE_ISIN","Description","Secteur","METHODE","STADE","mat_res",
                            "EAD_0",
                            "PD_TTC","PD_PIT_Vasicek",
                            "ECL_12M_Pess","ECL_12M_Cent","ECL_12M_Opti","ECL_12M_Pond",
                            "LGD_TTC","LGD_PIT",
                            "ECL_Life_Pess","ECL_Life_Cent","ECL_Life_Opti","ECL_Life_Pond",
                            "ECL_FINAL"]
                        _ecl_cols_ok=[c for c in _ecl_cols_full if c in _df_rp.columns]
                        # Renommer pour l'affichage
                        _col_rename_ecl={
                            "PD_TTC":"PD_TTC (%)",
                            "PD_PIT_Vasicek":"PD_PIT Vasicek (%)",
                            "PD_PIT_Pessimiste":"PD_PIT Pess. (%)",
                            "PD_PIT_Central":"PD_PIT Cent. (%)",
                            "PD_PIT_Optimiste":"PD_PIT Opti. (%)",
                            "LGD_TTC":"LGD_TTC (%)",
                            "LGD_PIT":"LGD_PIT FJ (%)",
                            "ECL_12M_Pess":"ECL_12M Pess. (MAD)",
                            "ECL_12M_Cent":"ECL_12M Cent. (MAD)",
                            "ECL_12M_Opti":"ECL_12M Opti. (MAD)",
                            "ECL_12M_Pond":"ECL_12M Pondere (MAD)",
                            "ECL_Life_Pess":"ECL_Life Pess. (MAD)",
                            "ECL_Life_Cent":"ECL_Life Cent. (MAD)",
                            "ECL_Life_Opti":"ECL_Life Opti. (MAD)",
                            "ECL_Life_Pond":"ECL_Life Pondere (MAD)",
                            "ECL_FINAL":"ECL FINAL (MAD)",
                        }
                        for _ci3,_cn3 in enumerate(_ecl_cols_ok,1):
                            _ws_ecl.cell(1,_ci3,_col_rename_ecl.get(_cn3,_cn3)).fill=_fill(_NAVY)
                            _ws_ecl.cell(1,_ci3).font=_font(True,9,_WHITE)
                            _ws_ecl.column_dimensions[get_column_letter(_ci3)].width=18
                        _ws_ecl.row_dimensions[1].height=24
                        _ws_ecl.freeze_panes="A2"
                        # Renommer headers pour lisibilite
                        _col_rename_ecl={"PD_TTC":"PD_TTC (%)","PD_PIT_Pessimiste":"PD_PIT_Pess (%)","PD_PIT_Central":"PD_PIT_Cent (%)","PD_PIT_Optimiste":"PD_PIT_Opti (%)","LGD_TTC":"LGD_TTC (%)","LGD_PIT":"LGD_PIT_FJ (%)","ECL_12M_Pond":"ECL_12M_Pond (MAD)","ECL_Life_Pond":"ECL_Life_Pond (MAD)","ECL_FINAL":"ECL_FINAL (MAD)"}
                        for _ci3r,_cn3r in enumerate(_ecl_cols_ok,1):
                            _ws_ecl.cell(1,_ci3r,_col_rename_ecl.get(_cn3r,_cn3r))
                        _pct_cols_ecl={"PD_TTC","PD_PIT_Vasicek","PD_PIT_Pessimiste","PD_PIT_Central",
                            "PD_PIT_Optimiste","LGD_TTC","LGD_PIT"}
                        for _ri3,_row3 in _df_rp[_ecl_cols_ok].iterrows():
                            _bg3={"S1":_LGRAY,"S2":"FFF3CD","S3":"F8D7DA"}.get(str(_row3.get("STADE","S1")),_WHITE)
                            for _ci3b,(_cn3b,_v3) in enumerate(zip(_ecl_cols_ok,_row3),1):
                                # Convertir PD/LGD en % pour lisibilite
                                if _cn3b in _pct_cols_ecl and isinstance(_v3,float):
                                    _v3_out=round(float(_v3)*100,4)
                                elif isinstance(_v3,float):
                                    _v3_out=round(float(_v3),2)
                                else:
                                    _v3_out=str(_v3)
                                _cl3=_ws_ecl.cell(_ri3+2,_ci3b,_v3_out)
                                _cl3.fill=_fill(_bg3); _cl3.border=_border(); _cl3.font=_font(sz=9)

                        # Resume ECL
                        _ws_ecl2=_wb.create_sheet("Resume ECL par Stade")
                        _stades_rp=["S1","S2","S3","Total"]
                        _rec_rp=[["Indicateur"]+_stades_rp]
                        _rec_rp.append(["N obligations"]+[str(int((_df_rp["STADE"]==s).sum())) if s!="Total" else str(len(_df_rp)) for s in _stades_rp[:-1]]+[str(len(_df_rp))])
                        _rec_rp.append(["EAD_0 (MAD)"]+[f"{_df_rp[_df_rp['STADE']==s]['EAD_0'].sum():,.0f}" if s!="Total" else f"{_df_rp['EAD_0'].sum():,.0f}" for s in _stades_rp[:-1]]+[f"{_df_rp['EAD_0'].sum():,.0f}"])
                        _rec_rp.append(["ECL_FINAL (MAD)"]+[f"{_df_rp[_df_rp['STADE']==s]['ECL_FINAL'].sum():,.0f}" if s!="Total" else f"{_df_rp['ECL_FINAL'].sum():,.0f}" for s in _stades_rp[:-1]]+[f"{_df_rp['ECL_FINAL'].sum():,.0f}"])
                        _rec_rp.append(["ECL/EAD (%)"]+[f"{_df_rp[_df_rp['STADE']==s]['ECL_FINAL'].sum()/_df_rp[_df_rp['STADE']==s]['EAD_0'].sum()*100:.4f}%" if s!="Total" and (_df_rp["STADE"]==s).sum()>0 else f"{_df_rp['ECL_FINAL'].sum()/_df_rp['EAD_0'].sum()*100:.4f}%" for s in _stades_rp])
                        for _ri4,_row4 in enumerate(_rec_rp,1):
                            for _ci4,_v4 in enumerate(_row4,1):
                                _cl4=_ws_ecl2.cell(_ri4,_ci4,_v4)
                                _cl4.fill=_fill(_NAVY if _ri4==1 else (_LGRAY if _ri4%2==0 else _WHITE))
                                _cl4.font=_font(True if _ri4==1 else False,10,_WHITE if _ri4==1 else "000000")
                                _cl4.border=_border(); _cl4.alignment=_align("center")
                                _ws_ecl2.column_dimensions[get_column_letter(_ci4)].width=20

                    # ── ONGLET UL/VaR ──────────────────────────────────────
                    if "Analyse de Risque (UL/VaR)" in _rp_sections and _var_rp:
                        _ws_var=_wb.create_sheet("Risque UL & VaR")
                        _risk_info=[
                            ("ANALYSE DE RISQUE DE CREDIT — UL & VaR","","",""),
                            ("","","",""),
                            ("Indicateur","Valeur (MAD)","% EAD","Commentaire"),
                            ("EL (Perte Attendue)",f"{_var_rp['el_total']:,.0f}",
                             f"{_var_rp['el_total']/_df_rp['EAD_0'].sum()*100:.3f}%" if _df_rp is not None else "—",
                             "Provision IFRS 9 requise"),
                            ("UL (Perte Inattendue σ)",f"{_var_rp['ul_mc']:,.0f}",
                             f"{_var_rp['ul_mc']/_df_rp['EAD_0'].sum()*100:.3f}%" if _df_rp is not None else "—",
                             "Ecart-type Monte Carlo"),
                            (f"VaR 99.9%",f"{_var_rp['var_mc']:,.0f}",
                             f"{_var_rp['var_mc']/_df_rp['EAD_0'].sum()*100:.3f}%" if _df_rp is not None else "—",
                             "Quantile 99.9% des pertes"),
                            ("CVaR (Expected Shortfall)",f"{_var_rp['cvar_mc']:,.0f}",
                             f"{_var_rp['cvar_mc']/_df_rp['EAD_0'].sum()*100:.3f}%" if _df_rp is not None else "—",
                             "Perte moyenne au-dela VaR"),
                            ("Capital Economique (Bale II)",f"{_var_rp['k_total']:,.0f}",
                             f"{_var_rp['k_total']/_df_rp['EAD_0'].sum()*100:.3f}%" if _df_rp is not None else "—",
                             "Pilier 1 — formule Vasicek"),
                        ]
                        for _ri5,_row5 in enumerate(_risk_info,1):
                            for _ci5,_v5 in enumerate(_row5,1):
                                _cl5=_ws_var.cell(_ri5,_ci5,_v5)
                                if _ri5==1: _cl5.fill=_fill(_NAVY); _cl5.font=_font(True,14,_WHITE)
                                elif _ri5==3: _cl5.fill=_fill(_CYAN); _cl5.font=_font(True,11,_WHITE)
                                else: _cl5.fill=_fill(_LGRAY if _ri5%2==0 else _WHITE); _cl5.font=_font(sz=10)
                                _cl5.border=_border(); _cl5.alignment=_align("center","center")
                            _ws_var.row_dimensions[_ri5].height=25
                        for _ci5b in range(1,5):
                            _ws_var.column_dimensions[get_column_letter(_ci5b)].width=32

                    # Sauvegarde
                    _buf_rp=io.BytesIO(); _wb.save(_buf_rp)
                    st.session_state["rapport_xlsx"]=_buf_rp.getvalue()
                    st.success("Rapport genere avec succes!")
                except Exception as _erp:
                    import traceback; st.error(f"Erreur: {_erp}"); st.code(traceback.format_exc())

    if "rapport_xlsx" in st.session_state:
        _kcols_rp=st.columns([1,3])
        with _kcols_rp[0]:
            st.download_button(
                "Telecharger le Rapport Excel",
                data=st.session_state["rapport_xlsx"],
                file_name=f"Rapport_IFRS9_Forvis_Mazars_{_rp_date.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_rapport", use_container_width=True)
        with _kcols_rp[1]:
            st.info("Le rapport contient : Page de garde · Portefeuille & Staging · ECL IFRS 9 · Resume par Stade · Analyse Risque UL/VaR")

    # Apercu du rapport
    st.markdown("---")
    st.subheader("Apercu du contenu du rapport")
    _preview_sections={
        "Page de Garde": ["Titre, auteur, date, version","Indicateurs cles (ECL total, N obligations, stades)","En-tete Forvis Mazars"],
        "Portefeuille & Staging": ["328 obligations avec code ISIN, secteur, methode","Maturite, nominal, PD, LGD","STADE IFRS 9 (colore S1/S2/S3)"],
        "ECL IFRS 9": ["ECL par obligation (12M et Lifetime)","PD_PIT, LGD_PIT, EAD_0","Scenarios Pessimiste, Central, Optimiste"],
        "Resume ECL par Stade": ["EAD, ECL, ECL/EAD par stade","N obligations S1/S2/S3","Total portefeuille"],
        "Risque UL & VaR": ["EL, UL, VaR 99.9%, CVaR","Capital Economique Bale II","% de l'EAD total"],
    }
    _pv_cols=st.columns(len(_preview_sections))
    for _pvi,(_pv_title,_pv_items) in enumerate(_preview_sections.items()):
        with _pv_cols[_pvi]:
            st.markdown(f"""
            <div style="background:white;border:1px solid #D8DFF0;border-top:3px solid #1B1F6B;
                 border-radius:8px;padding:14px;height:180px">
            <div style="font-weight:700;font-size:0.78rem;color:#1B1F6B;
                 text-transform:uppercase;letter-spacing:0.05em;margin-bottom:10px">{_pv_title}</div>
            {"".join(f'<div style="font-size:0.72rem;color:#555;padding:3px 0;border-bottom:1px solid #EEF0F8">{item}</div>' for item in _pv_items)}
            </div>""", unsafe_allow_html=True)
